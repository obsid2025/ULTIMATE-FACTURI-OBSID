"""
Ultimate Facturi OBSID - Dashboard Web
Aplicatie Streamlit pentru procesarea si gruparea facturilor
"""

import streamlit as st
import pandas as pd
import os
import tempfile
import shutil
from datetime import datetime
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Import utils
from utils.auth import login_form, logout, is_authenticated, get_user_name, check_auth_for_action
from utils.mt940_parser import extrage_referinte_op_din_mt940_folder, get_sursa_incasare
# NOTE: proceseaza_borderouri_gls, proceseaza_borderouri_sameday removed - now using API sync
from utils.processors import proceseaza_netopia
from utils.export import genereaza_export_excel
from utils.data_sync import (
    import_mt940_to_supabase,
    sync_oblio_invoices,
    get_profit_data,
    get_dashboard_stats,
    get_recent_sync_logs,
    get_transactions_for_period,
    get_invoices_for_period
)
from utils.matching_engine import (
    generate_opuri_report_data,
    export_opuri_to_excel,
    get_matching_statistics,
    search_invoice_by_amount_in_oblio
)
from utils.supabase_client import test_connection as test_supabase
from utils.oblio_api import test_connection as test_oblio
import plotly.express as px
import plotly.graph_objects as go

# Page config
st.set_page_config(
    page_title="Ultimate Facturi OBSID",
    page_icon="https://gomagcdn.ro/domains3/obsid.ro/files/company/parfumuri-arabesti8220.svg",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Premium CSS - GitHub Dark Aesthetic
st.markdown("""
<style>
    /* Import Inter font - clean, modern, highly readable */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

    /* CSS Variables - GitHub Dark Theme */
    :root {
        --bg-primary: #0d1117;
        --bg-secondary: #161b22;
        --bg-tertiary: #21262d;
        --bg-card: #1c2128;
        --border-subtle: #30363d;
        --border-accent: #484f58;
        --text-primary: #e6edf3;
        --text-secondary: #8b949e;
        --text-muted: #6e7681;
        --accent-primary: #8b949e;
        --accent-light: #c9d1d9;
        --accent-dark: #6e7681;
        --accent-emerald: #3fb950;
        --accent-rose: #f85149;
        --accent-blue: #58a6ff;
        --shadow-primary: rgba(139, 148, 158, 0.1);
        --shadow-dark: rgba(0, 0, 0, 0.5);
    }

    /* Global resets */
    .main {
        background-color: var(--bg-primary);
    }

    .stApp {
        background-color: var(--bg-primary);
    }

    /* Sidebar styling */
    [data-testid="stSidebar"] {
        background: var(--bg-secondary);
        border-right: 1px solid var(--border-subtle);
    }

    [data-testid="stSidebar"] > div:first-child {
        padding-top: 0;
    }

    /* Typography - Inter for clean readability */
    h1, h2, h3, h4, h5, h6 {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif !important;
        color: var(--text-primary) !important;
        letter-spacing: -0.01em;
    }

    h1 {
        font-size: 1.75rem !important;
        font-weight: 600 !important;
        color: var(--text-primary) !important;
        margin-bottom: 0.5rem !important;
    }

    h2 {
        font-size: 1.25rem !important;
        font-weight: 600 !important;
        color: var(--text-primary) !important;
    }

    h3 {
        font-size: 0.875rem !important;
        font-weight: 500 !important;
        color: var(--text-secondary) !important;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }

    p, span, div, label {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
        color: var(--text-secondary);
    }

    /* Brand header in sidebar */
    .brand-header {
        padding: 1.5rem 1rem;
        border-bottom: 1px solid var(--border-subtle);
        margin-bottom: 1rem;
    }

    .brand-logo {
        display: flex;
        align-items: center;
        gap: 1rem;
    }

    .brand-logo img {
        width: 48px;
        height: 48px;
        filter: grayscale(100%) brightness(1.2);
    }

    .brand-text {
        display: flex;
        flex-direction: column;
    }

    .brand-name {
        font-family: 'VCR OSD Mono', monospace;
        font-size: 1.125rem;
        font-weight: 400;
        color: var(--text-primary);
        letter-spacing: 0.05em;
        line-height: 1.2;
    }

    .brand-tagline {
        font-family: 'VCR OSD Mono', monospace;
        font-size: 0.625rem;
        color: var(--text-secondary);
        text-transform: uppercase;
        letter-spacing: 0.15em;
    }

    /* User profile section */
    .user-profile {
        display: flex;
        align-items: center;
        gap: 0.75rem;
        padding: 0.875rem 1rem;
        background: var(--bg-tertiary);
        border: 1px solid var(--border-subtle);
        border-radius: 6px;
        margin: 0 0.5rem 1rem 0.5rem;
    }

    .user-avatar {
        width: 36px;
        height: 36px;
        background: var(--border-accent);
        border-radius: 4px;
        display: flex;
        align-items: center;
        justify-content: center;
        font-family: 'VCR OSD Mono', monospace;
        font-weight: 400;
        font-size: 1rem;
        color: var(--text-primary);
    }

    .user-details {
        flex: 1;
    }

    .user-name {
        font-family: 'VCR OSD Mono', monospace;
        font-weight: 400;
        font-size: 0.875rem;
        color: var(--text-primary);
        line-height: 1.3;
    }

    .user-role {
        font-family: 'VCR OSD Mono', monospace;
        font-size: 0.625rem;
        color: var(--text-muted);
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }

    /* Navigation section */
    .nav-section {
        padding: 0 0.5rem;
    }

    .nav-label {
        font-family: 'VCR OSD Mono', monospace;
        font-size: 0.625rem;
        font-weight: 400;
        color: var(--text-muted);
        text-transform: uppercase;
        letter-spacing: 0.1em;
        padding: 0.5rem 0.75rem;
        margin-bottom: 0.25rem;
    }

    /* Navigation buttons - refined */
    [data-testid="stSidebar"] .stButton > button {
        background: transparent;
        color: var(--text-secondary);
        border: none;
        border-radius: 4px;
        padding: 0.75rem 1rem;
        margin-bottom: 2px;
        font-family: 'VCR OSD Mono', monospace;
        font-size: 0.875rem;
        font-weight: 400;
        text-align: left;
        justify-content: flex-start;
        transition: all 0.15s ease;
        position: relative;
        overflow: hidden;
    }

    [data-testid="stSidebar"] .stButton > button:hover {
        background: var(--bg-tertiary);
        color: var(--text-primary);
        transform: none;
        box-shadow: none;
    }

    [data-testid="stSidebar"] .stButton > button:active {
        background: var(--bg-card);
    }

    /* Active navigation state */
    [data-testid="stSidebar"] .nav-active > button {
        background: var(--bg-tertiary) !important;
        color: var(--text-primary) !important;
        border-left: 2px solid var(--text-secondary) !important;
        border-radius: 0 4px 4px 0 !important;
    }

    /* Logout button special styling */
    .logout-section {
        margin-top: auto;
        padding: 1rem 0.5rem;
        border-top: 1px solid var(--border-subtle);
    }

    .logout-section .stButton > button {
        background: transparent !important;
        border: 1px solid var(--border-subtle) !important;
        color: var(--text-muted) !important;
    }

    .logout-section .stButton > button:hover {
        border-color: var(--accent-rose) !important;
        color: var(--accent-rose) !important;
        background: rgba(248, 81, 73, 0.1) !important;
    }

    /* Main content area */
    .main .block-container {
        padding: 2rem 3rem;
        max-width: 1400px;
    }

    /* Page header */
    .page-header {
        margin-bottom: 2rem;
        padding-bottom: 1.5rem;
        border-bottom: 1px solid var(--border-subtle);
    }

    .page-title {
        font-family: 'VCR OSD Mono', monospace;
        font-size: 1.5rem;
        font-weight: 400;
        color: var(--text-primary);
        margin: 0 0 0.5rem 0;
        letter-spacing: 0.05em;
    }

    .page-subtitle {
        font-family: 'VCR OSD Mono', monospace;
        font-size: 0.875rem;
        color: var(--text-muted);
        margin: 0;
    }

    /* Metric cards - GitHub style */
    .metric-card {
        background: var(--bg-card);
        border: 1px solid var(--border-subtle);
        border-radius: 6px;
        padding: 1.25rem;
        position: relative;
        overflow: hidden;
        transition: border-color 0.15s ease;
    }

    .metric-card:hover {
        border-color: var(--border-accent);
    }

    .metric-label {
        font-family: 'VCR OSD Mono', monospace;
        font-size: 0.625rem;
        font-weight: 400;
        color: var(--text-muted);
        text-transform: uppercase;
        letter-spacing: 0.1em;
        margin-bottom: 0.5rem;
    }

    .metric-value {
        font-family: 'VCR OSD Mono', monospace;
        font-size: 1.5rem;
        font-weight: 400;
        color: var(--text-primary);
        line-height: 1;
        margin-bottom: 0.25rem;
    }

    .metric-value.gold {
        color: var(--accent-emerald);
    }

    .metric-change {
        font-family: 'VCR OSD Mono', monospace;
        font-size: 0.75rem;
        color: var(--accent-emerald);
    }

    .metric-change.negative {
        color: var(--accent-rose);
    }

    /* Section headers */
    .section-header {
        display: flex;
        align-items: center;
        gap: 0.75rem;
        margin: 2rem 0 1rem 0;
    }

    .section-title {
        font-family: 'VCR OSD Mono', monospace;
        font-size: 0.75rem;
        font-weight: 400;
        color: var(--text-muted);
        text-transform: uppercase;
        letter-spacing: 0.1em;
    }

    .section-line {
        flex: 1;
        height: 1px;
        background: var(--border-subtle);
    }

    /* File upload areas */
    [data-testid="stFileUploader"] {
        background: var(--bg-secondary);
        border: 1px dashed var(--border-accent);
        border-radius: 6px;
        padding: 1.25rem;
        transition: all 0.15s ease;
    }

    [data-testid="stFileUploader"]:hover {
        border-color: var(--text-secondary);
        background: var(--bg-tertiary);
    }

    [data-testid="stFileUploader"] label {
        font-family: 'VCR OSD Mono', monospace !important;
        font-weight: 400 !important;
        color: var(--text-secondary) !important;
    }

    /* Primary action buttons - GitHub style */
    .stButton > button {
        font-family: 'VCR OSD Mono', monospace;
        font-weight: 400;
        font-size: 0.875rem;
        letter-spacing: 0.02em;
        background: var(--bg-tertiary);
        color: var(--text-primary);
        border: 1px solid var(--border-accent);
        border-radius: 6px;
        padding: 0.625rem 1rem;
        transition: all 0.15s ease;
        box-shadow: none;
    }

    .stButton > button:hover {
        background: var(--border-accent);
        border-color: var(--text-muted);
    }

    .stButton > button:active {
        background: var(--bg-card);
    }

    .stButton > button:disabled {
        background: var(--bg-secondary);
        color: var(--text-muted);
        border-color: var(--border-subtle);
        cursor: not-allowed;
    }

    /* Download button special */
    .stDownloadButton > button {
        background: transparent;
        border: 1px solid var(--accent-emerald);
        color: var(--accent-emerald);
    }

    .stDownloadButton > button:hover {
        background: rgba(63, 185, 80, 0.1);
        border-color: var(--accent-emerald);
    }

    /* Data tables */
    .stDataFrame {
        border: 1px solid var(--border-subtle);
        border-radius: 6px;
        overflow: hidden;
    }

    .stDataFrame [data-testid="stDataFrameResizable"] {
        background: var(--bg-secondary);
    }

    /* Alerts and messages */
    .stAlert {
        background: var(--bg-card);
        border: 1px solid var(--border-subtle);
        border-radius: 6px;
        font-family: 'VCR OSD Mono', monospace;
    }

    .stAlert [data-testid="stMarkdownContainer"] p {
        color: var(--text-secondary);
    }

    /* Success state */
    .stSuccess {
        background: rgba(63, 185, 80, 0.1);
        border-color: var(--accent-emerald);
    }

    /* Warning state */
    .stWarning {
        background: rgba(139, 148, 158, 0.1);
        border-color: var(--accent-primary);
    }

    /* Error state */
    .stError {
        background: rgba(248, 81, 73, 0.1);
        border-color: var(--accent-rose);
    }

    /* Progress bar */
    .stProgress > div > div {
        background: var(--text-secondary);
        border-radius: 3px;
    }

    .stProgress > div {
        background: var(--bg-tertiary);
        border-radius: 3px;
    }

    /* Expander */
    .streamlit-expanderHeader {
        font-family: 'VCR OSD Mono', monospace;
        font-weight: 400;
        color: var(--text-secondary);
        background: var(--bg-card);
        border: 1px solid var(--border-subtle);
        border-radius: 6px;
    }

    .streamlit-expanderHeader:hover {
        color: var(--text-primary);
        border-color: var(--border-accent);
    }

    /* Multiselect */
    .stMultiSelect [data-baseweb="select"] {
        background: var(--bg-secondary);
        border: 1px solid var(--border-subtle);
        border-radius: 6px;
    }

    .stMultiSelect [data-baseweb="select"]:hover {
        border-color: var(--text-secondary);
    }

    /* Metrics from Streamlit */
    [data-testid="stMetricValue"] {
        font-family: 'VCR OSD Mono', monospace;
        font-size: 1.5rem;
        color: var(--text-primary);
    }

    [data-testid="stMetricLabel"] {
        font-family: 'VCR OSD Mono', monospace;
        font-size: 0.625rem;
        font-weight: 400;
        color: var(--text-muted);
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }

    /* Info box styling */
    .info-box {
        background: var(--bg-card);
        border: 1px solid var(--border-subtle);
        border-radius: 6px;
        padding: 1.25rem;
        font-family: 'VCR OSD Mono', monospace;
    }

    .info-box strong {
        color: var(--text-primary);
    }

    /* Dividers */
    hr {
        border: none;
        height: 1px;
        background: var(--border-subtle);
        margin: 1.5rem 0;
    }

    /* Scrollbar styling */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }

    ::-webkit-scrollbar-track {
        background: var(--bg-secondary);
    }

    ::-webkit-scrollbar-thumb {
        background: var(--border-accent);
        border-radius: 4px;
    }

    ::-webkit-scrollbar-thumb:hover {
        background: var(--text-muted);
    }

    /* Quick action cards */
    .action-card {
        background: var(--bg-card);
        border: 1px solid var(--border-subtle);
        border-radius: 6px;
        padding: 1.25rem;
        text-align: center;
        transition: border-color 0.15s ease;
        cursor: pointer;
    }

    .action-card:hover {
        border-color: var(--border-accent);
    }

    .action-icon {
        width: 40px;
        height: 40px;
        margin: 0 auto 0.75rem auto;
        background: var(--bg-tertiary);
        border-radius: 6px;
        display: flex;
        align-items: center;
        justify-content: center;
    }

    .action-title {
        font-family: 'VCR OSD Mono', monospace;
        font-weight: 400;
        font-size: 0.875rem;
        color: var(--text-primary);
        margin-bottom: 0.25rem;
    }

    .action-desc {
        font-family: 'VCR OSD Mono', monospace;
        font-size: 0.75rem;
        color: var(--text-muted);
    }

    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)


def get_page_slug(page_name: str) -> str:
    """Convert page name to URL slug."""
    slugs = {
        "Dashboard": "dashboard",
        "Profit Dashboard": "profit",
        "Raport OP-uri": "raport-opuri",
        "Procesare Facturi": "procesare",
        "Incasari MT940": "incasari",
        "Sincronizare Date": "sincronizare",
        "Setari": "setari"
    }
    return slugs.get(page_name, "dashboard")


def get_page_from_slug(slug: str) -> str:
    """Convert URL slug to page name."""
    pages = {
        "dashboard": "Dashboard",
        "profit": "Profit Dashboard",
        "raport-opuri": "Raport OP-uri",
        "procesare": "Procesare Facturi",
        "incasari": "Incasari MT940",
        "sincronizare": "Sincronizare Date",
        "setari": "Setari"
    }
    return pages.get(slug, "Raport OP-uri")


def navigate_to(page_name: str):
    """Navigate to a page and update URL."""
    st.session_state.current_page = page_name
    slug = get_page_slug(page_name)
    st.query_params["page"] = slug


def main():
    # Initialize session state for auth
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
        st.session_state.username = None
        st.session_state.name = None

    # Check authentication FIRST - block all access without login
    authenticated = is_authenticated()

    if not authenticated:
        # Show login form - no access without authentication
        login_form()
        return

    # All pages (only accessible after authentication)
    all_pages = [
        "Dashboard",
        "Profit Dashboard",
        "Raport OP-uri",
        "Procesare Facturi",
        "Incasari MT940",
        "Sincronizare Date",
        "Setari"
    ]

    # Get page from URL query params
    url_page = st.query_params.get("page", None)

    # Initialize or sync current page from URL
    if url_page:
        page_from_url = get_page_from_slug(url_page)
        if page_from_url in all_pages:
            st.session_state.current_page = page_from_url
        else:
            st.session_state.current_page = "Dashboard"
            st.query_params["page"] = "dashboard"
    elif 'current_page' not in st.session_state:
        st.session_state.current_page = "Dashboard"
        st.query_params["page"] = "dashboard"

    # Sidebar - always visible
    with st.sidebar:
        # Brand header
        st.markdown("""
        <div class="brand-header">
            <div class="brand-logo">
                <img src="https://gomagcdn.ro/domains3/obsid.ro/files/company/parfumuri-arabesti8220.svg" alt="OBSID">
                <div class="brand-text">
                    <span class="brand-name">Ultimate Facturi</span>
                    <span class="brand-tagline">OBSID Dashboard</span>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # User profile - authenticated user
        user_name = get_user_name()
        user_initial = user_name[0].upper() if user_name else 'A'
        user_role = "Administrator"
        st.markdown(f"""
        <div class="user-profile">
            <div class="user-avatar">{user_initial}</div>
            <div class="user-details">
                <div class="user-name">{user_name}</div>
                <div class="user-role">{user_role}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Navigation
        st.markdown('<div class="nav-section">', unsafe_allow_html=True)
        st.markdown('<div class="nav-label">Meniu Principal</div>', unsafe_allow_html=True)

        # Navigation items
        nav_items = [
            ("Dashboard", "Vedere generala"),
            ("Profit Dashboard", "Profit zilnic/lunar/anual"),
            ("Raport OP-uri", "Export contabilitate"),
            ("Procesare Facturi", "Incarca si proceseaza"),
            ("Incasari MT940", "Extrase bancare"),
            ("Sincronizare Date", "Oblio si MT940"),
            ("Setari", "Configurare")
        ]

        for page_name, _ in nav_items:
            is_active = st.session_state.current_page == page_name

            if is_active:
                st.markdown('<div class="nav-active">', unsafe_allow_html=True)

            if st.button(page_name, key=f"nav_{page_name}", use_container_width=True):
                navigate_to(page_name)
                st.rerun()

            if is_active:
                st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('</div>', unsafe_allow_html=True)

        # Logout at bottom
        st.markdown("---")
        st.markdown('<div class="logout-section">', unsafe_allow_html=True)
        if st.button("Deconectare", key="logout_btn", use_container_width=True):
            logout()
        st.markdown('</div>', unsafe_allow_html=True)

    # Main content
    page = st.session_state.get('current_page', 'Raport OP-uri')

    # Ensure URL is in sync with current page
    current_slug = st.query_params.get("page", "")
    expected_slug = get_page_slug(page)
    if current_slug != expected_slug:
        st.query_params["page"] = expected_slug

    if page == "Dashboard":
        show_dashboard()
    elif page == "Profit Dashboard":
        show_profit_dashboard()
    elif page == "Raport OP-uri":
        show_raport_opuri()
    elif page == "Procesare Facturi":
        show_procesare()
    elif page == "Incasari MT940":
        show_incasari()
    elif page == "Sincronizare Date":
        show_data_sync()
    elif page == "Setari":
        show_setari()


def show_dashboard():
    """Pagina principala cu sumar."""
    # Page header
    st.markdown("""
    <div class="page-header">
        <h1 class="page-title">Dashboard</h1>
        <p class="page-subtitle">Vedere generala asupra procesarii facturilor si incasarilor</p>
    </div>
    """, unsafe_allow_html=True)

    # Get stored data
    incasari = st.session_state.get('incasari_mt940', [])
    rezultate_gls = st.session_state.get('rezultate_gls', [])
    rezultate_sameday = st.session_state.get('rezultate_sameday', [])

    total_facturi = len(rezultate_gls) + len(rezultate_sameday)
    total_incasari = len(incasari)
    total_suma = sum(i[1] for i in incasari) if incasari else 0

    # Metrics row
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Facturi Procesate</div>
            <div class="metric-value">{total_facturi}</div>
        </div>
        """, unsafe_allow_html=True)

    with col2:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Incasari MT940</div>
            <div class="metric-value">{total_incasari}</div>
        </div>
        """, unsafe_allow_html=True)

    with col3:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Total Incasari</div>
            <div class="metric-value gold">{total_suma:,.2f} RON</div>
        </div>
        """, unsafe_allow_html=True)

    with col4:
        erori = len(st.session_state.get('erori', []))
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Erori</div>
            <div class="metric-value">{erori}</div>
        </div>
        """, unsafe_allow_html=True)

    # Section header
    st.markdown("""
    <div class="section-header">
        <span class="section-title">Actiuni Rapide</span>
        <div class="section-line"></div>
    </div>
    """, unsafe_allow_html=True)

    # Quick actions
    col1, col2, col3 = st.columns(3)

    with col1:
        if st.button("Procesare Noua", use_container_width=True, key="dash_process"):
            navigate_to('Procesare Facturi')
            st.rerun()

    with col2:
        if st.button("Vizualizeaza Incasari", use_container_width=True, key="dash_incasari"):
            navigate_to('Incasari MT940')
            st.rerun()

    with col3:
        if st.button("Export Raport", use_container_width=True, key="dash_export"):
            if not incasari:
                st.warning("Incarca mai intai fisierele pentru procesare")


def show_procesare():
    """Pagina de procesare facturi - Export OP-uri."""
    st.markdown("""
    <div class="page-header">
        <h1 class="page-title">Export OP-uri</h1>
        <p class="page-subtitle">Genereaza raportul de facturi grupate pe OP-uri bancare</p>
    </div>
    """, unsafe_allow_html=True)

    # Import processor
    try:
        from utils.opuri_processor import generate_opuri_export
    except ImportError as e:
        st.error(f"Eroare la import: {e}")
        return

    # Show data availability from Supabase
    from utils.supabase_client import get_client
    client = get_client()

    if client:
        try:
            gls_count = len(client.table("gls_parcels").select("id", count="exact").execute().data)
            sameday_count = len(client.table("sameday_parcels").select("id", count="exact").execute().data)
            netopia_count = len(client.table("netopia_transactions").select("id", count="exact").execute().data)
            invoices_count = len(client.table("invoices").select("id", count="exact").execute().data)
            mt940_count = len(client.table("bank_transactions").select("id", count="exact").execute().data)

            st.markdown("""
            <div class="section-header">
                <span class="section-title">Date Disponibile in Supabase</span>
                <div class="section-line"></div>
            </div>
            """, unsafe_allow_html=True)

            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                st.metric("GLS", gls_count)
            with col2:
                st.metric("Sameday", sameday_count)
            with col3:
                st.metric("Netopia", netopia_count)
            with col4:
                st.metric("Facturi Oblio", invoices_count)
            with col5:
                st.metric("Tranzactii MT940", mt940_count)

            if gls_count == 0 and sameday_count == 0:
                st.warning("Nu exista date in Supabase. Mergi la **Sincronizare Date** pentru a descarca datele.")
        except Exception as e:
            st.warning(f"Nu s-a putut verifica baza de date: {e}")

    st.markdown("---")

    # Period selection
    st.markdown("""
    <div class="section-header">
        <span class="section-title">Selecteaza Perioada</span>
        <div class="section-line"></div>
    </div>
    """, unsafe_allow_html=True)

    from datetime import date, timedelta
    today = date.today()
    first_day_of_month = today.replace(day=1)

    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        start_date = st.date_input("De la", value=first_day_of_month, key="proc_start")
    with col2:
        end_date = st.date_input("Pana la", value=today, key="proc_end")
    with col3:
        # Quick period buttons
        st.write("")  # Spacing
        col_q1, col_q2, col_q3 = st.columns(3)
        with col_q1:
            if st.button("Luna curenta", key="btn_proc_luna_curenta", use_container_width=True):
                st.session_state.proc_start = first_day_of_month
                st.session_state.proc_end = today
                st.rerun()
        with col_q2:
            if st.button("Luna trecuta", key="btn_proc_luna_trecuta", use_container_width=True):
                last_month_end = first_day_of_month - timedelta(days=1)
                last_month_start = last_month_end.replace(day=1)
                st.session_state.proc_start = last_month_start
                st.session_state.proc_end = last_month_end
                st.rerun()
        with col_q3:
            if st.button("Ultimele 60 zile", key="btn_proc_60zile", use_container_width=True):
                st.session_state.proc_start = today - timedelta(days=60)
                st.session_state.proc_end = today
                st.rerun()

    st.markdown("---")

    # Optional Gomag file
    st.markdown("""
    <div class="section-header">
        <span class="section-title">Fisier Gomag (Optional)</span>
        <div class="section-line"></div>
    </div>
    """, unsafe_allow_html=True)

    st.info("Fisierul Gomag este folosit pentru a potrivi AWB-urile cu Order ID-urile. Daca nu il incarci, Order ID-urile vor fi goale.")

    gomag_file = st.file_uploader(
        "Fisier Gomag (XLSX)",
        type=['xlsx'],
        key="gomag",
        help="Exportul comenzilor din Gomag"
    )

    st.markdown("---")

    # Generate button
    if st.button("GENEREAZA EXPORT OP-URI", use_container_width=True, type="primary"):
        with st.spinner("Se genereaza raportul..."):
            try:
                # Read Gomag if provided
                gomag_df = None
                if gomag_file:
                    gomag_df = pd.read_excel(gomag_file, dtype=str)
                    st.info(f"Gomag incarcat: {len(gomag_df)} randuri")

                # Generate export
                excel_buffer = generate_opuri_export(
                    start_date.strftime('%Y-%m-%d'),
                    end_date.strftime('%Y-%m-%d'),
                    gomag_df
                )

                st.success("Export generat cu succes!")

                # Download button
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                st.download_button(
                    label="DESCARCA RAPORT EXCEL",
                    data=excel_buffer,
                    file_name=f"opuri_export_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
            except Exception as e:
                st.error(f"Eroare la generare: {str(e)}")
                import traceback
                st.code(traceback.format_exc())


# NOTE: process_files function removed - now using opuri_processor.generate_opuri_export()


def show_incasari():
    """Pagina cu incasarile MT940."""
    st.markdown("""
    <div class="page-header">
        <h1 class="page-title">Incasari MT940</h1>
        <p class="page-subtitle">Vizualizare extrase bancare procesate</p>
    </div>
    """, unsafe_allow_html=True)

    if 'incasari_mt940' not in st.session_state or not st.session_state['incasari_mt940']:
        st.info("Nu exista incasari procesate. Mergi la 'Procesare Facturi' pentru a incarca fisierele MT940.")
        return

    incasari = st.session_state['incasari_mt940']

    # Summary metrics
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Incasari", len(incasari))
    with col2:
        total_suma = sum(i[1] for i in incasari)
        st.metric("Suma Totala", f"{total_suma:,.2f} RON")
    with col3:
        surse = {}
        for i in incasari:
            sursa = get_sursa_incasare(i[4])
            surse[sursa] = surse.get(sursa, 0) + 1
        st.metric("Surse", len(surse))

    st.markdown("---")

    # Filter
    surse_disponibile = list(set(get_sursa_incasare(i[4]) for i in incasari))
    sursa_filter = st.multiselect("Filtreaza dupa sursa", surse_disponibile, default=surse_disponibile)

    # Table
    data = []
    for op_ref, suma, data_op, batchid, details in incasari:
        sursa = get_sursa_incasare(details)
        if sursa in sursa_filter:
            data.append({
                'Data': data_op,
                'Referinta OP': op_ref,
                'Sursa': sursa,
                'Suma': f"{suma:,.2f} RON",
                'BatchID': batchid or '-'
            })

    if data:
        df = pd.DataFrame(data)
        st.dataframe(df, use_container_width=True, hide_index=True)
    else:
        st.info("Nu exista incasari pentru filtrele selectate.")


def show_setari():
    """Pagina de setari."""
    st.markdown("""
    <div class="page-header">
        <h1 class="page-title">Setari</h1>
        <p class="page-subtitle">Configurare si informatii despre aplicatie</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="section-header">
        <span class="section-title">Informatii Aplicatie</span>
        <div class="section-line"></div>
    </div>
    """, unsafe_allow_html=True)

    st.info("""
    **Ultimate Facturi OBSID**
    Versiune: 2.0.0

    Aplicatie pentru procesarea si gruparea facturilor cu sincronizare automata:
    - **GLS**: Colete cu ramburs din API MyGLS
    - **Sameday**: Colete cu ramburs din API Sameday
    - **Netopia**: Rapoarte decontare din email
    - **Oblio**: Facturi sincronizate din API
    - **MT940**: Extrase bancare Banca Transilvania
    """)

    st.markdown("""
    <div class="section-header">
        <span class="section-title">Despre</span>
        <div class="section-line"></div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    Aceasta aplicatie automatizeaza procesul de reconciliere a facturilor cu incasarile bancare.

    **Functionalitati:**
    - Parsare automata fisiere MT940 de la Banca Transilvania
    - Sincronizare automata colete GLS si Sameday din API
    - Sincronizare automata rapoarte Netopia din email
    - Sincronizare facturi Oblio din API
    - Grupare facturi pe OP-uri bancare
    - Export Excel cu toate datele procesate
    """)


def show_profit_dashboard():
    """Pagina cu profit pe zile/luni/ani."""
    st.markdown("""
    <div class="page-header">
        <h1 class="page-title">Profit Dashboard</h1>
        <p class="page-subtitle">Vizualizare profit pe perioade de timp</p>
    </div>
    """, unsafe_allow_html=True)

    # Period selector
    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        period = st.selectbox("Perioada", ["Lunar", "Zilnic", "Anual"], key="profit_period")
    with col2:
        from datetime import date, timedelta
        default_start = date.today() - timedelta(days=365)
        start_date = st.date_input("De la", value=default_start, key="profit_start")
    with col3:
        end_date = st.date_input("Pana la", value=date.today(), key="profit_end")

    # Map period to group_by
    group_map = {"Zilnic": "day", "Lunar": "month", "Anual": "year"}
    group_by = group_map[period]

    try:
        # Get profit data from Supabase
        profit_data = get_profit_data(start_date, end_date, group_by)

        if not profit_data:
            st.info("Nu exista date in baza de date. Sincronizeaza datele MT940 din pagina 'Sincronizare Date'.")
            return

        # Prepare data for chart
        dates = [d['date'] for d in profit_data]
        totals = [d['total'] for d in profit_data]

        # Summary metrics
        total_sum = sum(totals)
        avg_sum = total_sum / len(totals) if totals else 0
        max_sum = max(totals) if totals else 0
        min_sum = min(totals) if totals else 0

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">Total Perioada</div>
                <div class="metric-value gold">{total_sum:,.2f} RON</div>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">Medie {period.lower()}</div>
                <div class="metric-value">{avg_sum:,.2f} RON</div>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">Maxim</div>
                <div class="metric-value">{max_sum:,.2f} RON</div>
            </div>
            """, unsafe_allow_html=True)
        with col4:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">Minim</div>
                <div class="metric-value">{min_sum:,.2f} RON</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        # Main chart
        st.markdown("""
        <div class="section-header">
            <span class="section-title">Evolutie Profit</span>
            <div class="section-line"></div>
        </div>
        """, unsafe_allow_html=True)

        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=dates,
            y=totals,
            marker_color='#3fb950',
            name='Profit'
        ))
        fig.update_layout(
            plot_bgcolor='#0d1117',
            paper_bgcolor='#0d1117',
            font=dict(family='VCR OSD Mono, monospace', color='#8b949e'),
            xaxis=dict(
                gridcolor='#30363d',
                tickfont=dict(color='#8b949e')
            ),
            yaxis=dict(
                gridcolor='#30363d',
                tickfont=dict(color='#8b949e'),
                title='RON'
            ),
            margin=dict(l=40, r=40, t=40, b=40),
            showlegend=False
        )
        st.plotly_chart(fig, use_container_width=True)

        # Breakdown by source
        st.markdown("""
        <div class="section-header">
            <span class="section-title">Distributie pe Surse</span>
            <div class="section-line"></div>
        </div>
        """, unsafe_allow_html=True)

        # Aggregate by source
        source_totals = {}
        for d in profit_data:
            for source, amount in d.get('by_source', {}).items():
                source_totals[source] = source_totals.get(source, 0) + amount

        if source_totals:
            col1, col2 = st.columns([1, 1])

            with col1:
                # Pie chart
                fig_pie = go.Figure(data=[go.Pie(
                    labels=list(source_totals.keys()),
                    values=list(source_totals.values()),
                    hole=0.4,
                    marker=dict(colors=['#3fb950', '#58a6ff', '#8b949e', '#f85149', '#c9d1d9'])
                )])
                fig_pie.update_layout(
                    plot_bgcolor='#0d1117',
                    paper_bgcolor='#0d1117',
                    font=dict(family='VCR OSD Mono, monospace', color='#8b949e'),
                    margin=dict(l=20, r=20, t=20, b=20),
                    showlegend=True,
                    legend=dict(font=dict(color='#8b949e'))
                )
                st.plotly_chart(fig_pie, use_container_width=True)

            with col2:
                # Table
                source_data = [
                    {"Sursa": k, "Total": f"{v:,.2f} RON", "Procent": f"{v/total_sum*100:.1f}%"}
                    for k, v in sorted(source_totals.items(), key=lambda x: x[1], reverse=True)
                ]
                st.dataframe(pd.DataFrame(source_data), use_container_width=True, hide_index=True)

    except Exception as e:
        st.error(f"Eroare la incarcarea datelor: {str(e)}")
        st.info("Asigura-te ca conexiunea la Supabase este configurata corect.")


def show_data_sync():
    """Pagina pentru sincronizare date cu Supabase."""
    st.markdown("""
    <div class="page-header">
        <h1 class="page-title">Sincronizare Date</h1>
        <p class="page-subtitle">Sincronizare automata GLS, Sameday, Netopia, Oblio din API-uri</p>
    </div>
    """, unsafe_allow_html=True)

    # Import all required modules
    imap_available = False
    gls_available = False
    sameday_available = False

    try:
        from utils.email_imap import is_imap_configured, test_imap_connection, get_all_netopia_batch_ids
        from utils.netopia_api import (
            sync_netopia_batch,
            test_netopia_connection,
            save_netopia_transactions_to_supabase,
            save_netopia_batch_to_supabase,
            is_batch_already_synced,
            get_synced_batches_for_month
        )
        imap_available = True
    except ImportError as e:
        pass

    try:
        from utils.gls_api import (
            is_gls_configured,
            test_gls_connection,
            get_delivered_parcels_with_cod,
            save_gls_parcels_to_supabase,
            get_existing_gls_parcels
        )
        gls_available = True
    except ImportError as e:
        pass

    try:
        from utils.sameday_api import (
            is_sameday_configured,
            test_sameday_connection,
            get_sameday_deliveries_with_cod as get_sameday_parcels,
            save_sameday_parcels_to_supabase,
            get_existing_sameday_parcels
        )
        sameday_available = True
    except ImportError as e:
        pass

    # Connection status
    st.markdown("""
    <div class="section-header">
        <span class="section-title">Status Conexiuni</span>
        <div class="section-line"></div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3, col4, col5, col6 = st.columns(6)
    with col1:
        supabase_ok = test_supabase()
        if supabase_ok:
            st.success("Supabase: Conectat")
        else:
            st.error("Supabase: Deconectat")

    with col2:
        oblio_ok = test_oblio()
        if oblio_ok:
            st.success("Oblio API: Conectat")
        else:
            st.error("Oblio API: Deconectat")

    with col3:
        if imap_available:
            if is_imap_configured():
                imap_ok = test_imap_connection()
                if imap_ok:
                    st.success("Email IMAP: Conectat")
                else:
                    st.error("Email IMAP: Eroare conectare")
            else:
                st.warning("Email IMAP: Neconfigurat")
        else:
            st.error("Email IMAP: Indisponibil")

    with col4:
        if imap_available:
            netopia_ok = test_netopia_connection()
            if netopia_ok:
                st.success("Netopia API: Configurat")
            else:
                st.warning("Netopia API: Neconfigurat")
        else:
            st.error("Netopia API: Indisponibil")

    with col5:
        if gls_available:
            if is_gls_configured():
                st.success("GLS API: Configurat")
            else:
                st.warning("GLS API: Neconfigurat")
        else:
            st.error("GLS API: Indisponibil")

    with col6:
        if sameday_available:
            if is_sameday_configured():
                st.success("Sameday API: Configurat")
            else:
                st.warning("Sameday API: Neconfigurat")
        else:
            st.error("Sameday API: Indisponibil")

    st.markdown("---")

    # ============================================
    # SINCRONIZARE TOTALA - Un singur buton pentru tot
    # ============================================
    st.markdown("""
    <div class="section-header">
        <span class="section-title">Sincronizare Totala (Recomandat)</span>
        <div class="section-line"></div>
    </div>
    """, unsafe_allow_html=True)

    st.info("""
    **Un singur click sincronizeaza toate sursele de date:**
    - **GLS**: Colete livrate cu ramburs (ultimele 60 zile)
    - **Sameday**: Colete livrate cu ramburs (ultimele 30 zile)
    - **Netopia**: Rapoarte de decontare din email (ultimele 60 zile)
    - **Oblio**: Facturi emise (ultimele 60 zile)

    **Duplicatele sunt ignorate automat** - datele existente nu se suprascriu.
    """)

    # Show existing data counts
    from utils.supabase_client import get_client
    client = get_client()
    if client:
        try:
            gls_count = len(client.table("gls_parcels").select("id", count="exact").execute().data)
            sameday_count = len(client.table("sameday_parcels").select("id", count="exact").execute().data)
            netopia_count = len(client.table("netopia_transactions").select("id", count="exact").execute().data)
            invoices_count = len(client.table("invoices").select("id", count="exact").execute().data)

            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Colete GLS", gls_count)
            with col2:
                st.metric("Colete Sameday", sameday_count)
            with col3:
                st.metric("Tranzactii Netopia", netopia_count)
            with col4:
                st.metric("Facturi Oblio", invoices_count)
        except:
            pass

    from datetime import date, timedelta

    if st.button("SINCRONIZARE TOTALA", key="btn_sync_all", use_container_width=True, type="primary"):
        progress = st.progress(0)
        status = st.empty()
        results = st.container()

        total_stats = {
            'gls_inserted': 0, 'gls_skipped': 0,
            'sameday_inserted': 0, 'sameday_skipped': 0,
            'netopia_inserted': 0, 'netopia_skipped': 0,
            'oblio_inserted': 0,
            'errors': []
        }

        # 1. Sincronizare GLS (25%)
        if gls_available and is_gls_configured():
            status.text("Sincronizare GLS...")
            try:
                parcels = get_delivered_parcels_with_cod(days_back=60)
                if parcels:
                    stats = save_gls_parcels_to_supabase(parcels)
                    total_stats['gls_inserted'] = stats.get('inserted', 0)
                    total_stats['gls_skipped'] = stats.get('skipped', 0)
                    total_stats['errors'].extend(stats.get('errors', []))
            except Exception as e:
                total_stats['errors'].append(f"GLS: {str(e)}")
        progress.progress(25)

        # 2. Sincronizare Sameday (50%)
        if sameday_available and is_sameday_configured():
            status.text("Sincronizare Sameday (poate dura cateva minute)...")
            try:
                parcels = get_sameday_parcels(days_back=30)
                if parcels:
                    stats = save_sameday_parcels_to_supabase(parcels)
                    total_stats['sameday_inserted'] = stats.get('inserted', 0)
                    total_stats['sameday_skipped'] = stats.get('skipped', 0)
                    total_stats['errors'].extend(stats.get('errors', []))
            except Exception as e:
                total_stats['errors'].append(f"Sameday: {str(e)}")
        progress.progress(50)

        # 3. Sincronizare Netopia (75%)
        if imap_available and is_imap_configured():
            status.text("Sincronizare Netopia din email...")
            try:
                batch_ids = get_all_netopia_batch_ids(days_back=60)
                netopia_key = os.getenv('NETOPIA_API_KEY', '')
                if batch_ids and netopia_key:
                    for batch in batch_ids:
                        batch_id = batch['batch_id']
                        if is_batch_already_synced(batch_id):
                            total_stats['netopia_skipped'] += 1
                            continue

                        result = sync_netopia_batch(batch.get('report_id', batch_id), netopia_key)
                        if result['success']:
                            save_netopia_transactions_to_supabase(
                                result['transactions'],
                                batch_id,
                                batch.get('report_month', '')
                            )
                            save_netopia_batch_to_supabase({
                                'batch_id': batch_id,
                                'report_id': batch.get('report_id', batch_id),
                                'date': batch.get('date', ''),
                                'subject': batch.get('subject', ''),
                                'report_month': batch.get('report_month', ''),
                                'count': result['count'],
                                'total_amount': result['total_amount'],
                                'total_fees': result['total_fees'],
                                'net_amount': result['net_amount']
                            })
                            total_stats['netopia_inserted'] += 1
            except Exception as e:
                total_stats['errors'].append(f"Netopia: {str(e)}")
        progress.progress(75)

        # 4. Sincronizare Oblio (100%)
        if oblio_ok:
            status.text("Sincronizare Oblio...")
            try:
                oblio_start = date.today() - timedelta(days=60)
                oblio_end = date.today()
                stats = sync_oblio_invoices(oblio_start, oblio_end)
                total_stats['oblio_inserted'] = stats.get('inserted', 0)
                total_stats['errors'].extend(stats.get('errors', []))
            except Exception as e:
                total_stats['errors'].append(f"Oblio: {str(e)}")
        progress.progress(100)

        status.text("Sincronizare completa!")

        # Display results
        with results:
            st.success("Sincronizare finalizata!")

            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("GLS", f"+{total_stats['gls_inserted']}", f"{total_stats['gls_skipped']} existente")
            with col2:
                st.metric("Sameday", f"+{total_stats['sameday_inserted']}", f"{total_stats['sameday_skipped']} existente")
            with col3:
                st.metric("Netopia", f"+{total_stats['netopia_inserted']}", f"{total_stats['netopia_skipped']} existente")
            with col4:
                st.metric("Oblio", f"+{total_stats['oblio_inserted']}")

            if total_stats['errors']:
                with st.expander(f"Erori ({len(total_stats['errors'])})"):
                    for err in total_stats['errors'][:20]:
                        st.warning(err)

    st.markdown("---")

    # ============================================
    # Import MT940 - Separat pentru ca e din fisiere locale
    # ============================================
    st.markdown("""
    <div class="section-header">
        <span class="section-title">Import MT940 (Extrase Bancare)</span>
        <div class="section-line"></div>
    </div>
    """, unsafe_allow_html=True)

    st.info("Importa tranzactiile bancare din fisierele MT940. Acestea sunt necesare pentru gruparea facturilor pe OP-uri.")

    mt940_files_upload = st.file_uploader(
        "Incarca fisiere MT940 (TXT)",
        type=['txt'],
        accept_multiple_files=True,
        key="sync_mt940_files"
    )

    if st.button("Import MT940", key="btn_import_mt940", use_container_width=True, disabled=not supabase_ok):
        if mt940_files_upload:
            with st.spinner("Se importa tranzactiile..."):
                try:
                    with tempfile.TemporaryDirectory() as tmpdir:
                        file_names = []
                        for mt_file in mt940_files_upload:
                            file_path = os.path.join(tmpdir, mt_file.name)
                            with open(file_path, 'wb') as f:
                                f.write(mt_file.getbuffer())
                            file_names.append(mt_file.name)

                        stats = import_mt940_to_supabase(tmpdir, file_names)

                    st.success(f"Import finalizat!")
                    col_a, col_b, col_c = st.columns(3)
                    with col_a:
                        st.metric("Procesate", stats['processed'])
                    with col_b:
                        st.metric("Inserate", stats['inserted'])
                    with col_c:
                        st.metric("Ignorate (duplicate)", stats['skipped'])

                    if stats['errors']:
                        with st.expander(f"Erori ({len(stats['errors'])})"):
                            for err in stats['errors'][:10]:
                                st.warning(err)
                except Exception as e:
                    st.error(f"Eroare la import: {str(e)}")
        else:
            st.warning("Incarca fisiere MT940 pentru import")


    # Sync logs
    st.markdown("---")
    st.markdown("""
    <div class="section-header">
        <span class="section-title">Istoric Sincronizari</span>
        <div class="section-line"></div>
    </div>
    """, unsafe_allow_html=True)

    try:
        logs = get_recent_sync_logs(10)
        if logs:
            log_data = []
            for log in logs:
                log_data.append({
                    'Data': log.get('started_at', '-')[:19].replace('T', ' ') if log.get('started_at') else '-',
                    'Tip': log.get('sync_type', '-'),
                    'Status': log.get('status', '-'),
                    'Procesate': log.get('records_processed', 0),
                    'Inserate': log.get('records_inserted', 0),
                    'Ignorate': log.get('records_skipped', 0)
                })
            st.dataframe(pd.DataFrame(log_data), use_container_width=True, hide_index=True)
        else:
            st.info("Nu exista sincronizari anterioare.")
    except Exception as e:
        st.warning(f"Nu s-a putut incarca istoricul: {str(e)}")


def show_raport_opuri():
    """Pagina Raport OP-uri pentru contabilitate."""
    st.markdown("""
    <div class="page-header">
        <h1 class="page-title">Raport OP-uri</h1>
        <p class="page-subtitle">Raport pentru contabilitate - Export facturi grupate pe OP-uri cu matching automat</p>
    </div>
    """, unsafe_allow_html=True)

    # Period selector
    st.markdown("""
    <div class="section-header">
        <span class="section-title">Selecteaza Perioada</span>
        <div class="section-line"></div>
    </div>
    """, unsafe_allow_html=True)

    from datetime import date, timedelta

    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        # Default to current month
        today = date.today()
        first_day_of_month = today.replace(day=1)
        start_date = st.date_input("De la", value=first_day_of_month, key="raport_start")
    with col2:
        end_date = st.date_input("Pana la", value=today, key="raport_end")
    with col3:
        # Quick period buttons
        st.write("")  # Spacing
        col_q1, col_q2, col_q3 = st.columns(3)
        with col_q1:
            if st.button("Luna curenta", key="btn_luna_curenta", use_container_width=True):
                st.session_state.raport_start = first_day_of_month
                st.session_state.raport_end = today
                st.rerun()
        with col_q2:
            if st.button("Luna trecuta", key="btn_luna_trecuta", use_container_width=True):
                last_month_end = first_day_of_month - timedelta(days=1)
                last_month_start = last_month_end.replace(day=1)
                st.session_state.raport_start = last_month_start
                st.session_state.raport_end = last_month_end
                st.rerun()
        with col_q3:
            if st.button("Tot anul", key="btn_tot_anul", use_container_width=True):
                st.session_state.raport_start = date(today.year, 1, 1)
                st.session_state.raport_end = today
                st.rerun()

    st.markdown("---")

    try:
        # Get data for the period
        transactions = get_transactions_for_period(start_date, end_date)
        invoices = get_invoices_for_period(start_date, end_date)

        if not transactions and not invoices:
            st.info("Nu exista date pentru perioada selectata. Importa date MT940 si sincronizeaza facturi din pagina 'Sincronizare Date'.")
            return

        # Get matching statistics
        stats = get_matching_statistics(start_date, end_date)

        # Summary metrics
        col1, col2, col3, col4, col5 = st.columns(5)

        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">Total Incasari</div>
                <div class="metric-value gold">{stats['transactions_amount']:,.2f} RON</div>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">Tranzactii</div>
                <div class="metric-value">{stats['total_transactions']}</div>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">Facturi Oblio</div>
                <div class="metric-value">{stats['total_invoices']}</div>
            </div>
            """, unsafe_allow_html=True)
        with col4:
            diff = stats['difference']
            diff_color = "emerald" if abs(diff) < 1 else "rose"
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">Diferenta</div>
                <div class="metric-value {diff_color}">{diff:+,.2f} RON</div>
            </div>
            """, unsafe_allow_html=True)
        with col5:
            surse_count = len(stats.get('by_source', {}))
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">Surse Incasari</div>
                <div class="metric-value">{surse_count}</div>
            </div>
            """, unsafe_allow_html=True)

        # Source breakdown
        if stats.get('by_source'):
            st.markdown("""
            <div class="section-header">
                <span class="section-title">Incasari pe Surse</span>
                <div class="section-line"></div>
            </div>
            """, unsafe_allow_html=True)

            source_cols = st.columns(len(stats['by_source']))
            for i, (source, data) in enumerate(stats['by_source'].items()):
                with source_cols[i]:
                    color_map = {'GLS': 'blue', 'Sameday': 'rose', 'Netopia': 'cyan', 'eMag': 'orange'}
                    color = color_map.get(source, 'default')
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">{source}</div>
                        <div class="metric-value">{data['count']} tranz.</div>
                        <div class="metric-sublabel">{data['amount']:,.2f} RON</div>
                    </div>
                    """, unsafe_allow_html=True)

        # Matching section with button
        st.markdown("""
        <div class="section-header">
            <span class="section-title">Matching Tranzactii - Facturi</span>
            <div class="section-line"></div>
        </div>
        """, unsafe_allow_html=True)

        col_match1, col_match2 = st.columns([1, 3])
        with col_match1:
            run_matching = st.button("Ruleaza Matching", key="btn_run_matching", use_container_width=True, type="primary")

        if run_matching or st.session_state.get('matching_results'):
            with st.spinner("Se proceseaza matching-ul..."):
                # Generate report data with matching
                report_data = generate_opuri_report_data(start_date, end_date)
                st.session_state['matching_results'] = report_data

            if report_data:
                # Count matches and errors
                matched_count = sum(1 for r in report_data if r.get('numar_factura'))
                error_count = sum(1 for r in report_data if r.get('erori') == 'DA')

                st.success(f"Matching complet: {matched_count} potriviri, {error_count} erori")

                # Display matched data
                df_report = pd.DataFrame(report_data)
                df_report = df_report.rename(columns={
                    'data_op': 'Data OP',
                    'numar_op': 'Numar OP',
                    'nume_borderou': 'Borderou',
                    'curier': 'Curier',
                    'order_id': 'Order ID',
                    'numar_factura': 'Nr. Factura',
                    'suma': 'Suma',
                    'erori': 'Erori',
                    'diferenta_emag': 'Dif. eMag',
                    'facturi_comision_emag': 'Facturi Comision'
                })

                # Style the dataframe
                def highlight_errors(row):
                    if row.get('Erori') == 'DA':
                        return ['background-color: rgba(248, 81, 73, 0.2)'] * len(row)
                    return [''] * len(row)

                st.dataframe(
                    df_report.style.apply(highlight_errors, axis=1),
                    use_container_width=True,
                    hide_index=True
                )
            else:
                st.warning("Nu s-au gasit date pentru matching in perioada selectata.")

        # Transactions table (raw data)
        if transactions:
            with st.expander("Tranzactii Bancare (date brute)", expanded=False):
                display_data = []
                for t in transactions:
                    display_data.append({
                        'Data OP': t.get('transaction_date', ''),
                        'Numar OP': t.get('op_reference', ''),
                        'Curier': t.get('source', ''),
                        'Suma': f"{float(t.get('amount', 0)):,.2f} RON",
                        'Detalii': (t.get('details', '') or '')[:50] + '...' if t.get('details') and len(t.get('details', '')) > 50 else t.get('details', '')
                    })
                df_display = pd.DataFrame(display_data)
                st.dataframe(df_display, use_container_width=True, hide_index=True)

        # Invoices section
        if invoices:
            with st.expander("Facturi Oblio (date brute)", expanded=False):
                inv_data = []
                for inv in invoices:
                    inv_data.append({
                        'Data': inv.get('issue_date', ''),
                        'Serie': inv.get('series_name', ''),
                        'Numar': inv.get('invoice_number', ''),
                        'Client': inv.get('client_name', ''),
                        'Total': f"{float(inv.get('total', 0)):,.2f} RON",
                        'Tip': inv.get('invoice_type', ''),
                        'Incasata': 'DA' if inv.get('is_collected') else 'NU'
                    })
                df_inv = pd.DataFrame(inv_data)
                st.dataframe(df_inv, use_container_width=True, hide_index=True)

        # Export section
        st.markdown("""
        <div class="section-header">
            <span class="section-title">Export Excel</span>
            <div class="section-line"></div>
        </div>
        """, unsafe_allow_html=True)

        col1, col2 = st.columns(2)
        with col1:
            # Generate Excel with matching data
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            buffer = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "OP-uri"

            # Headers matching opuri_export.xlsx format
            headers = ["Data OP", "Numar OP", "Nume Borderou", "Curier", "Order ID", "Numar Factura", "Suma", "Erori", "Diferenta eMag", "Facturi Comision eMag"]
            ws.append(headers)

            # Style header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Color fills
            gls_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            sameday_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            netopia_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
            emag_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            error_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            # Use matching results if available, otherwise raw transactions
            export_data = st.session_state.get('matching_results', [])
            if not export_data:
                export_data = [{
                    'data_op': t.get('transaction_date', ''),
                    'numar_op': t.get('op_reference', ''),
                    'nume_borderou': t.get('file_name', ''),
                    'curier': t.get('source', ''),
                    'order_id': '',
                    'numar_factura': '',
                    'suma': float(t.get('amount', 0)),
                    'erori': 'NU',
                    'diferenta_emag': '',
                    'facturi_comision_emag': ''
                } for t in transactions]

            for data in export_data:
                ws.append([
                    data.get('data_op', ''),
                    data.get('numar_op', ''),
                    data.get('nume_borderou', ''),
                    data.get('curier', ''),
                    data.get('order_id', ''),
                    data.get('numar_factura', ''),
                    data.get('suma', 0),
                    data.get('erori', 'NU'),
                    data.get('diferenta_emag', ''),
                    data.get('facturi_comision_emag', '')
                ])

                row_idx = ws.max_row
                source = data.get('curier', '')

                # Color courier cell
                if source == 'GLS':
                    ws.cell(row=row_idx, column=4).fill = gls_fill
                    ws.cell(row=row_idx, column=4).font = Font(color="FFFFFF")
                elif source == 'Sameday':
                    ws.cell(row=row_idx, column=4).fill = sameday_fill
                    ws.cell(row=row_idx, column=4).font = Font(color="FFFFFF")
                elif source == 'Netopia':
                    ws.cell(row=row_idx, column=4).fill = netopia_fill
                elif source == 'eMag':
                    ws.cell(row=row_idx, column=4).fill = emag_fill

                # Color error cell
                if data.get('erori') == 'DA':
                    ws.cell(row=row_idx, column=8).fill = error_fill
                    ws.cell(row=row_idx, column=8).font = Font(color="FFFFFF")

            # Adjust column widths
            column_widths = [12, 22, 25, 12, 15, 15, 12, 8, 15, 35]
            for col_idx, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col_idx)].width = width

            wb.save(buffer)
            buffer.seek(0)

            st.download_button(
                label="Descarca opuri_export.xlsx",
                data=buffer,
                file_name=f"opuri_export_{start_date}_{end_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )

        with col2:
            st.markdown("""
            <div class="info-card">
                <p><strong>Acest raport contine:</strong></p>
                <ul>
                    <li>Tranzactii bancare din MT940</li>
                    <li>Matching automat cu facturi Oblio</li>
                    <li>Cautare in Oblio API pentru AWB-uri fara factura</li>
                    <li>Colorare pe surse (GLS, Sameday, Netopia, eMag)</li>
                    <li>Marcare erori pentru tranzactii nepotrivite</li>
                </ul>
            </div>
            """, unsafe_allow_html=True)

    except Exception as e:
        st.error(f"Eroare la incarcarea datelor: {str(e)}")
        import traceback
        st.code(traceback.format_exc())


if __name__ == "__main__":
    main()
