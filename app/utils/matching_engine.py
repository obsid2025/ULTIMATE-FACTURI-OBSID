"""
Matching Engine - Logica de potrivire MT940 + Borderouri + Facturi
Implementeaza logica din grupare facturi_obsid.py pentru versiunea web

Include algoritm inteligent de matching pentru:
- Gasirea combinatiei de colete care se potriveste cu suma OP-ului
- Identificarea coletelor "pending" care vor fi pe urmatorul borderou
"""

import re
from typing import List, Dict, Optional, Tuple, Set
from datetime import datetime, date, timedelta
from decimal import Decimal, ROUND_HALF_UP
from itertools import combinations

from .supabase_client import get_supabase_client
from .oblio_api import get_all_invoices, _get_access_token, _get_headers, OBLIO_API_BASE, OBLIO_CIF
from .smart_matching import find_parcel_combination, match_gls_parcels_to_bank_transactions, match_sameday_parcels_to_bank_transactions
import requests


def search_invoice_by_amount_in_oblio(
    amount: float,
    used_invoices: set = None,
    tolerance: float = 0.01
) -> Optional[Dict]:
    """
    Cauta factura in Oblio API pe baza sumei.

    Args:
        amount: Suma de cautat
        used_invoices: Set de facturi deja folosite (pentru a evita duplicatele)
        tolerance: Toleranta pentru comparatie (default 0.01)

    Returns:
        Dict cu informatii despre factura gasita sau None
    """
    if used_invoices is None:
        used_invoices = set()

    try:
        # Obtine token-ul de acces
        _get_access_token()

        # Cauta facturi in Oblio (ultimele 90 de zile)
        from datetime import timedelta
        end_date = date.today()
        start_date = end_date - timedelta(days=90)

        invoices = get_all_invoices(issued_after=start_date, issued_before=end_date)

        for invoice in invoices:
            total = float(invoice.get('total', 0))
            invoice_number = invoice.get('number', '')

            # Verifica daca suma se potriveste
            if abs(total - amount) < tolerance:
                # Extrage numarul numeric din factura
                match_numeric = re.search(r'\d+', str(invoice_number))
                if match_numeric:
                    numeric_part = match_numeric.group()

                    # Verifica daca factura nu a fost deja folosita
                    if numeric_part not in used_invoices:
                        return {
                            'invoice_number': numeric_part,
                            'full_number': invoice_number,
                            'total': total,
                            'issue_date': invoice.get('issueDate'),
                            'client_name': invoice.get('client', {}).get('name', ''),
                            'oblio_id': invoice.get('id')
                        }

        return None

    except Exception as e:
        print(f"Eroare la cautarea in Oblio API: {e}")
        return None


def search_invoice_by_number_in_oblio(invoice_number: str) -> Optional[Dict]:
    """
    Cauta factura in Oblio API pe baza numarului de factura.

    Args:
        invoice_number: Numarul facturii de cautat

    Returns:
        Dict cu informatii despre factura gasita sau None
    """
    try:
        _get_access_token()

        from datetime import timedelta
        end_date = date.today()
        start_date = end_date - timedelta(days=365)  # Cauta in ultimul an

        invoices = get_all_invoices(issued_after=start_date, issued_before=end_date)

        search_number = str(invoice_number).strip()

        for invoice in invoices:
            full_number = str(invoice.get('number', '')).strip()

            # Extrage partea numerica
            match_numeric = re.search(r'\d+', full_number)
            if match_numeric:
                numeric_part = match_numeric.group()

                # Verifica potrivirea exacta sau daca numarul e continut
                if numeric_part == search_number or search_number in full_number:
                    return {
                        'invoice_number': numeric_part,
                        'full_number': full_number,
                        'total': float(invoice.get('total', 0)),
                        'issue_date': invoice.get('issueDate'),
                        'client_name': invoice.get('client', {}).get('name', ''),
                        'oblio_id': invoice.get('id')
                    }

        return None

    except Exception as e:
        print(f"Eroare la cautarea facturii {invoice_number} in Oblio: {e}")
        return None


def match_transactions_with_invoices(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> Dict:
    """
    Potriveste tranzactiile bancare cu facturile din Oblio.

    Workflow:
    1. Ia tranzactiile bancare din Supabase (importate din MT940)
    2. Ia facturile din Supabase (sincronizate din Oblio)
    3. Potriveste dupa suma si sursa
    4. Pentru cele nepotrivite, cauta in Oblio API

    Returns:
        Dict cu rezultatele matchingului
    """
    supabase = get_supabase_client()

    # 1. Ia tranzactiile bancare
    trans_query = supabase.table('bank_transactions').select('*')
    if start_date:
        trans_query = trans_query.gte('transaction_date', start_date.isoformat())
    if end_date:
        trans_query = trans_query.lte('transaction_date', end_date.isoformat())

    transactions = trans_query.order('transaction_date').execute().data

    # 2. Ia facturile
    inv_query = supabase.table('invoices').select('*')
    if start_date:
        inv_query = inv_query.gte('issue_date', start_date.isoformat())
    if end_date:
        inv_query = inv_query.lte('issue_date', end_date.isoformat())

    invoices = inv_query.execute().data

    # Creeaza index pentru facturi dupa suma
    invoices_by_amount = {}
    for inv in invoices:
        total = float(inv.get('total', 0))
        if total not in invoices_by_amount:
            invoices_by_amount[total] = []
        invoices_by_amount[total].append(inv)

    # 3. Potriveste
    matched = []
    unmatched_transactions = []
    used_invoices = set()

    for trans in transactions:
        amount = float(trans.get('amount', 0))
        source = trans.get('source', '')

        # Cauta factura cu suma potrivita
        found_invoice = None

        # Cauta in facturile din DB
        for amt, inv_list in invoices_by_amount.items():
            if abs(amt - amount) < 0.01:
                for inv in inv_list:
                    inv_id = inv.get('oblio_id')
                    if inv_id not in used_invoices:
                        found_invoice = inv
                        used_invoices.add(inv_id)
                        break
                if found_invoice:
                    break

        if found_invoice:
            matched.append({
                'transaction': trans,
                'invoice': found_invoice,
                'match_type': 'exact_amount'
            })
        else:
            # Cauta in Oblio API
            oblio_result = search_invoice_by_amount_in_oblio(
                amount,
                used_invoices
            )

            if oblio_result:
                used_invoices.add(oblio_result['invoice_number'])
                matched.append({
                    'transaction': trans,
                    'invoice': oblio_result,
                    'match_type': 'oblio_api'
                })
            else:
                unmatched_transactions.append(trans)

    return {
        'matched': matched,
        'unmatched': unmatched_transactions,
        'total_transactions': len(transactions),
        'total_invoices': len(invoices),
        'match_rate': len(matched) / len(transactions) * 100 if transactions else 0
    }


def generate_opuri_report_data(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> List[Dict]:
    """
    Genereaza datele pentru raportul OP-uri.

    Combina:
    - Tranzactii bancare (din MT940/Supabase)
    - Facturi (din Oblio/Supabase)
    - Informatii de matching

    Returns:
        Lista de dict-uri cu datele raportului
    """
    supabase = get_supabase_client()

    # Ia tranzactiile bancare
    trans_query = supabase.table('bank_transactions').select('*')
    if start_date:
        trans_query = trans_query.gte('transaction_date', start_date.isoformat())
    if end_date:
        trans_query = trans_query.lte('transaction_date', end_date.isoformat())

    transactions = trans_query.order('transaction_date', desc=True).execute().data

    # Ia facturile pentru referinta
    inv_query = supabase.table('invoices').select('*')
    invoices_response = inv_query.execute()

    # Creeaza index facturi dupa suma pentru matching rapid
    invoices_by_amount = {}
    for inv in invoices_response.data:
        total = round(float(inv.get('total', 0)), 2)
        if total not in invoices_by_amount:
            invoices_by_amount[total] = []
        invoices_by_amount[total].append(inv)

    # Grupeaza tranzactiile dupa sursa si data
    report_data = []
    used_invoices = set()

    for trans in transactions:
        amount = round(float(trans.get('amount', 0)), 2)
        source = trans.get('source', 'Altul')

        # Cauta factura potrivita
        matched_invoice = None
        invoice_number = ''
        order_id = ''
        has_error = False

        # Incearca matching dupa suma exacta
        if amount in invoices_by_amount:
            for inv in invoices_by_amount[amount]:
                inv_id = inv.get('oblio_id')
                if inv_id and inv_id not in used_invoices:
                    matched_invoice = inv
                    used_invoices.add(inv_id)
                    invoice_number = inv.get('invoice_number', '')
                    break

        # Daca nu s-a gasit, marcheaza eroare
        if not matched_invoice:
            has_error = True
            # Incearca cautare in Oblio API
            oblio_result = search_invoice_by_amount_in_oblio(amount, used_invoices)
            if oblio_result:
                invoice_number = oblio_result.get('invoice_number', '')
                used_invoices.add(invoice_number)
                has_error = False

        report_data.append({
            'data_op': trans.get('transaction_date', ''),
            'numar_op': trans.get('op_reference', ''),
            'nume_borderou': trans.get('file_name', ''),
            'curier': source,
            'order_id': order_id,
            'numar_factura': invoice_number,
            'suma': amount,
            'erori': 'DA' if has_error else 'NU',
            'diferenta_emag': '',
            'facturi_comision_emag': ''
        })

    return report_data


def export_opuri_to_excel(
    report_data: List[Dict],
    output_path: str
) -> bool:
    """
    Exporta raportul OP-uri in format Excel.

    Args:
        report_data: Lista de date pentru raport
        output_path: Calea fisierului Excel de output

    Returns:
        True daca exportul a reusit, False altfel
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "OP-uri"

        # Defineste culorile
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        blue_fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")
        orange_fill = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")
        green_fill = PatternFill(start_color="FF90EE90", end_color="FF90EE90", fill_type="solid")
        netopia_fill = PatternFill(start_color="FFDAEEF3", end_color="FFDAEEF3", fill_type="solid")
        header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")

        # Header
        headers = [
            "Data OP", "Numar OP", "Nume Borderou", "Curier", "Order ID",
            "Numar Factura", "Suma", "Erori", "Diferenta eMag", "Facturi Comision eMag"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')

        # Date
        current_row = 2
        prev_source = None

        for data in report_data:
            source = data.get('curier', '')

            # Scrie datele
            ws.cell(row=current_row, column=1, value=data.get('data_op', ''))
            ws.cell(row=current_row, column=2, value=data.get('numar_op', ''))
            ws.cell(row=current_row, column=3, value=data.get('nume_borderou', ''))
            ws.cell(row=current_row, column=4, value=source)
            ws.cell(row=current_row, column=5, value=data.get('order_id', ''))
            ws.cell(row=current_row, column=6, value=data.get('numar_factura', ''))
            ws.cell(row=current_row, column=7, value=data.get('suma', 0))
            ws.cell(row=current_row, column=8, value=data.get('erori', 'NU'))
            ws.cell(row=current_row, column=9, value=data.get('diferenta_emag', ''))
            ws.cell(row=current_row, column=10, value=data.get('facturi_comision_emag', ''))

            # Coloreaza celula curierului
            curier_cell = ws.cell(row=current_row, column=4)
            if source == 'GLS':
                curier_cell.fill = blue_fill
                curier_cell.font = Font(color="FFFFFF")
            elif source == 'Sameday':
                curier_cell.fill = red_fill
                curier_cell.font = Font(color="FFFFFF")
            elif source == 'Netopia':
                curier_cell.fill = netopia_fill
            elif source == 'eMag':
                curier_cell.fill = orange_fill

            # Coloreaza erori
            if data.get('erori') == 'DA':
                ws.cell(row=current_row, column=8).fill = red_fill
                ws.cell(row=current_row, column=8).font = Font(color="FFFFFF")

            current_row += 1

        # Ajusteaza latimea coloanelor
        column_widths = [12, 20, 25, 12, 15, 15, 12, 8, 15, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        # Salveaza
        wb.save(output_path)
        return True

    except Exception as e:
        print(f"Eroare la export Excel: {e}")
        return False


def get_matching_statistics(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> Dict:
    """
    Calculeaza statistici despre matching-ul tranzactii-facturi.

    Returns:
        Dict cu statistici
    """
    supabase = get_supabase_client()

    # Numar tranzactii
    trans_query = supabase.table('bank_transactions').select('id, amount, source')
    if start_date:
        trans_query = trans_query.gte('transaction_date', start_date.isoformat())
    if end_date:
        trans_query = trans_query.lte('transaction_date', end_date.isoformat())

    transactions = trans_query.execute().data

    # Numar facturi
    inv_query = supabase.table('invoices').select('id, total')
    if start_date:
        inv_query = inv_query.gte('issue_date', start_date.isoformat())
    if end_date:
        inv_query = inv_query.lte('issue_date', end_date.isoformat())

    invoices = inv_query.execute().data

    # Calculeaza statistici pe surse
    by_source = {}
    for trans in transactions:
        source = trans.get('source', 'Altul')
        if source not in by_source:
            by_source[source] = {'count': 0, 'amount': 0}
        by_source[source]['count'] += 1
        by_source[source]['amount'] += float(trans.get('amount', 0))

    total_trans_amount = sum(float(t.get('amount', 0)) for t in transactions)
    total_inv_amount = sum(float(i.get('total', 0)) for i in invoices)

    return {
        'total_transactions': len(transactions),
        'total_invoices': len(invoices),
        'transactions_amount': total_trans_amount,
        'invoices_amount': total_inv_amount,
        'difference': total_trans_amount - total_inv_amount,
        'by_source': by_source
    }


def generate_smart_opuri_report(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> Dict:
    """
    Genereaza raportul OP-uri folosind algoritmul inteligent de matching.

    Acest algoritm:
    1. Ia toate tranzactiile bancare (OP-uri)
    2. Pentru fiecare OP, gaseste combinatia exacta de colete care se potriveste
    3. Coletele ramase sunt marcate ca "pending" pentru urmatorul OP
    4. Identifica coletele care nu au OP inca (vor veni in viitor)

    Returns:
        Dict cu rezultatele matching-ului si statistici
    """
    # Ruleaza smart matching pentru GLS si Sameday
    gls_results = match_gls_parcels_to_bank_transactions(start_date, end_date)
    sameday_results = match_sameday_parcels_to_bank_transactions(start_date, end_date)

    return {
        'gls': gls_results,
        'sameday': sameday_results,
        'summary': {
            'gls_matched': gls_results.get('matched_parcels', 0),
            'gls_pending': gls_results.get('pending_parcels', 0),
            'gls_pending_amount': gls_results.get('pending_total', 0),
            'sameday_matched': sameday_results.get('matched_parcels', 0),
            'sameday_pending': sameday_results.get('pending_parcels', 0),
            'sameday_pending_amount': sameday_results.get('pending_total', 0),
        }
    }


def get_pending_parcels_summary() -> Dict:
    """
    Returneaza un sumar al coletelor care nu au fost inca potrivite cu OP-uri.

    Aceste colete vor fi pe urmatoarele borderouri bancare.

    Returns:
        Dict cu coletele pending grupate pe sursa
    """
    smart_results = generate_smart_opuri_report()

    pending_summary = {
        'gls': {
            'count': smart_results['gls'].get('pending_parcels', 0),
            'total': smart_results['gls'].get('pending_total', 0),
            'parcels': smart_results['gls'].get('pending', [])
        },
        'sameday': {
            'count': smart_results['sameday'].get('pending_parcels', 0),
            'total': smart_results['sameday'].get('pending_total', 0),
            'parcels': smart_results['sameday'].get('pending', [])
        }
    }

    pending_summary['total_count'] = pending_summary['gls']['count'] + pending_summary['sameday']['count']
    pending_summary['total_amount'] = pending_summary['gls']['total'] + pending_summary['sameday']['total']

    return pending_summary


def analyze_parcel_discrepancy(
    source: str,
    op_reference: str,
    expected_amount: float
) -> Dict:
    """
    Analizeaza discrepanta pentru un OP specific.

    Util cand suma coletelor din API nu se potriveste cu suma din OP.
    Gaseste care colete fac parte din OP si care vor fi pe alt borderou.

    Args:
        source: 'GLS' sau 'Sameday'
        op_reference: Referinta OP-ului
        expected_amount: Suma asteptata (din OP)

    Returns:
        Dict cu analiza discrepantei
    """
    supabase = get_supabase_client()

    # Gaseste tranzactia bancara
    trans = supabase.table('bank_transactions') \
        .select('*') \
        .eq('op_reference', op_reference) \
        .execute().data

    if not trans:
        return {'error': f'OP-ul {op_reference} nu a fost gasit'}

    trans = trans[0]
    op_date = trans.get('transaction_date', '')
    op_amount = float(trans.get('amount', 0))

    # Calculeaza intervalul de livrare (1-5 zile inainte de OP)
    op_date_obj = datetime.strptime(op_date, '%Y-%m-%d').date()
    delivery_start = op_date_obj - timedelta(days=5)

    # Ia coletele din perioada
    if source == 'GLS':
        table = 'gls_parcels'
        id_field = 'parcel_number'
    else:
        table = 'sameday_parcels'
        id_field = 'awb_number'

    parcels = supabase.table(table).select('*') \
        .eq('is_delivered', True) \
        .gte('delivery_date', delivery_start.isoformat()) \
        .lte('delivery_date', op_date) \
        .execute().data

    if not parcels:
        return {
            'op_reference': op_reference,
            'op_amount': op_amount,
            'error': 'Nu s-au gasit colete pentru aceasta perioada'
        }

    total_parcels_sum = sum(float(p.get('cod_amount', 0)) for p in parcels)

    # Foloseste algoritmul de matching
    matched, remaining = find_parcel_combination(parcels, op_amount)

    result = {
        'op_reference': op_reference,
        'op_date': op_date,
        'op_amount': op_amount,
        'total_parcels_in_period': len(parcels),
        'total_parcels_sum': round(total_parcels_sum, 2),
        'difference': round(total_parcels_sum - op_amount, 2),
        'matched_parcels': len(matched),
        'matched_sum': round(sum(float(p.get('cod_amount', 0)) for p in matched), 2),
        'pending_parcels': len(remaining),
        'pending_sum': round(sum(float(p.get('cod_amount', 0)) for p in remaining), 2),
    }

    if remaining:
        result['pending_details'] = [
            {
                id_field: p.get(id_field),
                'cod_amount': float(p.get('cod_amount', 0)),
                'recipient_name': p.get('recipient_name', ''),
                'delivery_date': p.get('delivery_date'),
                'note': 'Va fi pe alt borderou'
            }
            for p in remaining
        ]

    return result
