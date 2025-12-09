"""
Procesor pentru generarea raportului OP-uri
Combina datele din Supabase (GLS, Sameday, Netopia) cu Gomag si MT940
pentru a genera export-ul in formatul original
"""

import pandas as pd
from typing import List, Dict, Tuple, Optional
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .supabase_client import get_supabase_client as get_client


def get_gls_parcels_for_period(start_date: str, end_date: str) -> List[Dict]:
    """
    Obtine coletele GLS livrate intr-o perioada.
    """
    client = get_client()
    if not client:
        return []

    try:
        result = client.table("gls_parcels").select("*").gte(
            "delivery_date", start_date
        ).lte(
            "delivery_date", end_date
        ).eq("is_delivered", True).execute()

        return result.data or []
    except Exception as e:
        print(f"Eroare la citirea GLS: {e}")
        return []


def get_sameday_parcels_for_period(start_date: str, end_date: str) -> List[Dict]:
    """
    Obtine coletele Sameday livrate intr-o perioada.
    """
    client = get_client()
    if not client:
        return []

    try:
        result = client.table("sameday_parcels").select("*").gte(
            "delivery_date", start_date
        ).lte(
            "delivery_date", end_date
        ).eq("is_delivered", True).execute()

        return result.data or []
    except Exception as e:
        print(f"Eroare la citirea Sameday: {e}")
        return []


def get_mt940_transactions_for_period(start_date: str, end_date: str) -> List[Dict]:
    """
    Obtine tranzactiile MT940 (OP-uri bancare) pentru o perioada.
    """
    client = get_client()
    if not client:
        return []

    try:
        result = client.table("bank_transactions").select("*").gte(
            "transaction_date", start_date
        ).lte(
            "transaction_date", end_date
        ).execute()

        return result.data or []
    except Exception as e:
        print(f"Eroare la citirea MT940: {e}")
        return []


def get_gls_borderouri_for_period(start_date: str, end_date: str) -> List[Dict]:
    """
    Obtine borderourile GLS din email pentru o perioada.
    Acestea contin gruparea reala a coletelor asa cum le trimite GLS.

    IMPORTANT: Include si borderouri din primele 7 zile ale lunii urmatoare,
    deoarece coletele livrate la sfarsitul lunii pot aparea pe borderou
    in luna urmatoare (ex: livrat 27.11, borderou 02.12).
    """
    client = get_client()
    if not client:
        return []

    try:
        # Extinde end_date cu 7 zile pentru a prinde borderouri care vin mai tarziu
        from datetime import datetime, timedelta
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
        extended_end = (end_date_obj + timedelta(days=7)).strftime('%Y-%m-%d')

        result = client.table("gls_borderouri").select("*").gte(
            "borderou_date", start_date
        ).lte(
            "borderou_date", extended_end
        ).order("borderou_date").execute()

        borderouri = result.data or []

        # Pentru fiecare borderou, obtine coletele asociate si verifica daca au livrari in perioada
        filtered_borderouri = []
        for borderou in borderouri:
            parcels_result = client.table("gls_borderou_parcels").select("*").eq(
                "borderou_id", borderou['id']
            ).execute()
            borderou['parcels'] = parcels_result.data or []

            # Verifica daca borderoul e in perioada originala SAU daca are colete livrate in perioada
            borderou_date = str(borderou.get('borderou_date', ''))
            if borderou_date >= start_date and borderou_date <= end_date:
                # Borderou in perioada - include direct
                filtered_borderouri.append(borderou)
            else:
                # Borderou in afara perioadei - include doar daca are colete cu delivery_date in perioada
                # Verifica in gls_parcels delivery_date pentru aceste colete
                parcel_numbers = [p.get('parcel_number') for p in borderou['parcels']]
                if parcel_numbers:
                    # Cauta delivery_date pentru aceste colete
                    parcels_with_dates = client.table("gls_parcels").select("parcel_number, delivery_date").in_(
                        "parcel_number", parcel_numbers
                    ).execute().data or []

                    # Verifica daca vreun colet a fost livrat in perioada selectata
                    has_delivery_in_period = False
                    for p in parcels_with_dates:
                        delivery_date = str(p.get('delivery_date', ''))
                        if delivery_date and delivery_date >= start_date and delivery_date <= end_date:
                            has_delivery_in_period = True
                            break

                    if has_delivery_in_period:
                        filtered_borderouri.append(borderou)

        return filtered_borderouri
    except Exception as e:
        print(f"Eroare la citirea GLS borderouri: {e}")
        return []


def get_gls_parcels_not_in_borderouri(start_date: str, end_date: str, borderou_parcels: List[str]) -> List[Dict]:
    """
    Obtine coletele GLS care nu sunt incluse in niciun borderou.
    Acestea sunt colete livrate dar neprimite inca in borderoul de ramburs.
    """
    client = get_client()
    if not client:
        return []

    try:
        # Obtine toate coletele GLS livrate in perioada
        result = client.table("gls_parcels").select("*").gte(
            "delivery_date", start_date
        ).lte(
            "delivery_date", end_date
        ).eq("is_delivered", True).execute()

        all_parcels = result.data or []

        # Filtreaza coletele care nu sunt in borderouri
        # borderou_parcels contine numerele de colet din borderouri
        pending_parcels = []
        for p in all_parcels:
            parcel_num = str(p.get('parcel_number', ''))
            # Verifica daca coletul NU este in niciun borderou
            if parcel_num not in borderou_parcels:
                pending_parcels.append(p)

        return pending_parcels
    except Exception as e:
        print(f"Eroare la citirea GLS parcels pending: {e}")
        return []


def group_parcels_by_delivery_date(parcels: List[Dict], curier: str) -> List[Dict]:
    """
    Grupeaza coletele pe data livrarii (simuleaza borderourile).
    """
    if not parcels:
        return []

    by_date = {}
    for p in parcels:
        date_str = p.get('delivery_date', 'Unknown')
        if date_str not in by_date:
            by_date[date_str] = []
        by_date[date_str].append(p)

    result = []
    for date_str, date_parcels in sorted(by_date.items()):
        suma_total = sum(float(p.get('cod_amount', 0) or 0) for p in date_parcels)

        if curier == "GLS":
            try:
                dt = datetime.strptime(date_str, '%Y-%m-%d')
                borderou_name = f"553005982_RON_{dt.strftime('%d%m%Y')}.xlsx"
            except:
                borderou_name = f"GLS_{date_str}.xlsx"
        else:
            try:
                borderou_name = f"{date_str}-obsid-s r l-cod-ledger-ron.xlsx"
            except:
                borderou_name = f"Sameday_{date_str}.xlsx"

        result.append({
            'borderou': borderou_name,
            'curier': curier,
            'delivery_date': date_str,
            'parcels': date_parcels,
            'suma_total': suma_total
        })

    return result


def match_awb_with_gomag(awb: str, gomag_df: pd.DataFrame, curier: str) -> Tuple[str, str]:
    """
    Cauta AWB-ul in Gomag si returneaza (Order ID, Numar Factura).

    Pentru GLS: AWB-ul din borderou poate avea 3 caractere extra la final.
    Pentru Sameday: AWB-ul poate avea 3 caractere la final (001, 002, etc.)
    """
    if gomag_df is None or gomag_df.empty:
        return '', ''

    # Normalizeaza AWB
    awb_norm = str(awb).strip().replace(' ', '').lstrip('0')

    # Pregateste Gomag pentru cautare
    gomag_df = gomag_df.copy()
    gomag_df.columns = gomag_df.columns.str.strip().str.lower()

    if 'awb' not in gomag_df.columns:
        return '', ''

    gomag_df['awb_norm'] = gomag_df['awb'].astype(str).str.strip().str.replace(' ', '').str.lstrip('0')

    # Cauta potrivire directa
    match = gomag_df[gomag_df['awb_norm'] == awb_norm]

    # Daca nu gaseste, incearca fara ultimele 3 caractere (pentru GLS)
    if match.empty and len(awb_norm) > 10:
        awb_short = awb_norm[:-3]
        match = gomag_df[gomag_df['awb_norm'] == awb_short]

    # Daca nu gaseste, incearca cu ultimele 3 caractere adaugate
    if match.empty:
        match = gomag_df[gomag_df['awb_norm'].str.startswith(awb_norm)]

    if not match.empty:
        row = match.iloc[0]
        order_id = str(row.get('numar comanda', row.get('id', ''))).strip()
        numar_factura = str(row.get('numar factura', '')).strip()

        # Curata numarul facturii
        if numar_factura and numar_factura != 'nan':
            try:
                numar_factura = str(int(float(numar_factura)))
            except:
                pass
        else:
            numar_factura = ''

        return order_id, numar_factura

    return '', ''


def match_op_with_borderou(suma_borderou: float, mt940_transactions: List[Dict], curier: str) -> Tuple[str, str]:
    """
    Gaseste OP-ul care se potriveste cu suma borderoului si curierul.
    """
    for trans in mt940_transactions:
        suma_op = float(trans.get('amount', 0) or 0)
        source = trans.get('source', '')

        # Verifica daca sursa se potriveste
        if curier.upper() not in source.upper():
            continue

        # Potrivire cu toleranta de 0.10 RON
        if abs(suma_op - suma_borderou) < 0.10:
            return trans.get('op_reference', ''), trans.get('transaction_date', '')

    return '', ''


def generate_opuri_export(
    start_date: str,
    end_date: str,
    gomag_df: Optional[pd.DataFrame] = None
) -> BytesIO:
    """
    Genereaza export-ul OP-uri in formatul original.

    IMPORTANT: Pentru GLS foloseste borderourile REALE din email (gls_borderouri table),
    nu gruparile simulate pe data livrarii. Borderourile GLS din email contin gruparea
    exacta a coletelor asa cum apar in desfasuratorul de ramburs.
    """
    # Obtine datele din Supabase
    sameday_parcels = get_sameday_parcels_for_period(start_date, end_date)

    # Extinde perioada pentru MT940 cu 7 zile pentru a gasi OP-uri care vin mai tarziu
    # (ex: borderou din 28.11 poate avea OP pe 02.12)
    from datetime import datetime, timedelta
    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
    extended_end = (end_date_obj + timedelta(days=7)).strftime('%Y-%m-%d')
    mt940_transactions = get_mt940_transactions_for_period(start_date, extended_end)

    # Obtine borderourile GLS REALE din email (nu simulate)
    gls_borderouri_real = get_gls_borderouri_for_period(start_date, end_date)

    # Formateaza borderourile GLS pentru procesare
    gls_borderouri = []
    borderou_parcel_numbers = set()  # Track parcels already in borderouri

    for borderou in gls_borderouri_real:
        parcels = borderou.get('parcels', [])
        suma_total = float(borderou.get('total_amount', 0))

        # Track parcel numbers
        for p in parcels:
            borderou_parcel_numbers.add(str(p.get('parcel_number', '')))

        # Formateaza borderou pentru procesare
        gls_borderouri.append({
            'borderou': borderou.get('file_name', f"GLS_{borderou['borderou_date']}.xlsx"),
            'curier': 'GLS',
            'delivery_date': borderou.get('borderou_date', ''),
            'parcels': parcels,
            'suma_total': suma_total,
            'op_reference': borderou.get('op_reference', ''),
            'op_date': borderou.get('op_date', ''),
            'op_matched': borderou.get('op_matched', False)
        })

    # Daca nu sunt borderouri din email, foloseste metoda veche (fallback)
    if not gls_borderouri:
        gls_parcels = get_gls_parcels_for_period(start_date, end_date)
        gls_borderouri = group_parcels_by_delivery_date(gls_parcels, "GLS")

    # Pentru Sameday folosim metoda veche (grupeaza pe data livrarii)
    sameday_borderouri = group_parcels_by_delivery_date(sameday_parcels, "Sameday")

    # Obtine facturile din Oblio
    client = get_client()
    invoices = []
    if client:
        try:
            result = client.table("invoices").select("*").execute()
            invoices = result.data or []
        except:
            pass

    # Creeaza workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "OP-uri"

    # Stiluri
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    gls_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    sameday_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    netopia_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    error_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Header
    headers = ["Data OP", "Număr OP", "Nume Borderou", "Curier", "Order ID",
               "Număr Factură", "Sumă", "Erori", "Diferență eMag", "Facturi Comision eMag"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Ajusteaza latimea coloanelor
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 25

    row_num = 2

    # Proceseaza GLS si Sameday
    for borderouri, curier, fill_color in [
        (gls_borderouri, "GLS", gls_fill),
        (sameday_borderouri, "Sameday", sameday_fill)
    ]:
        for borderou in borderouri:
            parcels = borderou['parcels']
            suma_total = borderou['suma_total']

            # Pentru GLS, foloseste OP-ul pre-potrivit daca exista
            if curier == "GLS" and borderou.get('op_matched'):
                numar_op = borderou.get('op_reference', '')
                data_op = borderou.get('op_date', '')
            else:
                # Gaseste OP-ul prin matching cu MT940
                numar_op, data_op = match_op_with_borderou(suma_total, mt940_transactions, curier)

            # Proceseaza fiecare colet
            facturi_ok = []
            facturi_ko = []

            for p in parcels:
                # Extrage AWB
                awb = p.get('parcel_number', p.get('awb_number', ''))
                cod_amount = float(p.get('cod_amount', 0) or 0)

                # Cauta in Gomag
                order_id, numar_factura = match_awb_with_gomag(awb, gomag_df, curier)

                if numar_factura:
                    facturi_ok.append({
                        'awb': awb,
                        'order_id': order_id,
                        'numar_factura': numar_factura,
                        'suma': cod_amount
                    })
                else:
                    facturi_ko.append({
                        'awb': awb,
                        'order_id': order_id,
                        'suma': cod_amount
                    })

            # Sorteaza facturile dupa numar
            facturi_ok.sort(key=lambda x: int(x['numar_factura']) if x['numar_factura'].isdigit() else 0)

            erori_exist = len(facturi_ko) > 0
            erori_text = "DA" if erori_exist else "NU"

            # Scrie facturile OK
            first_row = True
            for f in facturi_ok:
                ws.cell(row=row_num, column=1, value=data_op if first_row else "")
                ws.cell(row=row_num, column=2, value=numar_op if first_row else "")
                ws.cell(row=row_num, column=3, value=borderou['borderou'] if first_row else "")

                cell_curier = ws.cell(row=row_num, column=4, value=curier if first_row else "")
                if first_row:
                    cell_curier.fill = fill_color
                    cell_curier.font = Font(color="FFFFFF")

                ws.cell(row=row_num, column=5, value=f['order_id'])
                ws.cell(row=row_num, column=6, value=int(f['numar_factura']) if f['numar_factura'].isdigit() else f['numar_factura'])
                ws.cell(row=row_num, column=7, value=f['suma'])

                cell_erori = ws.cell(row=row_num, column=8, value=erori_text if first_row else "")
                if first_row and erori_exist:
                    cell_erori.fill = error_fill

                first_row = False
                row_num += 1

            # Daca nu sunt facturi OK, scrie doar header-ul
            if not facturi_ok:
                ws.cell(row=row_num, column=1, value=data_op)
                ws.cell(row=row_num, column=2, value=numar_op)
                ws.cell(row=row_num, column=3, value=borderou['borderou'])
                cell_curier = ws.cell(row=row_num, column=4, value=curier)
                cell_curier.fill = fill_color
                cell_curier.font = Font(color="FFFFFF")
                cell_erori = ws.cell(row=row_num, column=8, value=erori_text)
                if erori_exist:
                    cell_erori.fill = error_fill
                row_num += 1

            # Scrie AWB-urile fara factura
            if facturi_ko:
                ws.cell(row=row_num, column=6, value="AWB-uri fără factură:")
                row_num += 1

                for f in facturi_ko:
                    ws.cell(row=row_num, column=6, value=f['awb'])
                    ws.cell(row=row_num, column=7, value=f['suma'])
                    row_num += 1

            # Rand total
            ws.cell(row=row_num, column=6, value="Total")
            ws.cell(row=row_num, column=6).font = Font(bold=True)
            ws.cell(row=row_num, column=7, value=suma_total)
            ws.cell(row=row_num, column=7).font = Font(bold=True)
            row_num += 1

            # Rand gol
            row_num += 1

    # ============================================
    # Proceseaza Netopia
    # ============================================
    netopia_ops = [t for t in mt940_transactions if t.get('source', '').upper() == 'NETOPIA']

    for netopia_op in netopia_ops:
        batch_id = netopia_op.get('batch_id', '')
        suma_op = float(netopia_op.get('amount', 0) or 0)
        numar_op = netopia_op.get('op_reference', '')
        data_op = netopia_op.get('transaction_date', '')

        # Numele borderoului pentru Netopia
        borderou_name = f"batchId.{batch_id}.csv" if batch_id else f"Netopia_{data_op}.csv"

        # Obtine tranzactiile Netopia din Supabase pentru acest batch
        netopia_trans = []
        if client and batch_id:
            try:
                result = client.table("netopia_transactions").select("*").eq("batch_id", batch_id).execute()
                netopia_trans = result.data or []
            except:
                pass

        # Calculeaza comisioane si total facturi
        total_facturi = sum(float(t.get('amount', 0) or 0) for t in netopia_trans)
        total_comisioane = sum(float(t.get('fee', 0) or 0) for t in netopia_trans)

        if not netopia_trans:
            total_facturi = suma_op
            total_comisioane = 0

        # Scrie header-ul borderoului Netopia
        ws.cell(row=row_num, column=1, value=data_op)
        ws.cell(row=row_num, column=2, value=numar_op)
        ws.cell(row=row_num, column=3, value=borderou_name)
        cell_curier = ws.cell(row=row_num, column=4, value="Netopia")
        cell_curier.fill = netopia_fill

        first_row = True
        if netopia_trans:
            for trans in netopia_trans:
                order_id_raw = trans.get('order_id', '')
                amount = float(trans.get('amount', 0) or 0)

                # Extrage doar numarul din order_id (ex: "Comanda nr. 569 - www.obsid.ro" -> "569")
                order_id = order_id_raw
                if 'Comanda nr.' in str(order_id_raw):
                    import re
                    match = re.search(r'Comanda nr\.\s*(\d+)', str(order_id_raw))
                    if match:
                        order_id = match.group(1)
                elif 'Bank transfer' in str(order_id_raw):
                    # Skip bank transfer entries (sunt transferuri interne)
                    continue

                # Daca amount e 0, foloseste net_amount (suma - comision)
                if amount == 0:
                    amount = float(trans.get('net_amount', 0) or 0)
                    # Daca tot e 0, incearca sa calculam din fee (comisionul e negativ)
                    if amount == 0:
                        fee = float(trans.get('fee', 0) or 0)
                        if fee < 0:
                            amount = abs(fee)  # Folosim valoarea absoluta a comisionului ca aproximare

                # Cauta factura
                numar_factura = ''
                if gomag_df is not None and not gomag_df.empty and order_id:
                    gomag_temp = gomag_df.copy()
                    gomag_temp.columns = gomag_temp.columns.str.strip().str.lower()
                    match = gomag_temp[gomag_temp['numar comanda'].astype(str) == str(order_id)]
                    if not match.empty:
                        nf = match.iloc[0].get('numar factura', '')
                        if nf and str(nf) != 'nan':
                            try:
                                numar_factura = str(int(float(nf)))
                            except:
                                numar_factura = str(nf)

                if not first_row:
                    row_num += 1
                    ws.cell(row=row_num, column=1, value="")
                    ws.cell(row=row_num, column=2, value="")
                    ws.cell(row=row_num, column=3, value="")
                    ws.cell(row=row_num, column=4, value="")

                ws.cell(row=row_num, column=5, value=order_id)
                ws.cell(row=row_num, column=6, value=numar_factura)
                ws.cell(row=row_num, column=7, value=amount)
                ws.cell(row=row_num, column=8, value="NU" if numar_factura else "DA")

                first_row = False

        row_num += 1

        # Linie Comisioane
        ws.cell(row=row_num, column=6, value="Comisioane:")
        ws.cell(row=row_num, column=7, value=round(total_comisioane, 2))
        row_num += 1

        # Linie Total facturi
        ws.cell(row=row_num, column=6, value="Total facturi:")
        ws.cell(row=row_num, column=7, value=round(total_facturi, 2))
        row_num += 1

        # Linie Total OP
        ws.cell(row=row_num, column=6, value="Total OP:")
        ws.cell(row=row_num, column=6).font = Font(bold=True)
        ws.cell(row=row_num, column=7, value=round(suma_op, 2))
        ws.cell(row=row_num, column=7).font = Font(bold=True)
        row_num += 1

        # Rand gol
        row_num += 1

    # Salveaza in buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
