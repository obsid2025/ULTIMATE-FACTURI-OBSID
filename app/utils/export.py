"""
Export functionalitate pentru generarea fisierului Excel final
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Optional
from io import BytesIO
from datetime import datetime


def genereaza_export_excel(
    rezultate_gls: List[Dict],
    rezultate_sameday: List[Dict],
    rezultate_netopia: List[Dict],
    incasari_mt940: List[tuple]
) -> BytesIO:
    """
    Genereaza fisierul Excel cu toate facturile grupate pe OP-uri.

    Returns:
        BytesIO: Buffer cu fisierul Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturi Grupate"

    # Stiluri
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    gls_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    sameday_fill = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid")
    netopia_fill = PatternFill(start_color="06b6d4", end_color="06b6d4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header
    headers = ["Data OP", "Numar OP", "Borderou", "Curier", "AWB/Order ID", "Nr. Factura", "Suma", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Ajusteaza latimea coloanelor
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10

    row_num = 2

    # Creeaza dictionar cu incasari MT940 pentru potrivire
    incasari_dict = {}
    for op_ref, suma, data, batchid, details in incasari_mt940:
        details_upper = details.upper()
        if "GLS" in details_upper or "GENERAL LOGISTICS" in details_upper:
            sursa = "GLS"
        elif "DELIVERY SOLUTIONS" in details_upper:
            sursa = "Sameday"
        elif "NETOPIA" in details_upper or "BATCHID" in details_upper:
            sursa = "Netopia"
        else:
            sursa = "Altul"

        key = (sursa, round(suma, 2))
        incasari_dict[key] = (op_ref, data, batchid)

    # Proceseaza GLS
    for rezultat in rezultate_gls:
        borderou = rezultat['borderou']
        suma_total = rezultat['suma_total']
        potrivite = rezultat['potrivite']

        # Cauta OP-ul potrivit
        key = ("GLS", round(suma_total, 2))
        op_info = incasari_dict.get(key, ("", "", None))
        op_ref, data_op, _ = op_info

        first_row = True
        for _, row in potrivite.iterrows():
            awb = row.get('AWB_normalizat', '')
            factura = row.get('numar factura', '')
            suma = row.get('suma', 0)

            ws.cell(row=row_num, column=1, value=data_op if first_row else "")
            ws.cell(row=row_num, column=2, value=op_ref if first_row else "")
            ws.cell(row=row_num, column=3, value=borderou if first_row else "")
            cell_curier = ws.cell(row=row_num, column=4, value="GLS" if first_row else "")
            cell_curier.fill = gls_fill
            ws.cell(row=row_num, column=5, value=awb)
            ws.cell(row=row_num, column=6, value=factura if pd.notna(factura) else "")
            ws.cell(row=row_num, column=7, value=suma)
            ws.cell(row=row_num, column=8, value="OK" if pd.notna(factura) and factura else "LIPSA")

            first_row = False
            row_num += 1

        # Rand total
        ws.cell(row=row_num, column=6, value="TOTAL:")
        ws.cell(row=row_num, column=6).font = Font(bold=True)
        ws.cell(row=row_num, column=7, value=suma_total)
        ws.cell(row=row_num, column=7).font = Font(bold=True)
        row_num += 2

    # Proceseaza Sameday
    for rezultat in rezultate_sameday:
        borderou = rezultat['borderou']
        suma_total = rezultat['suma_total']
        potrivite = rezultat['potrivite']

        # Cauta OP-ul potrivit
        key = ("Sameday", round(suma_total, 2))
        op_info = incasari_dict.get(key, ("", "", None))
        op_ref, data_op, _ = op_info

        first_row = True
        for _, row in potrivite.iterrows():
            awb = row.get('AWB_normalizat', row.get('AWB', ''))
            factura = row.get('numar factura', '')
            suma = row.get('suma', 0)

            ws.cell(row=row_num, column=1, value=data_op if first_row else "")
            ws.cell(row=row_num, column=2, value=op_ref if first_row else "")
            ws.cell(row=row_num, column=3, value=borderou if first_row else "")
            cell_curier = ws.cell(row=row_num, column=4, value="Sameday" if first_row else "")
            cell_curier.fill = sameday_fill
            ws.cell(row=row_num, column=5, value=awb)
            ws.cell(row=row_num, column=6, value=factura if pd.notna(factura) else "")
            ws.cell(row=row_num, column=7, value=suma)
            ws.cell(row=row_num, column=8, value="OK" if pd.notna(factura) and factura else "LIPSA")

            first_row = False
            row_num += 1

        # Rand total
        ws.cell(row=row_num, column=6, value="TOTAL:")
        ws.cell(row=row_num, column=6).font = Font(bold=True)
        ws.cell(row=row_num, column=7, value=suma_total)
        ws.cell(row=row_num, column=7).font = Font(bold=True)
        row_num += 2

    # Proceseaza Netopia
    for rezultat in rezultate_netopia:
        borderou = rezultat['borderou']
        suma_total = rezultat['suma_total']
        batchid = rezultat.get('batchid')
        potrivite = rezultat['potrivite']

        # Cauta OP-ul potrivit prin batchid sau suma
        op_ref, data_op = "", ""
        for (op, suma, data, bid, details) in incasari_mt940:
            if batchid and bid == batchid:
                op_ref, data_op = op, data
                break
            elif "NETOPIA" in details.upper() and abs(suma - suma_total) < 1:
                op_ref, data_op = op, data
                break

        first_row = True
        for _, row in potrivite.iterrows():
            order_id = row.get('numar_comanda_norm', row.get('numar comanda', ''))
            factura = row.get('numar factura', '')
            suma = row.get('suma', 0)

            ws.cell(row=row_num, column=1, value=data_op if first_row else "")
            ws.cell(row=row_num, column=2, value=op_ref if first_row else "")
            ws.cell(row=row_num, column=3, value=borderou if first_row else "")
            cell_curier = ws.cell(row=row_num, column=4, value="Netopia" if first_row else "")
            cell_curier.fill = netopia_fill
            ws.cell(row=row_num, column=5, value=order_id)
            ws.cell(row=row_num, column=6, value=factura if pd.notna(factura) else "")
            ws.cell(row=row_num, column=7, value=suma)
            ws.cell(row=row_num, column=8, value="OK" if pd.notna(factura) and factura else "LIPSA")

            first_row = False
            row_num += 1

        # Rand total
        ws.cell(row=row_num, column=6, value="TOTAL:")
        ws.cell(row=row_num, column=6).font = Font(bold=True)
        ws.cell(row=row_num, column=7, value=suma_total)
        ws.cell(row=row_num, column=7).font = Font(bold=True)
        row_num += 2

    # Salveaza in buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
