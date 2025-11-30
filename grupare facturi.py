# File: grupare facturi.py
import os
import re
import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import warnings
import xml.etree.ElementTree as ET # Adăugăm importul pentru parsarea XML
import threading

# Încarcă xlrd pentru fișiere .xls (Excel vechi)
try:
    import xlrd
    print("xlrd disponibil pentru fisiere .xls")
except ImportError:
    print("xlrd nu este disponibil - fisierele .xls nu vor putea fi citite")

CONFIG_FILE = "config.txt"

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class FacturiApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Ultimate FACTURI - Panou de control")
        self.geometry("950x700")
        self.resizable(True, True)
        self.configure(bg='#f0f0f0')

        # Variabile pentru căi
        self.folder_gls = tk.StringVar()
        self.folder_sameday = tk.StringVar()
        self.path_gomag = tk.StringVar()
        self.path_extras = tk.StringVar()
        self.folder_netopia = tk.StringVar() # Modificat din path_netopia în folder_netopia
        self.folder_emag = tk.StringVar() # Adăugat pentru borderouri eMag
        self.path_easysales = tk.StringVar() # Adăugat pentru fișierul easySales
        self.path_oblio = tk.StringVar() # Adăugat pentru fișierul Oblio
        self.path_export = tk.StringVar() # Adăugat pentru calea de export

        # Credențiale API eMag pentru citirea facturilor
        self.emag_api_username = tk.StringVar()
        self.emag_api_password = tk.StringVar()

        self.erori = []
        
        # Variabile pentru progres
        self.progress_var = tk.DoubleVar()
        self.progress_text = tk.StringVar()
        self.progress_text.set("Gata pentru export...")
        
        # Setează calea de export implicită
        default_export_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "opuri_export.xlsx")
        self.path_export.set(default_export_path)

        # Tabs
        self.tab_control = ttk.Notebook(self)
        self.tab_main = ttk.Frame(self.tab_control)
        self.tab_erori = ttk.Frame(self.tab_control)
        self.tab_control.add(self.tab_main, text='Control')
        self.tab_control.add(self.tab_erori, text='Erori')
        self.tab_control.pack(expand=1, fill='both')

        # Panou principal
        self._build_main_tab()
        self._build_erori_tab()

        # Încarcă căile salvate (mutat aici)
        self._load_paths()
        # Actualizează starea butonului după încărcarea căilor
        self._update_export_state()

    def _build_main_tab(self):
        frm = self.tab_main
        # Nu putem seta bg pe ttk.Frame, doar pe tk.Frame
        
        # Main container cu padding
        main_container = tk.Frame(frm, bg='#f0f0f0')
        main_container.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Title
        title_label = tk.Label(main_container, text="Ultimate FACTURI", 
                              font=('Arial', 16, 'bold'), 
                              bg='#f0f0f0', fg='#2c3e50')
        title_label.pack(pady=(0, 20))
        
        # Canvas cu scroll pentru a putea face interfața mai mare
        canvas = tk.Canvas(main_container, bg='#f0f0f0', highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f0f0f0')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # --- GRUPA 1: BORDEROURI CURIER ---
        group1 = ttk.LabelFrame(scrollable_frame, text="📦 Borderouri Curier", padding=15)
        group1.pack(fill='x', pady=(0, 15))
        
        # GLS
        self._create_folder_row(group1, 0, "Folder borderouri GLS:", self.folder_gls, 
                               self.select_folder_gls, "📁 Selectează folder GLS...")
        
        # Sameday
        self._create_folder_row(group1, 1, "Folder borderouri Sameday:", self.folder_sameday, 
                               self.select_folder_sameday, "📁 Selectează folder Sameday...")
        
        # --- GRUPA 2: FIȘIERE PRINCIPALE ---
        group2 = ttk.LabelFrame(scrollable_frame, text="📄 Fișiere Principale", padding=15)
        group2.pack(fill='x', pady=(0, 15))
        
        # Gomag
        self._create_file_row(group2, 0, "Fișier Gomag (XLSX):", self.path_gomag, 
                             self.select_gomag, "📄 Selectează fișier Gomag...", [("Excel files", "*.xlsx")])
        
        # Extras bancar
        self._create_file_row(group2, 1, "Extras bancar (XML/TXT):", self.path_extras, 
                             self.select_extras, "📄 Selectează extras bancar...", 
                             [("XML files", "*.xml"), ("Text files", "*.txt"), ("All files", "*.*")])
        
        # --- GRUPA 3: PLĂȚI ONLINE ---
        group3 = ttk.LabelFrame(scrollable_frame, text="💳 Plăți Online", padding=15)
        group3.pack(fill='x', pady=(0, 15))
        
        # Netopia
        self._create_folder_row(group3, 0, "Folder Netopia (CSV-uri):", self.folder_netopia, 
                               self.select_netopia, "📁 Selectează folder Netopia...")
        
        # eMag
        self._create_folder_row(group3, 1, "Folder eMag (XLSX-uri):", self.folder_emag, 
                               self.select_emag, "📁 Selectează folder eMag...")
        
        # easySales
        self._create_file_row(group3, 2, "Fișier easySales (XLSX):", self.path_easysales, 
                             self.select_easysales, "📄 Selectează fișier easySales...", [("Excel files", "*.xlsx")])
        
        # Oblio
        self._create_file_row(group3, 3, "Fișier Oblio (XLS/XLSX):", self.path_oblio, 
                             self.select_oblio, "📄 Selectează fișier Oblio...", [("Excel files", "*.xlsx"), ("Excel files", "*.xls")])
        
        # --- GRUPA 4: EXPORT ---
        group4 = ttk.LabelFrame(scrollable_frame, text="💾 Export", padding=15)
        group4.pack(fill='x', pady=(0, 15))
        
        # Calea de export
        self._create_file_row(group4, 0, "Salvează în:", self.path_export, 
                             self.select_export_path, "💾 Selectează locația export...", [("Excel files", "*.xlsx")])
        
        # Progress bar
        progress_frame = tk.Frame(group4, bg='#f0f0f0')
        progress_frame.grid(row=1, column=0, columnspan=3, sticky='ew', pady=(10, 5))
        group4.grid_columnconfigure(0, weight=1)
        
        tk.Label(progress_frame, text="Progres:", font=('Arial', 9, 'bold'), bg='#f0f0f0').pack(anchor='w')
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, 
                                          maximum=100, length=400, mode='determinate')
        self.progress_bar.pack(fill='x', pady=(2, 5))
        
        self.progress_label = tk.Label(progress_frame, textvariable=self.progress_text, 
                                      font=('Arial', 8), fg='#7f8c8d', bg='#f0f0f0')
        self.progress_label.pack(anchor='w')
        
        # --- GRUPA 5: ACȚIUNI ---
        group5 = tk.Frame(scrollable_frame, bg='#f0f0f0')
        group5.pack(fill='x', pady=(10, 0))
        
        # Butoane cu design frumos
        button_frame = tk.Frame(group5, bg='#f0f0f0')
        button_frame.pack(expand=True)
        
        self.btn_export = tk.Button(button_frame, text="🚀 EXPORT", 
                                   command=self.export_threaded,
                                   font=('Arial', 12, 'bold'),
                                   bg='#27ae60', fg='white',
                                   relief='flat', padx=30, pady=10,
                                   cursor='hand2',
                                   state='disabled')
        self.btn_export.pack(side='left', padx=(0, 10))
        
        btn_close = tk.Button(button_frame, text="❌ ÎNCHIDE", 
                             command=self.quit,
                             font=('Arial', 12, 'bold'),
                             bg='#e74c3c', fg='white',
                             relief='flat', padx=30, pady=10,
                             cursor='hand2')
        btn_close.pack(side='left')
        
        # Pack canvas și scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Actualizează starea butonului Export când se schimbă ceva
        for var in [self.folder_gls, self.folder_sameday, self.path_gomag, self.path_extras, 
                   self.folder_netopia, self.folder_emag, self.path_easysales, self.path_oblio, self.path_export]:
            var.trace_add('write', self._update_export_state)

        # Salvează căile la închiderea aplicației
        self.protocol("WM_DELETE_WINDOW", self._on_closing)
    
    def _create_folder_row(self, parent, row, label_text, var, command, button_text):
        """Creează un rând pentru selectarea unui folder"""
        tk.Label(parent, text=label_text, font=('Arial', 9, 'bold')).grid(
            row=row, column=0, sticky="w", padx=(0, 10), pady=5)
        
        entry = tk.Entry(parent, textvariable=var, width=60, 
                        font=('Arial', 9), relief='solid', bd=1)
        entry.grid(row=row, column=1, sticky='ew', padx=(0, 10), pady=5)
        
        btn = tk.Button(parent, text=button_text, command=command,
                       font=('Arial', 8), bg='#3498db', fg='white',
                       relief='flat', padx=10, pady=5, cursor='hand2')
        btn.grid(row=row, column=2, padx=(0, 0), pady=5)
        
        parent.grid_columnconfigure(1, weight=1)
    
    def _create_file_row(self, parent, row, label_text, var, command, button_text, filetypes):
        """Creează un rând pentru selectarea unui fișier"""
        tk.Label(parent, text=label_text, font=('Arial', 9, 'bold')).grid(
            row=row, column=0, sticky="w", padx=(0, 10), pady=5)
        
        entry = tk.Entry(parent, textvariable=var, width=60, 
                        font=('Arial', 9), relief='solid', bd=1)
        entry.grid(row=row, column=1, sticky='ew', padx=(0, 10), pady=5)
        
        btn = tk.Button(parent, text=button_text, 
                       command=lambda: command(filetypes),
                       font=('Arial', 8), bg='#3498db', fg='white',
                       relief='flat', padx=10, pady=5, cursor='hand2')
        btn.grid(row=row, column=2, padx=(0, 0), pady=5)
        
        parent.grid_columnconfigure(1, weight=1)

    def _build_erori_tab(self):
        self.txt_erori = tk.Text(self.tab_erori, wrap='word', state='disabled', bg="#f8d7da")
        self.txt_erori.pack(expand=1, fill='both', padx=10, pady=10)

    def select_folder_gls(self):
        path = filedialog.askdirectory(title="Selectează folderul cu borderouri GLS")
        if path:
            self.folder_gls.set(path)

    def select_folder_sameday(self):
        path = filedialog.askdirectory(title="Selectează folderul cu borderouri Sameday")
        if path:
            self.folder_sameday.set(path)

    def select_gomag(self, filetypes):
        path = filedialog.askopenfilename(title="Selectează fișierul Gomag (XLSX)", filetypes=filetypes)
        if path:
            self.path_gomag.set(path)

    def select_extras(self, filetypes):
        path = filedialog.askopenfilename(title="Selectează extrasul bancar", filetypes=filetypes)
        if path:
            self.path_extras.set(path)

    def select_netopia(self):
        path = filedialog.askdirectory(title="Selectează folderul cu fișiere Netopia (CSV)")
        if path:
            self.folder_netopia.set(path)

    def select_emag(self):
        path = filedialog.askdirectory(title="Selectează folderul cu fișiere eMag (XLSX)")
        if path:
            self.folder_emag.set(path)

    def select_easysales(self, filetypes):
        path = filedialog.askopenfilename(title="Selectează fișierul easySales (XLSX)", filetypes=filetypes)
        if path:
            self.path_easysales.set(path)

    def select_oblio(self, filetypes):
        path = filedialog.askopenfilename(title="Selectează fișierul Oblio (XLS/XLSX)", filetypes=filetypes)
        if path:
            self.path_oblio.set(path)
    
    def select_export_path(self, filetypes):
        path = filedialog.asksaveasfilename(
            title="Selectează locația pentru export", 
            filetypes=filetypes,
            defaultextension=".xlsx",
            initialfile="opuri_export.xlsx"
        )
        if path:
            self.path_export.set(path)

    def _update_export_state(self, *args):
        if all([self.folder_gls.get(), self.folder_sameday.get(), self.path_gomag.get(), self.path_extras.get(), self.folder_netopia.get(), self.folder_emag.get(), self.path_easysales.get(), self.path_oblio.get(), self.path_export.get()]):
            self.btn_export.config(state='normal', bg='#27ae60')
        else:
            self.btn_export.config(state='disabled', bg='#95a5a6')
    
    def export_threaded(self):
        """Rulează exportul într-un thread separat pentru a nu bloca interfața"""
        self.btn_export.config(state='disabled', text='⏳ PROCESEAZĂ...', bg='#f39c12')
        self.progress_var.set(0)
        self.progress_text.set("Începe procesarea...")
        
        # Rulează exportul într-un thread separat
        thread = threading.Thread(target=self.export)
        thread.daemon = True
        thread.start()

    def export(self):
        self.erori.clear()
        self._show_erori("")  # Golește tab-ul de erori

        try:
            # Actualizează progresul
            self.progress_var.set(10)
            self.progress_text.set("Procesează borderouri GLS...")
            
            rezultate_gls, erori_gls = self.proceseaza_borderouri(self.folder_gls.get(), self.path_gomag.get(), self.path_extras.get(), "GLS")
            self.erori.extend(erori_gls)
            
            self.progress_var.set(25)
            self.progress_text.set("Procesează borderouri Sameday...")
            
            rezultate_sameday, erori_sameday = self.proceseaza_borderouri(self.folder_sameday.get(), self.path_gomag.get(), self.path_extras.get(), "Sameday")
            self.erori.extend(erori_sameday)

            # CHECK FINAL: Caută în Oblio pentru AWB-uri rămase cu erori
            self.progress_var.set(40)
            self.progress_text.set("Verificare finală în Oblio...")
            self._cautare_finala_oblio(rezultate_gls, rezultate_sameday)

            self.progress_var.set(50)
            self.progress_text.set("Procesează Netopia...")
            
            # Procesează Netopia
            print("*" * 60)
            print("ÎNAINTE DE PROCESAREA NETOPIA")
            print("*" * 60)
            tranzactii_netopia, erori_netopia = self.proceseaza_netopia(self.folder_netopia.get(), self.path_gomag.get())
            print(f"DEBUG MAIN: Netopia processing completed. Results: {len(tranzactii_netopia)}, Errors: {len(erori_netopia)}")
            if tranzactii_netopia:
                print(f"DEBUG MAIN: First Netopia result: {tranzactii_netopia[0]}")
            else:
                print("DEBUG MAIN: NO Netopia results returned!")
            print("*" * 60)
            print("DUPĂ PROCESAREA NETOPIA")
            print("*" * 60)
            self.erori.extend(erori_netopia)

            self.progress_var.set(75)
            self.progress_text.set("Procesează eMag...")
            
            # Procesează eMag
            rezultate_emag, erori_emag = self.proceseaza_emag(self.folder_emag.get(), self.path_easysales.get())
            print(f"DEBUG MAIN: eMag processing completed. Results: {len(rezultate_emag)}, Errors: {len(erori_emag)}")
            if rezultate_emag:
                print(f"DEBUG MAIN: First eMag result: {rezultate_emag[0]}")
            else:
                print("DEBUG MAIN: NO eMag results returned!")
            self.erori.extend(erori_emag)

            self.progress_var.set(90)
            self.progress_text.set("Generează fișierul Excel...")
            
            # Exportă OP-urile chiar dacă există erori! - folosește calea selectată de utilizator
            print(f"DEBUG: Apelează export_opuri cu rezultate_emag: {len(rezultate_emag) if rezultate_emag else 'None'}")
            self.export_opuri(rezultate_gls, rezultate_sameday, tranzactii_netopia, rezultate_emag, self.path_export.get(), self.folder_netopia.get())
            
            self.progress_var.set(100)
            self.progress_text.set("Export finalizat cu succes!")
            
        except Exception as e:
            self.erori.append(f"Eroare generală la export: {e}")
            self.progress_text.set(f"Eroare: {e}")

        # Resetează butonul export
        self.btn_export.config(state='normal', text='🚀 EXPORT', bg='#27ae60')
        
        if self.erori:
            self._show_erori("\n".join(self.erori))
            messagebox.showwarning("Exportat cu erori", f"Exportul a fost realizat, dar au apărut erori. Verifică tab-ul Erori.\nFișierul a fost salvat în: {self.path_export.get()}")
        else:
            messagebox.showinfo("Export reușit", f"Exportul a avut succes!\nFișierul a fost salvat în: {self.path_export.get()}")
            self._save_paths() # Salvează căile după un export reușit

    def _show_erori(self, text):
        self.txt_erori.config(state='normal')
        self.txt_erori.delete(1.0, tk.END)
        self.txt_erori.insert(tk.END, text)
        self.txt_erori.config(state='disabled')

    def _on_closing(self):
        self._save_paths()
        self.destroy()

    def _save_paths(self):
        paths = {
            "folder_gls": self.folder_gls.get(),
            "folder_sameday": self.folder_sameday.get(),
            "path_gomag": self.path_gomag.get(),
            "path_extras": self.path_extras.get(),
            "folder_netopia": self.folder_netopia.get(),
            "folder_emag": self.folder_emag.get(),
            "path_easysales": self.path_easysales.get(),
            "path_oblio": self.path_oblio.get(),
            "path_export": self.path_export.get(),
            "emag_api_username": self.emag_api_username.get(),
            "emag_api_password": self.emag_api_password.get()
        }
        try:
            with open(CONFIG_FILE, "w") as f:
                for key, value in paths.items():
                    f.write(f"{key}={value}\n")
        except Exception as e:
            print(f"Eroare la salvarea căilor: {e}")

    def _load_paths(self):
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, "r") as f:
                    for line in f:
                        line = line.strip()
                        if "=" in line:
                            key, value = line.split("=", 1)
                            if key == "folder_gls":
                                self.folder_gls.set(value)
                            elif key == "folder_sameday":
                                self.folder_sameday.set(value)
                            elif key == "path_gomag":
                                self.path_gomag.set(value)
                            elif key == "path_extras":
                                self.path_extras.set(value)
                            elif key == "folder_netopia":
                                self.folder_netopia.set(value)
                            elif key == "folder_emag":
                                self.folder_emag.set(value)
                            elif key == "path_easysales":
                                self.path_easysales.set(value)
                            elif key == "path_oblio":
                                self.path_oblio.set(value)
                            elif key == "path_export":
                                self.path_export.set(value)
                            elif key == "emag_api_username":
                                self.emag_api_username.set(value)
                            elif key == "emag_api_password":
                                self.emag_api_password.set(value)
        except Exception as e:
            print(f"Eroare la încărcarea căilor: {e}")

    def _cautare_finala_oblio(self, rezultate_gls, rezultate_sameday):
        """
        Căutare FINALĂ în Oblio pentru AWB-urile rămase cu erori.
        Se execută după procesarea tuturor borderourilor GLS și Sameday.
        Caută facturi care NU au fost deja folosite.
        """
        print("\n" + "="*80)
        print("CĂUTARE FINALĂ ÎN OBLIO PENTRU AWB-URI CU ERORI")
        print("="*80)

        path_oblio = self.path_oblio.get()
        if not path_oblio or not os.path.exists(path_oblio):
            print("Oblio: Fișier nu există, sărim căutarea finală")
            return

        # Colectează toate facturile deja folosite
        facturi_folosite = set()
        for rezultate in [rezultate_gls, rezultate_sameday]:
            for rez in rezultate:
                potrivite = rez.get('potrivite', pd.DataFrame())
                for idx, row in potrivite.iterrows():
                    numar_factura = row.get('numar factura', None)
                    if numar_factura and not pd.isna(numar_factura) and numar_factura != 0:
                        try:
                            facturi_folosite.add(str(int(float(numar_factura))))
                        except (ValueError, TypeError):
                            facturi_folosite.add(str(numar_factura))

        print(f"Facturi deja folosite: {len(facturi_folosite)}")

        # Încarcă Oblio
        try:
            # Pentru fișiere .xls (Excel vechi), folosește engine='xlrd'
            # Dacă fișierul este corupt, încearcă cu calamine (pentru fișiere Excel vechi corupte)
            try:
                if path_oblio.endswith('.xls'):
                    # Încearcă cu xlrd mai întâi
                    try:
                        oblio_df = pd.read_excel(path_oblio, header=4, engine='xlrd')
                        print("Oblio: Fișier .xls citit cu xlrd")
                    except:
                        # Dacă xlrd eșuează, încearcă cu calamine (python-calamine)
                        try:
                            oblio_df = pd.read_excel(path_oblio, header=4, engine='calamine')
                            print("Oblio: Fișier .xls citit cu calamine")
                        except:
                            # Ultimate fallback - convertește cu win32com dacă e pe Windows
                            import tempfile
                            temp_xlsx = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name

                            # Încearcă conversia cu win32com
                            try:
                                import win32com.client
                                excel = win32com.client.Dispatch("Excel.Application")
                                excel.Visible = False
                                wb = excel.Workbooks.Open(os.path.abspath(path_oblio))
                                wb.SaveAs(temp_xlsx, FileFormat=51)  # 51 = xlsx
                                wb.Close()
                                excel.Quit()

                                oblio_df = pd.read_excel(temp_xlsx, header=4)
                                os.unlink(temp_xlsx)
                                print("Oblio: Fișier .xls convertit cu Excel COM și citit")
                            except:
                                raise Exception("Nu s-a putut citi fișierul Oblio corupt")
                else:
                    oblio_df = pd.read_excel(path_oblio, header=4)
                    print("Oblio: Fișier .xlsx citit cu openpyxl")
            except Exception as e_engine:
                print(f"Oblio: Eroare cu engine-ul specific: {e_engine}")
                raise

            if 'Total valoare' not in oblio_df.columns or 'Factura' not in oblio_df.columns:
                print("Oblio: Coloane lipsă, sărim căutarea finală")
                return

            # Pregătește datele Oblio
            oblio_df['Total_valoare_numeric'] = pd.to_numeric(oblio_df['Total valoare'], errors='coerce')
            oblio_df['Factura_clean'] = oblio_df['Factura'].astype(str).str.strip()
            oblio_valide = oblio_df.dropna(subset=['Total_valoare_numeric'])
            oblio_valide = oblio_valide[oblio_valide['Factura_clean'] != 'nan']
            oblio_valide = oblio_valide[oblio_valide['Factura_clean'] != '']

            print(f"Oblio: {len(oblio_valide)} facturi valide în total")

        except Exception as e:
            print(f"Oblio: Eroare la citire: {e}")
            return

        # Caută AWB-uri cu erori și încearcă să le rezolvi
        awb_rezolvate = 0
        for rezultate, tip in [(rezultate_gls, "GLS"), (rezultate_sameday, "Sameday")]:
            for rez in rezultate:
                potrivite = rez.get('potrivite', pd.DataFrame())
                borderou_name = rez.get('borderou', 'N/A')

                # Găsește AWB-uri fără factură
                awb_fara_factura_mask = (potrivite['numar factura'].isna()) | (potrivite['numar factura'] == 0)
                awb_fara_factura_indices = potrivite[awb_fara_factura_mask].index

                if len(awb_fara_factura_indices) == 0:
                    continue

                print(f"\n{tip} - {borderou_name}: {len(awb_fara_factura_indices)} AWB-uri fără factură")

                for idx in awb_fara_factura_indices:
                    row = potrivite.loc[idx]
                    awb = row.get('AWB_normalizat', 'N/A')
                    suma = row.get('Sumă ramburs') or row.get('Suma ramburs')

                    if not suma:
                        print(f"  ✗ AWB {awb} - SĂRIT (fără sumă)")
                        continue

                    try:
                        suma_float = float(suma)
                    except (ValueError, TypeError):
                        print(f"  ✗ AWB {awb} - SĂRIT (suma '{suma}' nu e numerică)")
                        continue

                    print(f"  ? Caută AWB {awb} cu suma {suma_float}...")

                    # Caută în Oblio facturi care SE POTRIVESC după sumă și NU sunt folosite
                    potriviri = oblio_valide[abs(oblio_valide['Total_valoare_numeric'] - suma_float) < 0.01]

                    print(f"    → {len(potriviri)} potriviri după sumă în Oblio")

                    factura_gasita = False
                    for _, match_row in potriviri.iterrows():
                        factura_completa = match_row['Factura_clean']
                        suma_oblio = match_row['Total_valoare_numeric']

                        # Extrage numărul facturii
                        import re
                        match_numeric = re.search(r'\d+', factura_completa)
                        if not match_numeric:
                            print(f"    ✗ Factură '{factura_completa}' - nu se poate extrage număr")
                            continue

                        numar_factura = match_numeric.group()

                        # Verifică dacă factura NU a fost folosită
                        if numar_factura in facturi_folosite:
                            print(f"    ✗ Factură {numar_factura} (suma {suma_oblio}) - DEJA FOLOSITĂ")
                        else:
                            # GĂSIT! Asociază factura cu AWB-ul
                            potrivite.at[idx, 'numar factura'] = numar_factura
                            facturi_folosite.add(numar_factura)
                            awb_rezolvate += 1
                            print(f"  ✓ AWB {awb} - REZOLVAT cu Factură {numar_factura} (suma {suma})")
                            factura_gasita = True
                            break

                    if not factura_gasita and len(potriviri) == 0:
                        print(f"  ✗ AWB {awb} - NU GĂSIT în Oblio (suma {suma_float} nu există)")
                    elif not factura_gasita:
                        print(f"  ✗ AWB {awb} - NU REZOLVAT (toate facturile cu suma {suma_float} sunt deja folosite)")

        print(f"\n{'='*80}")
        print(f"CĂUTARE FINALĂ COMPLETĂ: {awb_rezolvate} AWB-uri rezolvate")
        print(f"{'='*80}\n")

    def _cauta_in_oblio(self, suma_cautata, nume_client, data_livrare, tip_curier, awb):
        """
        Caută factură în fișierul Oblio pe baza sumei
        Mapare: Coloana P ("Total valoare") -> Coloana D ("Factura")
        Header pe rândul 5, datele încep de pe rândul 6
        """
        import re
        print(f"{tip_curier} Oblio: Caută AWB {awb} - Suma: {suma_cautata}")
        
        path_oblio = self.path_oblio.get()
        print(f"🔍 DEBUG: Calea fișierului Oblio selectat: '{path_oblio}'")
        
        if not path_oblio or not os.path.exists(path_oblio):
            print(f"{tip_curier} Oblio: Fișier nu există la {path_oblio}")
            return None
            
        try:
            # Header pe rândul 5 (index 4), datele încep de pe rândul 6
            # Pentru fișiere .xls (Excel vechi), folosește engine='xlrd' sau convertește cu Excel COM
            try:
                if path_oblio.endswith('.xls'):
                    # Încearcă cu xlrd mai întâi
                    try:
                        oblio_df = pd.read_excel(path_oblio, header=4, engine='xlrd')
                        print(f"{tip_curier} Oblio: Fișier .xls citit cu xlrd")
                    except:
                        # Dacă xlrd eșuează, încearcă cu calamine (python-calamine)
                        try:
                            oblio_df = pd.read_excel(path_oblio, header=4, engine='calamine')
                            print(f"{tip_curier} Oblio: Fișier .xls citit cu calamine")
                        except:
                            # Ultimate fallback - convertește cu win32com dacă e pe Windows
                            import tempfile
                            temp_xlsx = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name

                            # Încearcă conversia cu win32com
                            import win32com.client
                            excel = win32com.client.Dispatch("Excel.Application")
                            excel.Visible = False
                            wb = excel.Workbooks.Open(os.path.abspath(path_oblio))
                            wb.SaveAs(temp_xlsx, FileFormat=51)  # 51 = xlsx
                            wb.Close()
                            excel.Quit()

                            oblio_df = pd.read_excel(temp_xlsx, header=4)
                            os.unlink(temp_xlsx)
                            print(f"{tip_curier} Oblio: Fișier .xls convertit cu Excel COM și citit")
                else:
                    oblio_df = pd.read_excel(path_oblio, header=4)
                    print(f"{tip_curier} Oblio: Fișier .xlsx citit cu openpyxl")
            except Exception as e_engine:
                print(f"{tip_curier} Oblio: Eroare cu engine-ul specific: {e_engine}")
                return None
                
            print(f"🔍 DEBUG COMPLET pentru fișierul Oblio:")
            print(f"   Fișier: {path_oblio}")
            print(f"   Shape: {oblio_df.shape}")
            print(f"   Coloane: {list(oblio_df.columns)}")
            
            # DEBUG: Afișează TOATE rândurile pentru a vedea structura completă
            print(f"🔍 DEBUG: TOATE rândurile din fișier (primele 50):")
            for i, (idx, row) in enumerate(oblio_df.head(50).iterrows()):
                if i < 30:  # Limitează la 30 pentru a nu fi prea lung
                    print(f"  Rând {i+6}: {dict(row)}")
            
            # Verifică coloanele necesare
            if 'Total valoare' not in oblio_df.columns:
                print(f"🔍 DEBUG: Nu există coloana 'Total valoare'!")
                print(f"🔍 DEBUG: Coloane disponibile: {list(oblio_df.columns)}")
                
                # Încearcă să găsească coloane similare
                possible_cols = [col for col in oblio_df.columns if 'total' in col.lower() or 'valoare' in col.lower()]
                print(f"🔍 DEBUG: Coloane posibile pentru Total valoare: {possible_cols}")
                return None
                
            if 'Factura' not in oblio_df.columns:
                print(f"🔍 DEBUG: Nu există coloana 'Factura'!")
                print(f"🔍 DEBUG: Coloane disponibile: {list(oblio_df.columns)}")
                
                # Încearcă să găsească coloane similare
                possible_cols = [col for col in oblio_df.columns if 'factura' in col.lower() or 'nr' in col.lower()]
                print(f"🔍 DEBUG: Coloane posibile pentru Factura: {possible_cols}")
                return None
            
            # Convertește suma la numeric
            oblio_df['Total_valoare_numeric'] = pd.to_numeric(oblio_df['Total valoare'], errors='coerce')
            oblio_df['Factura_clean'] = oblio_df['Factura'].astype(str).str.strip()
            
            # Elimină rândurile cu valori invalide
            oblio_valide = oblio_df.dropna(subset=['Total_valoare_numeric'])
            oblio_valide = oblio_valide[oblio_valide['Factura_clean'] != 'nan']
            oblio_valide = oblio_valide[oblio_valide['Factura_clean'] != '']
            
            print(f"{tip_curier} Oblio: {len(oblio_valide)} rânduri valide găsite")
            
            # DEBUG SPECIAL pentru suma 86.8
            suma_cautata_float = float(suma_cautata)
            if abs(suma_cautata_float - 86.8) < 0.1:
                print(f"🔍 DEBUG SPECIAL pentru suma ~86.8:")
                print(f"   Suma căutată exact: {suma_cautata_float}")
                print(f"   Primele 20 de valori din coloana 'Total valoare':")
                for i, (idx, row) in enumerate(oblio_valide.head(20).iterrows()):
                    val = row['Total_valoare_numeric']
                    factura = row['Factura_clean']
                    print(f"     Rând {i+6}: {val} -> {factura}")
                    if abs(val - 86.8) < 0.02:
                        print(f"     ⭐ POSIBILĂ POTRIVIRE: {val} (diferență: {abs(val - 86.8)})")
            
            # Caută suma exactă (toleranță 0.01)
            potriviri = oblio_valide[abs(oblio_valide['Total_valoare_numeric'] - suma_cautata_float) < 0.01]
            
            print(f"{tip_curier} Oblio: {len(potriviri)} potriviri pentru suma {suma_cautata}")
            
            # DEBUG SUPLIMENTAR pentru suma 86.8
            if abs(suma_cautata_float - 86.8) < 0.1:
                print(f"🔍 DEBUG CĂUTARE pentru suma ~86.8:")
                print(f"   Toleranța folosită: ±0.01")
                for i, (idx, row) in enumerate(oblio_valide.iterrows()):
                    val = row['Total_valoare_numeric']
                    diff = abs(val - suma_cautata_float)
                    if diff < 0.1:  # Arată toate valorile apropiate
                        match_status = "✓ MATCH" if diff < 0.01 else "✗ nu match"
                        print(f"     Valoare {val}: diferență={diff:.4f} -> {match_status}")
            
            if potriviri.empty:
                print(f"{tip_curier} Oblio: ✗ Nu s-a găsit suma {suma_cautata}")
                return None
            
            # Ia prima potrivire
            first_match = potriviri.iloc[0]
            factura_completa = first_match['Factura_clean']
            
            # Extrage doar partea numerică din factură (ex: NRTTF233054 -> 233054)
            import re
            match_numeric = re.search(r'\d+', factura_completa)
            if match_numeric:
                numar_factura = match_numeric.group()
                print(f"{tip_curier} Oblio: ✓ GĂSIT! Suma {suma_cautata} -> Factură {factura_completa} -> Număr {numar_factura}")
                return numar_factura
            else:
                print(f"{tip_curier} Oblio: ✗ Nu s-a putut extrage numărul din factură {factura_completa}")
                return None
                
        except Exception as e:
            print(f"{tip_curier} Oblio: Eroare: {e}")
            return None
    
    def _cauta_factura_in_oblio_by_number(self, numar_factura_cautat):
        """
        Caută factură în fișierul Oblio pe baza numărului de factură
        Returnează numărul de factură dacă este găsit, None altfel
        """
        print(f"eMag Oblio: Caută numărul facturii {numar_factura_cautat}")
        
        path_oblio = self.path_oblio.get()
        if not path_oblio or not os.path.exists(path_oblio):
            print(f"eMag Oblio: Fișier nu există la {path_oblio}")
            return None
            
        try:
            # Citește fișierul Oblio
            if path_oblio.endswith('.xls'):
                oblio_df = pd.read_excel(path_oblio, header=4, engine='xlrd')
            else:
                oblio_df = pd.read_excel(path_oblio, header=4)
                
            if 'Factura' not in oblio_df.columns:
                print(f"eMag Oblio: Nu există coloana 'Factura'")
                return None
            
            # Curăță și normalizează datele
            oblio_df['Factura_clean'] = oblio_df['Factura'].astype(str).str.strip()
            
            # Elimină rândurile invalide
            oblio_valide = oblio_df[oblio_df['Factura_clean'] != 'nan']
            oblio_valide = oblio_valide[oblio_valide['Factura_clean'] != '']
            
            # Extrage părțile numerice din facturile Oblio
            import re
            numar_cautat_str = str(numar_factura_cautat).strip()
            
            print(f"eMag Oblio: Caută numărul '{numar_cautat_str}' în {len(oblio_valide)} facturi")
            
            for idx, row in oblio_valide.iterrows():
                factura_completa = row['Factura_clean']
                
                # Extrage partea numerică din factură
                match_numeric = re.search(r'\d+', factura_completa)
                if match_numeric:
                    numar_din_factura = match_numeric.group()
                    
                    # Verifică dacă numerele se potrivesc
                    if numar_din_factura == numar_cautat_str:
                        print(f"eMag Oblio: ✓ GĂSIT! Număr {numar_cautat_str} -> Factură {factura_completa}")
                        return numar_din_factura
                    
                    # Verifică și dacă numărul căutat este conținut în factură
                    if numar_cautat_str in factura_completa:
                        print(f"eMag Oblio: ✓ GĂSIT (conținut)! Număr {numar_cautat_str} în Factură {factura_completa}")
                        return numar_din_factura
            
            print(f"eMag Oblio: ✗ Nu s-a găsit numărul {numar_cautat_str}")
            return None
                
        except Exception as e:
            print(f"eMag Oblio: Eroare la căutarea numărului {numar_factura_cautat}: {e}")
            return None

    def proceseaza_borderouri(self, folder, path_gomag, path_extras, tip):
        rezultate = []
        erori = []
        # Citește Gomag XLSX
        try:
            gomag = pd.read_excel(path_gomag)
        except Exception as e:
            erori.append(f"Eroare la citirea Gomag: {e}")
            return rezultate, erori

        gomag.columns = gomag.columns.str.strip().str.lower()
        if 'awb' not in gomag.columns:
            erori.append("Fișierul Gomag nu conține coloana 'AWB'. Coloane găsite: " + ", ".join(gomag.columns))
            return rezultate, erori
        if 'numar factura' not in gomag.columns:
            erori.append("Fișierul Gomag nu conține coloana 'Numar Factura'. Coloane găsite: " + ", ".join(gomag.columns))
            return rezultate, erori
        gomag['awb_normalizat'] = gomag['awb'].astype(str).str.replace(' ', '').str.lstrip('0')

        # Verifică dacă fișierul Oblio este disponibil (folosim calea selectată din GUI)
        facturi_oblio_disponibile = bool(self.path_oblio.get() and os.path.exists(self.path_oblio.get()))
        print(f"{tip}: Oblio disponibil: {facturi_oblio_disponibile} (Cale: {self.path_oblio.get()})")

        for file in os.listdir(folder):
            if not (file.endswith('.xlsx') or file.endswith('.csv')):
                continue
            path = os.path.join(folder, file)
            suma_total = None  # Inițializare pentru fiecare fișier
            
            try:
                if tip == "GLS":
                    borderou = pd.read_excel(path, header=7, dtype={'Număr colet': str})
                    if not {'Număr colet', 'Sumă ramburs'}.issubset(borderou.columns):
                        erori.append(f"{tip}: {file} - Nu am găsit coloanele 'Număr colet' sau 'Sumă ramburs'. Coloane găsite: {list(borderou.columns)}")
                        continue
                    awb_col = 'Număr colet'
                    suma_col = 'Sumă ramburs'
                    borderou['AWB_normalizat'] = borderou[awb_col].astype(str).str.replace(r'\.0$', '', regex=True).str.replace(' ', '').str.lstrip('0')
                    gomag['AWB_normalizat'] = gomag['awb'].astype(str).str.replace(' ', '').str.lstrip('0')
                elif tip == "Sameday":
                    xls = pd.ExcelFile(path)
                    # Extrage totalul din sheet-ul 'client'
                    if "client" not in xls.sheet_names:
                        erori.append(f"{tip}: {file} - Nu am găsit sheet-ul 'client'. Sheet-uri găsite: {xls.sheet_names}")
                        continue
                    client_sheet = pd.read_excel(xls, sheet_name="client")
                    client_sheet.columns = client_sheet.columns.str.strip() # Normalizăm numele coloanelor
                    print(f"Sameday: {file} - Coloane sheet 'client': {list(client_sheet.columns)}") # Debug print
                    if 'Suma totala' not in client_sheet.columns:
                        erori.append(f"{tip}: {file} - Sheet-ul 'client' nu conține coloana 'Suma totala'. Coloane găsite: {list(client_sheet.columns)}")
                        continue
                    try:
                        suma_total_raw = client_sheet['Suma totala'].iloc[1]
                        suma_total = pd.to_numeric(suma_total_raw, errors='coerce')
                        if pd.isna(suma_total):
                            raise ValueError(f"Nu s-a putut converti '{suma_total_raw}' la un număr.")
                        print(f"Sameday: {file} - Suma totala extrasa: {suma_total}") # Debug print
                    except (IndexError, ValueError, TypeError) as e:
                        erori.append(f"{tip}: {file} - Eroare la extragerea 'Suma totala' din sheet-ul 'client': {e}")
                        suma_total = None # Asigură că suma_total este None dacă extragerea eșuează
                        print(f"Sameday: {file} - Eroare la extragerea sumei: {e}") # Debug print

                    # Citește sheet-ul 'expeditii' pentru datele principale
                    if "expeditii" not in xls.sheet_names:
                        erori.append(f"{tip}: {file} - Nu am găsit sheet-ul 'expeditii'. Sheet-uri găsite: {xls.sheet_names}")
                        continue
                    borderou = pd.read_excel(xls, sheet_name="expeditii")
                    if not {'AWB', 'Suma ramburs'}.issubset(borderou.columns):
                        erori.append(f"{tip}: {file} - Nu am găsit coloanele 'AWB' sau 'Suma ramburs'. Coloane găsite: {list(borderou.columns)}")
                        continue
                    awb_col = 'AWB'
                    suma_col = 'Suma ramburs'
                    borderou['AWB_normalizat'] = borderou[awb_col].astype(str).str.strip()
                    gomag['AWB_normalizat'] = gomag['awb'].astype(str).str.strip()

                    # NOU: Dacă suma_total nu a putut fi extrasă din sheet-ul 'client', o calculăm din 'expeditii'
                    if suma_total is None:
                        try:
                            # Asigură-te că coloana 'Suma ramburs' este numerică
                            borderou['Suma ramburs'] = pd.to_numeric(borderou['Suma ramburs'], errors='coerce')
                            suma_total = borderou['Suma ramburs'].sum()
                            print(f"Sameday: {file} - Suma totala calculata din expeditii: {suma_total}") # Debug print
                        except Exception as e:
                            erori.append(f"{tip}: {file} - Eroare la calcularea sumei totale din sheet-ul 'expeditii': {e}")
                            suma_total = None # Reset to None if calculation fails

                else:
                    erori.append(f"{tip}: {file} - Tip borderou necunoscut.")
                    continue
            except Exception as e:
                erori.append(f"Eroare la citirea borderoului {file}: {e}")
                continue

            # Această secțiune este acum doar pentru GLS, deoarece Sameday își extrage totalul din sheet-ul 'client'
            if tip == "GLS":
                total_row = borderou[borderou[awb_col].isna() & borderou[suma_col].notna()]
                if not total_row.empty:
                    suma_total = float(total_row[suma_col].values[0])
                    # Elimină rândul de total din borderou pentru potrivire
                    borderou = borderou[~(borderou[awb_col].isna() & borderou[suma_col].notna())]
            elif tip == "Sameday" and suma_total is None:
                erori.append(f"{tip}: {file} - Suma totală nu a putut fi extrasă corect din sheet-ul 'client'. Verifică formatul fișierului.")

            potrivite = pd.merge(borderou, gomag, on='AWB_normalizat', how='left', suffixes=('_borderou', '_gomag'))

            # LOGICĂ NOUĂ: Caută în Oblio pentru AWB-urile fără factură
            print(f"{tip}: {file} - Începe căutarea în Oblio pentru AWB-uri fără factură...")
            # NU mai făcem copy(), ci lucrăm direct cu indexurile din potrivite
            awb_fara_factura_mask = (potrivite['numar factura'].isna()) | (potrivite['numar factura'] == 0)
            awb_fara_factura_indices = potrivite[awb_fara_factura_mask].index
            
            if len(awb_fara_factura_indices) > 0:
                print(f"{tip}: {file} - {len(awb_fara_factura_indices)} AWB-uri fără factură, caută în Oblio...")
                
                for idx in awb_fara_factura_indices:
                    row = potrivite.loc[idx]
                    awb = row[awb_col]
                    suma = row[suma_col]
                    
                    # Extrage numele clientului și data livrării din borderou
                    nume_client = 'NECUNOSCUT'
                    data_livrare = ''
                    
                    if tip == "GLS":
                        # Pentru GLS, numele clientului poate fi în altă coloană
                        nume_client = row.get('Nume', '') or row.get('Client', '') or row.get('Destinatar', '') or 'NECUNOSCUT'
                        data_livrare = row.get('Data livrare', '') or row.get('Data', '') or ''
                    elif tip == "Sameday":
                        # Pentru Sameday, caută în coloanele disponibile
                        nume_client = row.get('Destinatar', '') or row.get('Nume', '') or row.get('Client', '') or 'NECUNOSCUT'
                        data_livrare = row.get('Data livrare', '') or row.get('Data', '') or ''
                    
                    if suma and facturi_oblio_disponibile:
                        # Încearcă căutarea în Oblio doar pe baza sumei (mapare Q -> D)
                        numar_factura_oblio = self._cauta_in_oblio(
                            suma, nume_client, data_livrare, tip, awb
                        )
                        
                        if numar_factura_oblio:
                            # Actualizează în potrivite folosind index-ul corect
                            print(f"DEBUG OBLIO: ÎNAINTE de actualizare - potrivite.at[{idx}, 'numar factura'] = {potrivite.at[idx, 'numar factura']}")
                            potrivite.at[idx, 'numar factura'] = numar_factura_oblio
                            print(f"DEBUG OBLIO: DUPĂ actualizare - potrivite.at[{idx}, 'numar factura'] = {potrivite.at[idx, 'numar factura']}")
                            print(f"{tip}: ✓ AWB {awb} REZOLVAT din Oblio - Factură: {numar_factura_oblio}")
                        else:
                            print(f"{tip}: ✗ AWB {awb} nu a fost găsit nici în Oblio")
                    else:
                        print(f"{tip}: Sărit AWB {awb} - date insuficiente sau Oblio indisponibil")

            # Verifică din nou dacă mai lipsesc facturi după căutarea în Oblio
            print(f"DEBUG: {tip}: {file} - Verifică AWB-uri după căutarea în Oblio...")
            facturi_ramase_fara_numar = potrivite[potrivite['numar factura'].isna() | (potrivite['numar factura'] == 0)]
            print(f"DEBUG: {tip}: {file} - AWB-uri rămase fără factură după Oblio: {len(facturi_ramase_fara_numar)}")
            
            for idx, row in potrivite.iterrows():
                if pd.isna(row.get('numar factura', None)) or row.get('numar factura', 0) == 0:
                    erori.append(f"{tip}: {file} - Pentru AWB {row[awb_col]} nu am identificat factura nici în Gomag, nici în Oblio.")

            rezultate.append({'borderou': file, 'potrivite': potrivite, 'suma_total': suma_total})

        return rezultate, erori

    def proceseaza_netopia(self, folder_netopia, path_gomag):
        erori = []
        tranzactii_netopia = []

        if not folder_netopia or not os.path.exists(folder_netopia):
            erori.append(f"Folderul Netopia nu există sau nu este valid: {folder_netopia}")
            return [], erori

        # Citește Gomag
        try:
            gomag = pd.read_excel(path_gomag, dtype=str)
            gomag.columns = gomag.columns.str.strip().str.lower()
            gomag['numar comanda'] = gomag['numar comanda'].astype(str).str.strip()
        except Exception as e:
            erori.append(f"Eroare la citirea Gomag pentru Netopia: {e}")
            return [], erori

        for file in os.listdir(folder_netopia):
            if not file.endswith('.csv'):
                continue
            path = os.path.join(folder_netopia, file)
            try:
                netopia_df = pd.read_csv(path, sep=',', encoding='utf-8', dtype=str, skip_blank_lines=True)
                netopia_df.columns = netopia_df.columns.str.strip().str.replace('"', '').str.replace("'", "")
                netopia_df = netopia_df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)

                # Caută coloanele necesare
                col_procesat = None
                col_creditat = None
                for col in netopia_df.columns:
                    if col.lower() == "procesat":
                        col_procesat = col
                    elif col.lower() == "creditat":
                        col_creditat = col

                if not col_procesat:
                    erori.append(f"Fișierul Netopia {file} nu conține coloana 'Procesat'. Coloane găsite: {list(netopia_df.columns)}")
                    continue
                if not col_creditat:
                    erori.append(f"Fișierul Netopia {file} nu conține coloana 'Creditat'. Coloane găsite: {list(netopia_df.columns)}")
                    continue

                # Convertește valorile numerice
                netopia_df[col_procesat] = pd.to_numeric(netopia_df[col_procesat].str.replace(',', '.'), errors='coerce')
                netopia_df[col_creditat] = pd.to_numeric(netopia_df[col_creditat].str.replace(',', '.'), errors='coerce')

                # Umple valorile NaN cu 0
                netopia_df[col_procesat] = netopia_df[col_procesat].fillna(0)
                netopia_df[col_creditat] = netopia_df[col_creditat].fillna(0)

                # Extrage numărul de comandă din descriere pentru toate rândurile
                netopia_df['numar_comanda_extras'] = netopia_df['Descriere'].str.extract(r'Comanda nr\. (\d+)')
                netopia_df['numar_comanda_extras'] = netopia_df['numar_comanda_extras'].astype(str).str.strip()

                # Grupează pe comandă și calculează suma netă (procesat + creditat, unde creditat poate fi negativ pentru refunduri)
                comenzi_grupate = netopia_df.groupby('numar_comanda_extras').agg({
                    col_procesat: 'sum',
                    col_creditat: 'sum',
                    'Descriere': 'first'  # păstrează prima descriere
                }).reset_index()

                # Calculează suma netă
                comenzi_grupate['suma_neta'] = comenzi_grupate[col_procesat] + comenzi_grupate[col_creditat]

                # Filtrează doar comenzile cu suma netă pozitivă (nu avem refund total)
                tranzactii = comenzi_grupate[comenzi_grupate['suma_neta'] > 0].copy()

                # Asociază cu Gomag după număr comandă
                merge = tranzactii.merge(
                    gomag,
                    left_on='numar_comanda_extras',
                    right_on='numar comanda',
                    how='left',
                    suffixes=('', '_gomag')
                )

                for idx, row in merge.iterrows():
                    tranzactii_netopia.append({
                        'fisier': file,
                        'numar_op': '',  # va fi completat la export
                        'curier': 'Netopia',
                        'numar_factura': row.get('numar factura', ''),
                        'suma': row['suma_neta'],
                        'numar_comanda': row['numar_comanda_extras'],
                        'descriere': row.get('Descriere', ''),
                    })

            except Exception as e:
                erori.append(f"Eroare la citirea sau procesarea fișierului Netopia {file}: {e}")

        return tranzactii_netopia, erori

    def verifica_status_comanda_easysales(self, order_id, path_easysales):
        """
        Verifică statusul unei comenzi în fișierul easySales.
        Returnează 'Canceled' dacă comanda este anulată, altfel returnează None.
        """
        # Funcția nu mai e folosită - post-procesarea se face în _completeaza_comenzi_anulate_emag
        return None

    # Adăugat pentru eMag
    def proceseaza_emag(self, folder_emag, path_easysales):
        print(f"eMag: Începe procesarea. Folder: {folder_emag}, easySales: {path_easysales}")
        erori = []
        rezultate_emag = []

        # Citește fișierul easySales pentru maparea ID comandă -> Număr factură
        try:
            print(f"eMag: Citește fișierul easySales...")
            easysales = pd.read_excel(path_easysales, dtype=str)
            easysales.columns = easysales.columns.str.strip()
            print(f"eMag: easySales citit cu succes. Shape: {easysales.shape}")
            
            # Verifică coloanele necesare
            if 'ID comandă' not in easysales.columns:
                erori.append("Fișierul easySales nu conține coloana 'ID comandă'")
                print(f"eMag: EROARE - Nu există coloana 'ID comandă'. Coloane: {list(easysales.columns)}")
                return [], erori
            if 'Valoarea totală Cu taxă' not in easysales.columns:
                erori.append("Fișierul easySales nu conține coloana 'Valoarea totală Cu taxă'")
                print(f"eMag: EROARE - Nu există coloana 'Valoarea totală Cu taxă'. Coloane: {list(easysales.columns)}")
                return [], erori
            if 'Numărul facturii' not in easysales.columns:
                erori.append("Fișierul easySales nu conține coloana 'Numărul facturii'")
                print(f"eMag: EROARE - Nu există coloana 'Numărul facturii'. Coloane: {list(easysales.columns)}")
                return [], erori
            if 'Status' not in easysales.columns:
                erori.append("Fișierul easySales nu conține coloana 'Status'")
                print(f"eMag: EROARE - Nu există coloana 'Status'. Coloane: {list(easysales.columns)}")
                return [], erori
                
            # Normalizează ID-urile comenzilor și numerele facturilor - elimină apostroful din față
            easysales['ID comandă'] = easysales['ID comandă'].astype(str).str.strip()
            easysales['ID comandă'] = easysales['ID comandă'].str.lstrip("'")
            easysales['ID comandă'] = easysales['ID comandă'].str.lstrip("`")
            easysales['ID comandă'] = easysales['ID comandă'].str.lstrip("'")
            easysales = easysales[easysales['ID comandă'] != '']
            
            # Normalizează coloana Status
            easysales['Status'] = easysales['Status'].astype(str).str.strip()
            
            # Afișează statistici despre statusuri
            total_inainte_filtrare = len(easysales)
            status_counts = easysales['Status'].value_counts()
            print(f"eMag: Statistici statusuri înainte de filtrare:")
            for status, count in status_counts.items():
                print(f"  - {status}: {count}")
            
            # Păstrează o copie COMPLETĂ pentru verificarea statusurilor (inclusiv Canceled)
            easysales_status = easysales.copy()

            # FILTREAZĂ COMENZILE ANULATE pentru maparea facturilor - exclude comenzile cu status "Canceled"
            easysales_filtrat = easysales[easysales['Status'] != 'Canceled'].copy()
            total_dupa_filtrare = len(easysales_filtrat)
            comenzi_anulate = total_inainte_filtrare - total_dupa_filtrare
            
            print(f"eMag: FILTRARE COMENZI:")
            print(f"  - Total comenzi înainte: {total_inainte_filtrare}")
            print(f"  - Comenzi anulate (Canceled): {comenzi_anulate}")
            print(f"  - Comenzi valide după filtrare: {total_dupa_filtrare}")
            
            # Folosește datele FILTRATE pentru mapări de facturi (fără Canceled)
            easysales_mapare = easysales_filtrat
            easysales_mapare['Numărul facturii'] = easysales_mapare['Numărul facturii'].astype(str).str.strip()
            easysales_mapare['Numărul facturii'] = easysales_mapare['Numărul facturii'].str.lstrip("'")
            easysales_mapare['Numărul facturii'] = easysales_mapare['Numărul facturii'].str.lstrip("`")
            easysales_mapare['Numărul facturii'] = easysales_mapare['Numărul facturii'].str.lstrip("'")
            
            print(f"eMag: easySales normalizat și filtrat. Rânduri finale valide pentru mapare: {len(easysales_mapare)}")
            
            # Creează dicționar pentru mapare rapidă Order ID -> Număr factură (doar pentru comenzile valide)
            mapare_facturi = dict(zip(easysales_mapare['ID comandă'], easysales_mapare['Numărul facturii']))
            print(f"eMag: Creat dicționar de mapare cu {len(mapare_facturi)} intrări (exclude comenzile anulate)")
            
        except Exception as e:
            erori.append(f"Eroare la citirea fișierului easySales: {e}")
            print(f"eMag: EXCEPȚIE la citirea easySales: {e}")
            return [], erori

        # ====== NOUA LOGICĂ DE CALCUL EMAG ======
        # Folosește calculatorul îmbunătățit pentru gruparea pe perioade
        print("\n=== CALCULUL EMAG IMBUNATATIT ===")
        rezultate_calcule = self._calculeaza_emag_pe_perioade(folder_emag, erori)

        # Salvează rezultatele pentru folosire ulterioară
        self.rezultate_emag_perioade = rezultate_calcule

        print(f"eMag: Calculat pentru {len(rezultate_calcule)} perioade")
        for rezultat in rezultate_calcule:
            print(f"  Perioada {rezultat['period']}: {rezultat['total_final']:.2f} RON")

        # *** PARTEA PRINCIPALĂ: PROCESAREA FIȘIERELOR DP CONFORM SPECIFICAȚIILOR ***
        # Colectează toate datele din fișierele DP și procesează perioada de referință
        toate_datele_emag = []
        files_in_folder = os.listdir(folder_emag)
        print(f"eMag: Fișiere în folder: {files_in_folder}")
        
        for file in files_in_folder:
            print(f"eMag: Analizez fișierul: {file}")
            if not file.endswith('.xlsx') or not file.startswith('nortia_dp_'):
                print(f"eMag: Sărim fișierul {file} (nu este DP .xlsx)")
                continue
            
            print("eMag: Procesează fișierul DP " + file)
            path_emag = os.path.join(folder_emag, file)
            
            try:
                emag = pd.read_excel(path_emag)
                emag.columns = emag.columns.str.strip()
                print("eMag: " + file + " citit cu succes. Shape: " + str(emag.shape))
                print("eMag: " + file + " - Coloane disponibile: " + str(list(emag.columns)))
                
                # Verifică coloanele necesare conform specificațiilor
                required_cols = ['Payout date', 'Reference period start', 'Reference period end', 
                               'Order ID', 'Fraction type', 'Client name', 'Fraction value']
                missing_cols = [col for col in required_cols if col not in emag.columns]
                if missing_cols:
                    erori.append("eMag: " + file + " - Lipsesc coloanele: " + str(missing_cols))
                    print("eMag: " + file + " - EROARE - Lipsesc coloanele: " + str(missing_cols))
                    continue
                    
                print("eMag: " + file + " - Toate coloanele necesare sunt prezente")
                
                # Verifică poziția coloanelor conform specificațiilor (Q=Fraction value, K=Fraction type, L=Client name)
                cols_list = list(emag.columns)
                if len(cols_list) >= 17 and cols_list[16] != 'Fraction value':
                    print(f"eMag: WARNING - Coloana Q (16) nu este 'Fraction value': {cols_list[16]}")
                if len(cols_list) >= 11 and cols_list[10] != 'Fraction type':
                    print(f"eMag: WARNING - Coloana K (10) nu este 'Fraction type': {cols_list[10]}")
                if len(cols_list) >= 12 and cols_list[11] != 'Client name':
                    print(f"eMag: WARNING - Coloana L (11) nu este 'Client name': {cols_list[11]}")
                
                # Convertește la tipurile corecte
                emag['Order ID'] = emag['Order ID'].astype(str).str.strip()
                emag['Fraction type'] = emag['Fraction type'].astype(str).str.strip()
                emag['Client name'] = emag['Client name'].astype(str).str.strip()
                emag['Fraction value'] = pd.to_numeric(emag['Fraction value'], errors='coerce')
                
                # Elimină rândurile cu valori NaN
                emag = emag.dropna(subset=['Fraction value'])
                
                # CALCULARE SUMA TOTALĂ DP din coloana Q (Fraction value)
                suma_totala_dp = emag['Fraction value'].sum()
                print("eMag: " + file + " - SUMA TOTALĂ DP (Fraction value): " + f"{suma_totala_dp:.2f}" + " RON")
                
                # ANALIZĂ FRACTION TYPE
                fraction_types = emag['Fraction type'].value_counts()
                print(f"eMag: " + file + " - Tipuri fracții găsite: {dict(fraction_types)}")
                
                # DETECTARE REFUND COD și POTRIVIRE CU COD CASHING
                refund_mask = emag['Fraction type'].str.contains('Refund', na=False, case=False)
                cod_mask = emag['Fraction type'].str.contains('COD', na=False, case=False) & ~refund_mask
                
                refund_entries = emag[refund_mask].copy()
                cod_entries = emag[cod_mask].copy()
                
                print(f"eMag: " + file + " - COD Cashing entries: {len(cod_entries)}")
                print(f"eMag: " + file + " - Refund entries: {len(refund_entries)}")
                
                # Marchează intrările pentru storno
                emag['este_storno'] = False
                emag['factura_storno'] = ''
                
                # Procesează fiecare Refund pentru a găsi perechea COD corespunzătoare
                for refund_idx, refund_row in refund_entries.iterrows():
                    client_refund = refund_row['Client name']
                    suma_refund = abs(refund_row['Fraction value'])  # Valoarea pozitivă
                    
                    print(f"eMag: " + file + " - Procesez Refund: Client='{client_refund}', Suma={suma_refund:.2f}")
                    
                    # Caută COD Cashing pentru același client cu aceeași sumă
                    matching_cod = cod_entries[
                        (cod_entries['Client name'] == client_refund) &
                        (abs(cod_entries['Fraction value'] - suma_refund) < 0.01)  # toleranță mică
                    ]
                    
                    if not matching_cod.empty:
                        cod_idx = matching_cod.index[0]
                        factura_originala = emag.loc[cod_idx, 'Factura'] if 'Factura' in emag.columns else 'N/A'
                        
                        # Marchează ambele intrări ca storno
                        emag.at[cod_idx, 'este_storno'] = True
                        emag.at[refund_idx, 'este_storno'] = True
                        emag.at[refund_idx, 'factura_storno'] = f"Storno + {factura_originala}"
                        
                        print(f"eMag: " + file + " - ✓ DETECTAT STORNO pentru client '{client_refund}': factura {factura_originala}")
                    else:
                        print(f"eMag: " + file + " - ✗ NU s-a găsit pereche COD pentru Refund client '{client_refund}'")
                
                emag_suma_totala = suma_totala_dp  # Păstrează pentru mai târziu
                
                # ACTUALIZEAZĂ FIȘIERUL CU COLOANA FACTURĂ
                try:
                    print("eMag: " + file + " - Actualizez cu coloana Factura...")
                    if 'Factura' not in emag.columns:
                        emag['Factura'] = ''
                        print("eMag: " + file + " - Coloana Factura adăugată")
                    
                    emag['Order ID'] = emag['Order ID'].astype(str).str.strip()
                    facturi_mapate = 0
                    
                    # FILTRARE PENTRU MAPARE FACTURI: Păstrează doar rândurile cu Order ID valid care nu sunt anulate
                    valid_order_ids = set(mapare_facturi.keys())  # Doar Order ID-urile din easySales filtrat
                    emag_initial_count = len(emag)
                    
                    for idx, row in emag.iterrows():
                        order_id = row['Order ID']
                        if order_id and order_id not in ['', 'nan']:
                            factura = mapare_facturi.get(order_id, '')
                            
                            # Dacă nu s-a găsit factură, va fi procesată în post-procesare
                            # (funcția _completeaza_comenzi_anulate_emag o va completa cu "Canceled" dacă e anulată)
                            
                            emag.at[idx, 'Factura'] = factura
                            if factura and factura != 'Canceled':
                                facturi_mapate += 1
                    
                    print("eMag: " + file + " - " + str(facturi_mapate) + " facturi mapate")
                    
                    # NU MAI ELIMINĂM rândurile din fișierul original!
                    # Păstrăm toate rândurile pentru calculele corecte
                    # DOAR adăugăm coloana Factura fără să eliminăm date
                    
                    # Salvează cu coloana Factura adăugată dar TOATE rândurile păstrate
                    emag.to_excel(path_emag, index=False)
                    print("eMag: " + file + " - Fișier salvat cu coloana Factura (TOATE rândurile păstrate)!")
                    
                    # Re-citire pentru a fi siguri că avem structura actualizată
                    emag = pd.read_excel(path_emag, engine='openpyxl')
                except Exception as e:
                    erori.append("eMag: " + file + " - Eroare la actualizarea cu coloana Factura: " + str(e))
                    print("eMag: " + file + " - EXCEPȚIE la actualizarea coloanei Factura: " + str(e))
                
                print("eMag: === ANALIZĂ FIȘIER DP: " + file + " ===")
                print("eMag: Coloane disponibile: " + str(list(emag.columns)))
                
                # DETECTEAZĂ COMENZI CU MODIFICĂRI/RAMBURSURI
                order_counts = emag['Order ID'].value_counts()
                comenzi_multiple = order_counts[order_counts > 1]
                if len(comenzi_multiple) > 0:
                    print(f"eMag: 🔍 COMENZI CU INTRĂRI MULTIPLE (modificări/rambursuri):")
                    for order_id, count in comenzi_multiple.items():
                        print(f"eMag:   Order {order_id}: {count} intrări")
                        order_rows = emag[emag['Order ID'] == order_id]
                        for idx, row in order_rows.iterrows():
                            fraction_val = row.get('Fraction value', 'N/A')
                            transaction_type = row.get('Transaction type', 'N/A')
                            print(f"eMag:     - {transaction_type}: {fraction_val}")
                
                # ANALIZĂ DETALIATĂ pentru primele câteva înregistrări
                print(f"eMag: PRIMELE 3 ÎNREGISTRĂRI PENTRU ANALIZĂ:")
                for idx in range(min(3, len(emag))):
                    row = emag.iloc[idx]
                    print(f"eMag:   Row {idx+1}:")
                    for col in ['Order ID', 'Transaction type', 'Fraction value', 'Transaction date']:
                        if col in emag.columns and pd.notna(row[col]):
                            print(f"eMag:     {col}: {row[col]}")
                    print("eMag:   ---")
                
                # COLECTEAZĂ DATELE PENTRU GRUPARE PE PERIOADE
                emag['Order ID'] = emag['Order ID'].astype(str).str.strip()
                emag = emag[emag['Order ID'] != '']
                emag['Fraction value'] = pd.to_numeric(emag['Fraction value'], errors='coerce')
                emag = emag.dropna(subset=['Fraction value'])
                
                # PROCESEAZĂ EXACT CONFORM SPECIFICAȚIILOR eMag
                # Verifică dacă este fișier DP și procesează COD/Refund
                emag['Fraction type'] = emag['Fraction type'].astype(str).str.strip() if 'Fraction type' in emag.columns else ''
                emag['Client name'] = emag['Client name'].astype(str).str.strip() if 'Client name' in emag.columns else ''
                
                # Detectează perechi COD Cashing / Refund COD (inclusiv variațiile CO Cashing / Refund CO)
                cod_refund_pairs = []
                refund_indices = emag[emag['Fraction type'].isin(['Refund COD', 'Refund CO'])].index
                
                for idx in refund_indices:
                    refund_row = emag.loc[idx]
                    customer_name = refund_row['Client name']
                    refund_value = abs(refund_row['Fraction value'])  # Valoarea pozitivă a refund-ului
                    
                    # Caută COD Cashing corespunzător pentru același client cu aceeași sumă
                    matching_cod = emag[
                        (emag['Fraction type'].isin(['COD Cashing', 'CO Cashing'])) &
                        (emag['Client name'] == customer_name) &
                        (abs(emag['Fraction value']) == refund_value)
                    ]
                    
                    if not matching_cod.empty:
                        cod_idx = matching_cod.index[0]
                        cod_refund_pairs.append((cod_idx, idx))
                        print("eMag: " + file + " - DETECTAT STORNO: Client " + customer_name + ", Sumă " + f"{refund_value:.2f}")
                
                # Marchează pentru storno în export
                emag['este_storno'] = False
                for cod_idx, refund_idx in cod_refund_pairs:
                    emag.loc[cod_idx, 'este_storno'] = True
                    emag.loc[refund_idx, 'este_storno'] = True
                
                # DEBUG: Afișează suma din acest fișier - folosește suma completă calculată înainte de filtrare
                suma_fisier = emag_suma_totala  # Suma completă cu toate rândurile
                print("eMag: " + file + " - SUMA DIN ACEST FIȘIER: " + f"{suma_fisier:.2f}" + " RON")
                print("eMag: " + file + " - Numărul de înregistrări valide: " + str(len(emag)))
                print("eMag: " + file + " - Perechi COD/Refund detectate: " + str(len(cod_refund_pairs)))
                
                # PĂSTREAZĂ valorile cu semn din DP (inclusiv refund-urile negative)
                print("eMag: " + file + " - Păstrez valorile originale din DP (cu refund-uri negative)")
                
                # Adaugă numele fișierului pentru tracking
                emag['source_file'] = file
                toate_datele_emag.append(emag)
                print("eMag: " + file + " - " + str(len(emag)) + " înregistrări colectate pentru grupare")
                    
            except Exception as e:
                erori.append("Eroare la procesarea fișierului eMag " + file + ": " + str(e))
                print("eMag: EXCEPȚIE la procesarea fișierului " + file + ": " + str(e))
                continue

        # GRUPARE PE PERIOADE BILUNARE
        if toate_datele_emag:
            # Concatenează toate datele
            df_complet = pd.concat(toate_datele_emag, ignore_index=True)
            print(f"eMag: Total înregistrări colectate: {len(df_complet)}")
            
            # VERIFICARE CRITICĂ: Suma totală din toate fișierele DP
            suma_totala_toate_fisierele = df_complet['Fraction value'].sum()
            print(f"eMag: 🎯 VERIFICARE CRITICĂ - SUMA TOTALĂ din TOATE fișierele DP: {suma_totala_toate_fisierele:.2f} RON")
            if abs(suma_totala_toate_fisierele - 8475.08) < 0.01:
                print(f"eMag: ✅ PERFECT! Suma totală {suma_totala_toate_fisierele:.2f} corespunde cu 8,475.08")
            else:
                print(f"eMag: ❌ PROBLEMĂ! Suma totală {suma_totala_toate_fisierele:.2f} NU corespunde cu 8,475.08!")
                print(f"eMag: Diferența: {suma_totala_toate_fisierele - 8475.08:.2f}")
                # Afișează breakdown per fișier
                for fisier_grup in df_complet['source_file'].unique():
                    suma_fisier = df_complet[df_complet['source_file'] == fisier_grup]['Fraction value'].sum()
                    print(f"eMag:   - {fisier_grup}: {suma_fisier:.2f} RON")
            
            # Grupează după perioada de referință (bilunară)
            grupuri_perioade = df_complet.groupby(['Payout date', 'Reference period start', 'Reference period end']).agg({
                'Order ID': lambda x: list(x),
                'Fraction value': lambda x: list(x),  # Păstrează toate valorile pentru calcul manual
                'source_file': lambda x: list(set(x))  # Lista fișierelor sursa
            }).reset_index()
            
            print(f"eMag: Grupuri de perioade create: {len(grupuri_perioade)}")
            
            # Pentru fiecare perioadă bilunară, creează un rezultat
            for idx, row in grupuri_perioade.iterrows():
                payout_date = row['Payout date']
                ref_start = row['Reference period start'] 
                ref_end = row['Reference period end']
                order_ids_raw = row['Order ID']
                fraction_values_raw = row['Fraction value']
                source_files = row['source_file']
                
                # CALCULEAZĂ SUMA din TOATE valorile Fraction value (inclusiv negative)
                # CONFORM CERINȚELOR: Adună toate valorile din coloana Q, chiar și cele cu minus
                suma_platita = sum(fraction_values_raw)  # Suma din TOATE valorile, inclusiv negative
                
                # Pentru maparea facturilor, creează dicționar cu Order ID unice
                unique_orders = {}
                orders_with_negative_values = set()  # Order ID-uri cu valori negative
                
                for i, order_id in enumerate(order_ids_raw):
                    fraction_val = fraction_values_raw[i]
                    if order_id not in unique_orders:
                        unique_orders[order_id] = fraction_val
                    
                    # Marchează Order ID-urile cu valori negative pentru a lăsa factura goală
                    if fraction_val < 0:
                        orders_with_negative_values.add(order_id)
                
                order_ids_unique = list(unique_orders.keys())
                
                order_ids_duplicate_count = len(order_ids_raw) - len(order_ids_unique)
                
                if order_ids_duplicate_count > 0:
                    suma_cu_duplicate = sum(fraction_values_raw)
                    print(f"eMag: ⚠️ ATENȚIE - Găsite {order_ids_duplicate_count} Order ID-uri duplicate în perioada {ref_start} - {ref_end}")
                    print(f"eMag: Total Order IDs: {len(order_ids_raw)} → Unice: {len(order_ids_unique)}")
                    print(f"eMag: Sumă cu duplicate: {suma_cu_duplicate:.2f} → Sumă corectă: {suma_platita:.2f}")
                    
                    # Afișează care sunt duplicate
                    from collections import Counter
                    duplicates = Counter(order_ids_raw)
                    for order_id, count in duplicates.items():
                        if count > 1:
                            print(f"eMag:   - Order ID {order_id}: apare de {count} ori")
                
                order_ids = order_ids_unique  # Folosește lista deduplicată
                
                # SUMA DP NU PRIMEȘTE TVA - aceasta este suma efectiv încasată din fișierele DP
                ref_year = int(ref_start[:4])  # Extrage anul din YYYY-MM-DD
                ref_month = ref_start[:7]  # Extrage anul și luna din YYYY-MM-DD (ex: 2025-07)
                
                print(f"eMag: Perioada {ref_start} - {ref_end} (luna {ref_month})")
                print(f"eMag: Suma DP din fișiere (TOATE valorile inclusiv negative): {suma_platita:.2f} RON")
                
                # VERIFICARE CRITICĂ: Suma trebuie să fie 8,475.08
                if abs(suma_platita - 8475.08) < 0.01:
                    print(f"eMag: ✅ VERIFICARE OK - Suma {suma_platita:.2f} corespunde cu valoarea așteptată de 8,475.08")
                else:
                    print(f"eMag: ⚠️ VERIFICARE EȘUATĂ - Suma {suma_platita:.2f} NU corespunde cu 8,475.08! Diferența: {suma_platita - 8475.08:.2f}")
                    print(f"eMag: 📊 Analiză detalii frazione:")
                    print(f"eMag:     - Total intrări în DP: {len(fraction_values_raw)}")
                    print(f"eMag:     - Valori pozitive: {[v for v in fraction_values_raw if v > 0]}")
                    print(f"eMag:     - Valori negative: {[v for v in fraction_values_raw if v < 0]}")
                    print(f"eMag:     - Suma pozitive: {sum([v for v in fraction_values_raw if v > 0]):.2f}")
                    print(f"eMag:     - Suma negative: {sum([v for v in fraction_values_raw if v < 0]):.2f}")
                
                print(f"eMag: Procesează perioada {ref_start} - {ref_end}")
                print(f"eMag: Suma plătită finală: {suma_platita:.2f}, Comenzi: {len(order_ids)}, Fișiere: {source_files}")
                
                # Calculează comisionul pentru această perioadă - LOGICĂ NOUĂ PLECÂND DE LA XML
                ref_month = ref_start[:7]  # 2025-07-01 -> 2025-07 sau 2025-06-16 -> 2025-06
                
                # PASUL 1: Caută în XML suma pentru această perioadă de la DANTE INTERNATIONAL
                suma_din_xml = None
                op_gasit_xml = ""
                data_op_xml = ""
                
                # Citește din XML pentru a găsi suma corectă
                try:
                    referinte_op = extrage_referinte_op_din_extras(self.path_extras.get())
                    
                    for op, suma_op, data, batchid_details, details_text in referinte_op:
                        if "DANTE INTERNATIONAL SA" in details_text:
                            # Verifică dacă data OP-ului corespunde cu perioada
                            if data:
                                data_op_obj = pd.to_datetime(data)
                                ref_start_obj = pd.to_datetime(ref_start)
                                ref_end_obj = pd.to_datetime(ref_end)
                                
                                # Verifică dacă data OP-ului este în intervalul de plată pentru perioada respectivă
                                # Pentru perioada 2025-06-16 - 2025-06-30, plata vine în iulie (2025-07-02)
                                # Pentru perioada 2025-07-01 - 2025-07-15, plata vine în iulie (2025-07-18)
                                diferenta_zile = (data_op_obj - ref_end_obj).days
                                
                                if 0 <= diferenta_zile <= 20:  # Plata vine în 1-20 zile după sfârșitul perioadei
                                    suma_din_xml = suma_op
                                    op_gasit_xml = op
                                    data_op_xml = data
                                    print(f"eMag: ✓ GĂSIT OP în XML pentru perioada {ref_start} - {ref_end}")
                                    print(f"eMag: OP: {op}, Data: {data}, Suma XML: {suma_din_xml}")
                                    break
                except Exception as e:
                    print(f"eMag: Eroare la citirea XML pentru OP-uri: {e}")
                
                print(f"eMag: *** FOLOSEȘTE CALCULUL ÎMBUNĂTĂȚIT PE PERIOADE pentru {ref_month} ***")

                # DEBUG: Afișează perioade disponibile
                print(f"eMag DEBUG: Căutăm rezultat pentru perioada {ref_start} - {ref_end}")
                if hasattr(self, 'rezultate_emag_perioade'):
                    print(f"eMag DEBUG: Perioade disponibile în rezultate:")
                    for r in self.rezultate_emag_perioade:
                        print(f"  - {r['period']}: comisioane_total={r.get('comisioane_total', 'N/A')}, ded_total={r.get('ded_total', 'N/A')}")
                else:
                    print(f"eMag DEBUG: rezultate_emag_perioade NU EXISTĂ!")

                # Caută rezultatul pentru această perioadă exactă în rezultatele calculate
                # IMPORTANT: Rezultatele au 'period' format ca "YYYY-MM-DD - YYYY-MM-DD"
                # Trebuie să potrivim pe baza datelor exacte ref_start și ref_end
                candidate_results = []
                if hasattr(self, 'rezultate_emag_perioade'):
                    for result in self.rezultate_emag_perioade:
                        # Extragem datele din result['period'] format "2025-09-01 - 2025-09-15"
                        if ' - ' in result['period']:
                            result_start_str, result_end_str = result['period'].split(' - ')
                            # Comparăm ca stringuri (ambele sunt YYYY-MM-DD)
                            if result_start_str == ref_start and result_end_str == ref_end:
                                candidate_results.append(result)
                                print(f"eMag: ✓ POTRIVIRE EXACTĂ găsită pentru perioada {ref_start} - {ref_end}")
                                break  # Avem potrivire exactă, nu căutăm mai departe

                chosen = None
                if candidate_results:
                    print(f"eMag: ✓ GĂSIT rezultat calculat pentru perioada {ref_start} - {ref_end}")
                    # Folosim rezultatul găsit
                    chosen = candidate_results[0]
                    if suma_din_xml is not None:
                        # Validăm cu suma XML
                        predicted = float(suma_platita) + (float(chosen['dv_total']) - float(chosen.get('dvs_total', 0))) - float(chosen['comisioane_total']) + abs(float(chosen['dcs_total']))
                        diff = abs(predicted - float(suma_din_xml))
                        print(f"eMag: Validare cu XML: Predicted {predicted:.2f} vs XML {suma_din_xml:.2f} | diff {diff:.2f}")

                    # Acum extragem componentele din rezultatul ales și CALCULĂM cu DP din fișierele perioadei
                    voucher_total = float(chosen['dv_total'])
                    dvs_total = float(chosen.get('dvs_total', 0))  # DVS - voucher storno (se scade din DV)

                    # Formula eMag: (DP) + (DV-DVS) - (DC+DCCD+DCCO-DCS) - (DY+DED)
                    dc_total = float(chosen.get('dc_total', 0))
                    dccd_total = float(chosen.get('dccd_total', 0))
                    dcco_total = float(chosen.get('dcco_total', 0))
                    dy_total = float(chosen.get('dy_total', 0))
                    ded_total = float(chosen.get('ded_total', 0))
                    dcs_total = float(chosen.get('dcs_total', 0))  # DCS este negativ (storno)

                    comision_total = dc_total + dccd_total + dcco_total + dy_total + ded_total
                    storno_total = abs(dcs_total)
                    # dp_total din clusterul lunar nu e relevant aici; folosim suma_platita (DP real al perioadei)
                    dp_total = float(suma_platita)

                    # LOG DETALIAT pentru debugging
                    print(f"\n{'='*70}")
                    print(f"eMag: LOG DETALIAT CALCUL pentru perioada {ref_start} - {ref_end}")
                    print(f"{'='*70}")
                    print(f"DEBUG: ded_total extras din chosen: {ded_total}")
                    print(f"DEBUG: chosen dict keys: {chosen.keys()}")
                    print(f"DEBUG: chosen['ded_total'] value: {chosen.get('ded_total', 'NOT FOUND')}")
                    print(f"Componente extrase din rezultatul calculat:")
                    print(f"  DC (comision):           {dc_total:.2f} RON")
                    print(f"  DCCD (comenzi anulate):  {dccd_total:.2f} RON")
                    print(f"  DCCO (comision):         {dcco_total:.2f} RON")
                    print(f"  DY (discount voucher):   {dy_total:.2f} RON")
                    print(f"  DED (alte facturi):      {ded_total:.2f} RON")
                    print(f"  DV (vouchere):           {voucher_total:.2f} RON")
                    print(f"  DVS (voucher storno):    {dvs_total:.2f} RON")
                    print(f"  DCS (comision storno):   {dcs_total:.2f} RON (abs={storno_total:.2f})")
                    print(f"\nDP din fisierele perioadei: {dp_total:.2f} RON")
                    print(f"\nCalcul componente:")
                    print(f"  (DV - DVS) =                      {voucher_total:.2f} - {dvs_total:.2f} = {voucher_total - dvs_total:.2f} RON")
                    print(f"  (DC + DCCD + DCCO + DY + DED) =   {dc_total:.2f} + {dccd_total:.2f} + {dcco_total:.2f} + {dy_total:.2f} + {ded_total:.2f} = {comision_total:.2f} RON")
                    print(f"  DCS (storno) =                    {storno_total:.2f} RON")
                    print(f"{'='*70}\n")
                else:
                    print(f"eMag: ❌ NU s-a găsit rezultat pentru perioada {ref_start} - {ref_end}")
                    print(f"eMag: Perioade disponibile: {[r['period'] for r in self.rezultate_emag_perioade] if hasattr(self, 'rezultate_emag_perioade') else 'N/A'}")
                    # Fallback la sistemul vechi
                    dp_total = suma_platita
                    comision_total = 0
                    voucher_total = 0
                    dvs_total = 0
                    storno_total = 0
                
                # Folosește rezultatul calculat în noul sistem
                if candidate_results:
                    # CALCUL FINAL: (DP) + (DV - DVS) - (DC+DCCD+DY+DED) + DCS
                    # Formula eMag: incasari + vouchere - comision - alte facturi = suma transferata
                    suma_finala_calculata = float(dp_total) + (float(voucher_total) - float(dvs_total)) - float(comision_total) + float(storno_total)

                    print(f"\nFORMULA FINALA:")
                    print(f"  DP + (DV - DVS) - (DC + DCCD + DCCO + DY + DED) + DCS")
                    print(f"  {dp_total:.2f} + ({voucher_total:.2f} - {dvs_total:.2f}) - {comision_total:.2f} + {storno_total:.2f}")
                    print(f"  = {dp_total:.2f} + {voucher_total - dvs_total:.2f} - {comision_total:.2f} + {storno_total:.2f}")
                    print(f"  = {suma_finala_calculata:.2f} RON")
                    print(f"\neMag: Rezultat din calculul pe clustere: {suma_finala_calculata:.2f} RON")
                else:
                    # Fallback: formula veche
                    suma_finala_calculata = dp_total - (comision_total - storno_total) + voucher_total
                    print(f"eMag: Rezultat din calculul fallback: {suma_finala_calculata:.2f} RON")
                
                print(f"eMag: === DETALII CALCUL pentru {ref_month} ===")
                print(f"eMag: DP total: {dp_total:.2f} RON")
                print(f"eMag: DV total (voucher): +{voucher_total:.2f} RON")
                print(f"eMag: Comisioane (fără DCS): -{comision_total:.2f} RON")
                print(f"eMag: DCS (storno - se adună): +{storno_total:.2f} RON")
                print(f"eMag: REZULTAT FINAL: {suma_finala_calculata:.2f} RON")
                
                if suma_din_xml:
                    print(f"eMag: Suma din XML Netopia: {suma_din_xml:.2f} RON")
                    print(f"eMag: Diferența: {suma_finala_calculata - suma_din_xml:.2f} RON")
                else:
                    print(f"eMag: Nu s-a găsit OP corespunzător în XML pentru {ref_month}")
                
                # Folosește ÎNTOTDEAUNA suma calculată cu formula exactă
                comision_cu_tva = comision_total
                suma_finala_pentru_op = suma_finala_calculata
                
                print(f"eMag: === REZULTAT FINAL pentru perioada {ref_start} - {ref_end} ===")
                print(f"eMag: Suma plătită: {suma_platita}")
                print(f"eMag: Comision: {comision_cu_tva}")
                print(f"eMag: Suma finală pentru OP: {suma_finala_pentru_op}")
                print(f"eMag: OP găsit: {op_gasit_xml}, Data: {data_op_xml}")

                # Mapează comenzile cu facturile pentru raport (folosește easySales doar pentru mapare)
                comenzi_mapate = []
                suma_recalculata_easysales = 0.0
                
                print(f"eMag: DEBUG DETALIAT - Mapare Order IDs pentru perioada {ref_start} - {ref_end}:")
                print(f"eMag: Total Order IDs din DP: {len(order_ids)}")
                
                # Creează un dicționar cu valorile din DP pentru comparație
                # FOLOSEȘTE VALORILE DEDUPLICATE din unique_orders în loc de toate rândurile
                dp_values = unique_orders.copy()  # Folosește valorile deduplicate
                
                # PRIMA ETAPĂ: Procesează comenzile care au factură în easySales
                for order_id in order_ids:
                    # DEBUG SPECIAL pentru Order ID 431642847
                    if str(order_id) == "431642847":
                        print(f"eMag: 🔍 DEBUGGING Order ID 431642847:")
                        print(f"eMag:   Order ID în DP: {order_id}")
                        print(f"eMag:   Tip: {type(order_id)}")
                        print(f"eMag:   Valoare DP: {dp_values.get(order_id, 'N/A')}")
                    
                    # Caută în setul pentru mapare (fără Canceled) pentru a găsi factura și valoarea
                    matching_factura = easysales_mapare[easysales_mapare['ID comandă'] == order_id]
                    
                    # DEBUG SPECIAL pentru Order ID 431642847 - verifică căutarea în easySales
                    if str(order_id) == "431642847":
                        print(f"eMag:   Căutare în easySales pentru '{order_id}':")
                        print(f"eMag:   Rezultate găsite: {len(matching_factura)}")
                        if not matching_factura.empty:
                            status = matching_factura.iloc[0].get('Status', 'N/A')
                            print(f"eMag:   Status găsit: '{status}'")
                        
                        # Verifică și alte variante ale Order ID-ului
                        for possible_id in [str(order_id), int(float(str(order_id))) if str(order_id).replace('.', '').isdigit() else None]:
                            if possible_id is not None:
                                test_match = easysales_mapare[easysales_mapare['ID comandă'] == possible_id]
                                print(f"eMag:   Test cu ID '{possible_id}' (tip {type(possible_id)}): {len(test_match)} rezultate")
                    
                    if not matching_factura.empty:
                        nr_factura = matching_factura.iloc[0]['Numărul facturii']
                        valoare = matching_factura.iloc[0]['Valoarea totală Cu taxă']
                        valoare_numeric = pd.to_numeric(valoare, errors='coerce')
                        if not pd.isna(valoare_numeric):
                            suma_recalculata_easysales += valoare_numeric
                        
                        # DEBUGGING: Compară valorile
                        dp_value = dp_values.get(order_id, "N/A")
                        
                        # CAZ SPECIAL: Analiză detaliată pentru factură 233319
                        if nr_factura == "233319" or order_id == "233319":
                            print(f"eMag: 🔍 CAZ SPECIAL - FACTURA 233319:")
                            print(f"eMag:     Order ID: {order_id}")
                            print(f"eMag:     Număr factură: {nr_factura}")
                            print(f"eMag:     easySales - Valoarea totală Cu taxă: {valoare_numeric}")
                            print(f"eMag:     eMag DP - Fraction value: {dp_value}")
                            
                            # Caută alte coloane în easySales pentru această factură
                            factura_row = matching_factura.iloc[0]
                            for col in matching_factura.columns:
                                if 'preț' in col.lower() or 'price' in col.lower() or 'valoare' in col.lower():
                                    print(f"eMag:     easySales - {col}: {factura_row[col]}")
                        
                        print(f"eMag:   Order {order_id}: DP={dp_value} | easySales={valoare_numeric} | Fact={nr_factura}")
                        
                        # Calculează diferența (DP - easySales)
                        diferenta_comanda = None
                        if dp_value != "N/A" and valoare_numeric is not None and not pd.isna(valoare_numeric):
                            diferenta_comanda = float(dp_value) - float(valoare_numeric)
                        
                        # VERIFICĂ STATUSUL ÎN EASYSALES PENTRU COMENZI ANULATE
                        # ATENȚIE: Statusul se caută în setul COMPLET (inclusiv Canceled)
                        status_rows = easysales_status[easysales_status['ID comandă'] == order_id]
                        status_comanda = status_rows.iloc[0].get('Status', '').strip() if not status_rows.empty else ''
                        
                        # NOUĂ LOGICĂ: Pentru valorile negative, lasă factura goală
                        numar_factura_final = nr_factura
                        
                        # PRIORITATE 1: Verifică dacă comanda este ANULATĂ (indiferent de alte condiții)
                        if status_comanda == 'Canceled':
                            numar_factura_final = "Canceled"  # Folosește consistent "Canceled"
                            # Pentru comenzile anulate, PĂSTREAZĂ valoarea din easySales pentru a apărea în export
                            valoare_pentru_export = valoare  # Folosește valoarea din easySales, nu 0!
                            print(f"eMag:   Order {order_id}: Status=Canceled în easySales → Canceled (valoare păstrată: {valoare_pentru_export})")
                        elif order_id in orders_with_negative_values:
                            numar_factura_final = ""  # Lasă factura goală pentru valorile negative
                            valoare_pentru_export = valoare  # Folosește valoarea din easySales
                            print(f"eMag:   Order {order_id} are valoare negativă - factura va fi GOALĂ")
                        elif pd.isna(nr_factura) or str(nr_factura).strip() == '' or str(nr_factura) == 'nan':
                            # Nu există numărul facturii în easySales
                            numar_factura_final = ""  # Lasă gol
                            valoare_pentru_export = valoare  # Folosește valoarea din easySales
                        else:
                            # Caz normal - folosește valoarea din easySales
                            valoare_pentru_export = valoare
                        
                        # DEBUG FINAL pentru Order ID 431642847
                        if str(order_id) == "431642847":
                            print(f"eMag:   Factură finală DUPĂ verificări: '{numar_factura_final}'")
                            print(f"eMag:   Valoare pentru export: {valoare_pentru_export}")
                        
                        # DEBUG pentru Order ID specific menționat
                        if str(order_id) == "431642847":
                            print(f"eMag: 🔍 DEBUG Order ID 431642847:")
                            print(f"eMag:   Nr factură easySales (mapare): '{nr_factura}' (tip: {type(nr_factura)})")
                            print(f"eMag:   Status (din easySales complet): '{status_comanda}' (tip: {type(status_comanda)})")
                            print(f"eMag:   Valoare easySales: {valoare_numeric}")
                            print(f"eMag:   Valoare DP: {dp_values.get(order_id, 0)}")
                            print(f"eMag:   Factură finală ÎNAINTE: '{numar_factura_final}'")
                        
                        comenzi_mapate.append({
                            'order_id': order_id,
                            'numar_factura': numar_factura_final,
                            'valoare': valoare_pentru_export,
                            'diferenta': diferenta_comanda
                        })
                    else:
                        # Pentru order ID-uri care nu sunt în easySales, încearcă căutare în Oblio
                        numar_factura_final = ""
                        valoare_finala = dp_values.get(order_id, 0)
                        
                        # Pentru valorile negative, lasă factură goală
                        if order_id in orders_with_negative_values:
                            print(f"eMag:   Order {order_id} are valoare negativă - factura va fi GOALĂ")
                        else:
                            # Încearcă să caute în Oblio pe baza numărului de factură (order_id poate fi numărul facturii)
                            if order_id and str(order_id).strip() and str(order_id) != 'nan':
                                factura_din_oblio = self._cauta_factura_in_oblio_by_number(order_id)
                                if factura_din_oblio:
                                    numar_factura_final = factura_din_oblio
                                    print(f"eMag:   Order {order_id}: easySales=LIPSĂ | Oblio=GĂSIT ({factura_din_oblio})")
                                else:
                                    # NU S-A GĂSIT NICĂIERI - marchează ca Canceled
                                    numar_factura_final = "Canceled"
                                    erori.append(f"eMag: Order ID {order_id} nu a fost găsit nici în easySales, nici în Oblio - marcat ca Canceled")
                                    print(f"eMag:   Order {order_id}: easySales=LIPSĂ | Oblio=LIPSĂ → Canceled")
                            else:
                                # ID invalid - marchează ca Canceled
                                numar_factura_final = "Canceled"
                                print(f"eMag:   Order {order_id}: ID invalid → Canceled")
                        
                        # Adaugă order_id-ul la rezultate
                        comenzi_mapate.append({
                            'order_id': order_id,
                            'numar_factura': numar_factura_final,
                            'valoare': valoare_finala,
                            'diferenta': None
                        })
                
                # LOGICA SIMPLĂ PENTRU COMENZI FĂRĂ FACTURĂ - caută Status "Canceled"
                print(f"eMag: === VERIFICARE COMENZI FĂRĂ FACTURĂ ===")
                comenzi_fara_factura_actualizate = 0
                
                for comanda in comenzi_mapate:
                    # Dacă comanda nu are factură (câmp gol sau doar spații)
                    if not comanda['numar_factura'] or str(comanda['numar_factura']).strip() == '':
                        order_id = comanda['order_id']
                        
                        # Caută în easySales pe coloana B (ID comandă)
                        # Pentru status folosim setul COMPLET (inclusiv Canceled)
                        matching_easysales = easysales_status[easysales_status['ID comandă'] == order_id]
                        
                        if not matching_easysales.empty:
                            # Verifică coloana C (Status)
                            status = matching_easysales.iloc[0].get('Status', '').strip()
                            
                            if status == 'Canceled':
                                comanda['numar_factura'] = "Canceled"
                                comenzi_fara_factura_actualizate += 1
                                print(f"eMag:   Order {order_id}: Fără factură → easySales Status=Canceled → Canceled")
                            else:
                                print(f"eMag:   Order {order_id}: Fără factură → easySales Status='{status}' → Rămâne gol")
                        else:
                            print(f"eMag:   Order {order_id}: Fără factură → Nu găsit în easySales → Rămâne gol")
                
                print(f"eMag: Total comenzi actualizate cu Canceled: {comenzi_fara_factura_actualizate}")
                
                # INFORMATIV: Afișează comparația sumelor cu explicații
                print(f"eMag: COMPARAȚIE SUME pentru perioada {ref_start} - {ref_end}:")
                print(f"eMag: - Sumă din fișiere DP (folosită pentru calcule): {suma_platita:.2f} RON")
                print(f"eMag: - Sumă din easySales (facturile originale): {suma_recalculata_easysales:.2f} RON")
                diferenta = abs(suma_platita - suma_recalculata_easysales)
                if diferenta > 1.0:
                    print(f"eMag: ℹ️ Diferența {diferenta:.2f} RON - EXPLICAȚII POSIBILE:")
                    print(f"eMag:   • Modificări comenzi (schimbare produse)")
                    print(f"eMag:   • Rambursuri parțiale")
                    print(f"eMag:   • Discount-uri aplicate post-facturare")
                    print(f"eMag:   • Comenzile au intrări multiple în DP (pozitive + negative)")
                else:
                    print(f"eMag: ✓ Sumele se potrivesc perfect!")
                print(f"eMag: Folosesc suma din DP ({suma_platita:.2f}) pentru că reflectă încasările reale")
                
                # Creează numele descriptiv pentru grupa de fișiere
                fisiere_nume = " + ".join(source_files)
                
                rezultate_emag.append({
                    'fisier': f"eMag Perioada {ref_start} - {ref_end} ({fisiere_nume})",
                    'payout_date': payout_date,
                    'ref_period': f"{ref_start} - {ref_end}",
                    'suma_platita': suma_platita,
                    'comision_cu_tva': comision_cu_tva,
                    'voucher_total': voucher_total,
                    'storno_total': storno_total,
                    'suma_finala_pentru_op': suma_finala_pentru_op,
                    'comenzi': comenzi_mapate,
                    'source_files': source_files,
                    # Adaugă componentele individuale pentru debug
                    'dc_total': dc_total if candidate_results else 0,
                    'dccd_total': dccd_total if candidate_results else 0,
                    'dcco_total': dcco_total if candidate_results else 0,
                    'dy_total': dy_total if candidate_results else 0,
                    'ded_total': ded_total if candidate_results else 0,
                    'dvs_total': dvs_total if candidate_results else 0
                })
                print(f"eMag: Perioada {ref_start} - {ref_end} adăugată la rezultate")

        print(f"eMag: Procesare completă. Total perioade: {len(rezultate_emag)}, Total erori: {len(erori)}")
        return rezultate_emag, erori

    def _calculeaza_emag_pe_perioade(self, folder_emag, erori):
        """Calculează eMag pe perioade folosind formula: DP + DV - (DCCO + DCCD + DC + DED + DCS)"""

        tva_rates = {
            "2025-07": 1.19,  # TVA 19% pentru iulie
            "2025-08": 1.21,  # TVA 21% pentru august și ulterior
            "2025-09": 1.21,
        }

        def get_file_period(file_path, file_type):
            """Determină perioada unui fișier eMag bazat pe tipul său"""
            try:
                if file_type == 'dp':
                    # Pentru DP, citește Reference period start/end
                    df = pd.read_excel(file_path, dtype=str)
                    if 'Reference period start' in df.columns and len(df) > 0:
                        period_start = pd.to_datetime(df['Reference period start'].iloc[0])
                        return f"{period_start.year}-{period_start.month:02d}"
                else:
                    # Pentru DC, DCCO, DV, etc - citește coloana Luna
                    df = pd.read_excel(file_path, dtype=str)
                    if 'Luna' in df.columns and len(df) > 0:
                        luna_text = str(df['Luna'].iloc[0]).strip()
                        # Parsează "iulie 2025" -> "2025-07"
                        month_map = {
                            'ianuarie': '01', 'februarie': '02', 'martie': '03',
                            'aprilie': '04', 'mai': '05', 'iunie': '06',
                            'iulie': '07', 'august': '08', 'septembrie': '09',
                            'octombrie': '10', 'noiembrie': '11', 'decembrie': '12'
                        }
                        for month_name, month_num in month_map.items():
                            if month_name in luna_text.lower():
                                year = re.search(r'\d{4}', luna_text)
                                if year:
                                    return f"{year.group()}-{month_num}"
            except Exception as e:
                print(f"Eroare la determinarea perioadei pentru {file_path}: {e}")

            # Fallback: încearcă să extragă din numele fișierului
            match = re.search(r'(\d{2})(\d{4})', os.path.basename(file_path))
            if match:
                month = match.group(1)
                year = match.group(2)
                return f"{year}-{month}"
            return None

        def group_files_by_period():
            """Grupează fișierele eMag pe perioade de referință din DP"""
            from collections import defaultdict
            
            # Structură: {period_key: {'dp_files': [...], 'start': date, 'end': date, 'other_files': {...}}}
            periods = defaultdict(lambda: {
                'dp_files': [], 
                'start': None, 
                'end': None,
                'dp_total': 0,
                'other_files': defaultdict(list)  # {file_type: [{'file': path, 'records': [...]}]}
            })
            
            def parse_date(date_str):
                """Convertește string-ul de dată în obiect datetime"""
                try:
                    if pd.isna(date_str) or date_str == '':
                        return None
                    # Format: 2025-09-01 12:34:56 sau 2025-09-01
                    return pd.to_datetime(str(date_str).split()[0])
                except:
                    return None
            
            def find_period_for_date(date_obj, periods_dict):
                """Găsește perioada DP în care se încadrează o dată"""
                if not date_obj:
                    return None
                
                for period_key, data in periods_dict.items():
                    start = parse_date(data['start'])
                    end = parse_date(data['end'])
                    
                    if start and end and start <= date_obj <= end:
                        return period_key
                
                return None
            
            # Semnături pentru deduplicare DP
            dp_signatures_per_period = {}
            
            def compute_dp_signature(file_path):
                """Calculează o semnătură de conținut pentru un fișier DP pentru deduplicare"""
                try:
                    df = pd.read_excel(file_path, dtype=str)
                    if 'Fraction value' not in df.columns:
                        return None
                    values = pd.to_numeric(df['Fraction value'], errors='coerce')
                    values = values.dropna()
                    row_count = int(values.shape[0])
                    sum_val = round(float(values.sum()), 2)
                    sum_abs = round(float(values.abs().sum()), 2)
                    unique_orders = 0
                    if 'Order ID' in df.columns:
                        unique_orders = int(df['Order ID'].astype(str).str.strip().replace('nan', '').replace('None', '').nunique())
                    return (row_count, sum_val, sum_abs, unique_orders)
                except Exception as e:
                    print(f"Eroare la calcularea semnăturii DP pentru {file_path}: {e}")
                    return None
            
            # Caută în folder-ul principal și în eMag 2
            folders_to_check = [folder_emag]
            folder_emag2 = os.path.join(os.path.dirname(folder_emag), "eMag 2")
            if os.path.exists(folder_emag2):
                folders_to_check.append(folder_emag2)
                print(f"Gasit si folder-ul eMag 2: {folder_emag2}")
            
            # PASUL 1: Citim toate fișierele DP și identificăm perioadele
            print("\n=== PASUL 1: IDENTIFICARE PERIOADE DP ===")
            for current_folder in folders_to_check:
                if not os.path.exists(current_folder):
                    continue
                
                for file in os.listdir(current_folder):
                    if not file.endswith('.xlsx') or not file.startswith('nortia_dp_'):
                        continue
                    
                    file_path = os.path.join(current_folder, file)
                    
                    try:
                        df = pd.read_excel(file_path, dtype=str)
                        
                        # Extrage perioada de referință
                        if 'Reference period start' in df.columns and 'Reference period end' in df.columns:
                            period_start = df['Reference period start'].iloc[0]
                            period_end = df['Reference period end'].iloc[0]
                            period_key = f"{period_start}_{period_end}"
                            
                            # Calculează totalul DP
                            if 'Fraction value' in df.columns:
                                total = pd.to_numeric(df['Fraction value'], errors='coerce').sum()
                                
                                # Verifică duplicatele
                                signature = compute_dp_signature(file_path)
                                if period_key not in dp_signatures_per_period:
                                    dp_signatures_per_period[period_key] = set()
                                
                                if signature and signature in dp_signatures_per_period[period_key]:
                                    print(f"  Duplicat DP (continut): {file} -> perioada {period_start} → {period_end} (sarit)")
                                    continue
                                
                                # Adaugă fișierul
                                periods[period_key]['dp_files'].append(file_path)
                                periods[period_key]['start'] = period_start
                                periods[period_key]['end'] = period_end
                                periods[period_key]['dp_total'] += total
                                
                                if signature:
                                    dp_signatures_per_period[period_key].add(signature)
                                
                                print(f"  Adaugat DP: {file} -> {period_start} → {period_end} (Total: {total:.2f})")
                    
                    except Exception as e:
                        print(f"  Eroare la procesarea DP {file}: {e}")
            
            # PASUL 2: Pentru fiecare tip de fișier, mapăm înregistrările la perioade
            print("\n=== PASUL 2: MAPARE FISIERE LA PERIOADE ===")
            file_type_mapping = {
                'nortia_dc_': ('dc', 'Data finalizare comanda'),
                'nortia_dv_': ('dv', 'Data finalizare comanda'),
                'nortia_dvs_': ('dvs', 'Data finalizare comanda'),
                'nortia_dy_': ('dy', 'Data finalizare retur'),
                'nortia_dcs_': ('dcs', 'Data stornare comanda'),
                'nortia_dcco_': ('dcco', 'Data anulare comanda'),
                'nortia_dccd_': ('dccd', 'Data anulare comanda'),
                'nortia_ded_': ('ded', 'Data finalizare comanda')
            }
            
            for current_folder in folders_to_check:
                if not os.path.exists(current_folder):
                    continue
                
                for file in os.listdir(current_folder):
                    if not file.endswith('.xlsx'):
                        continue
                    
                    # Identifică tipul fișierului
                    file_type = None
                    date_col = None
                    
                    for prefix, (ftype, dcol) in file_type_mapping.items():
                        if file.startswith(prefix):
                            # Verificare specială pentru DC (exclude DCCD și DCCO)
                            if prefix == 'nortia_dc_':
                                if not file.startswith('nortia_dccd_') and not file.startswith('nortia_dcco_'):
                                    file_type = ftype
                                    date_col = dcol
                                    break
                            else:
                                file_type = ftype
                                date_col = dcol
                                break
                    
                    if not file_type:
                        continue
                    
                    file_path = os.path.join(current_folder, file)
                    
                    try:
                        df = pd.read_excel(file_path, dtype=str)
                        
                        if date_col not in df.columns:
                            print(f"  ⚠️  {file_type.upper()} {file}: Nu are coloana {date_col}")
                            continue
                        
                        # Grupăm înregistrările după perioadă
                        period_records = defaultdict(list)
                        
                        for idx, row in df.iterrows():
                            date_str = row.get(date_col)
                            date_obj = parse_date(date_str)
                            
                            if date_obj:
                                period_key = find_period_for_date(date_obj, periods)
                                if period_key:
                                    period_records[period_key].append(row)
                        
                        # Adaugă fișierul la perioade
                        for period_key, records in period_records.items():
                            periods[period_key]['other_files'][file_type].append({
                                'file': file_path,
                                'records': records
                            })
                            print(f"  Adaugat {file_type.upper()}: {file} -> {periods[period_key]['start']} → {periods[period_key]['end']} ({len(records)} înregistrări)")
                    
                    except Exception as e:
                        print(f"  Eroare la procesarea {file_type.upper()} {file}: {e}")
            
            return dict(periods)

        def calculate_dp_total(dp_files):
            """Calculează totalul din fișierele DP"""
            total = 0.0
            for file_path in dp_files:
                try:
                    df = pd.read_excel(file_path, dtype=str)
                    if 'Fraction value' in df.columns:
                        values = pd.to_numeric(df['Fraction value'], errors='coerce')
                        total += values.sum()
                        print(f"DP {os.path.basename(file_path)}: {values.sum():.2f}")
                except Exception as e:
                    print(f"Eroare la procesarea DP {file_path}: {e}")
            return total

        def calculate_dv_total_from_records(dv_file_data):
            """Calculează totalul voucher-elor din înregistrările mapate - FĂRĂ TVA!
            Valorile din fișierele eMag sunt deja nete și se folosesc direct în formula"""
            total = 0.0
            for file_info in dv_file_data:
                file_path = file_info['file']
                records = file_info['records']
                
                for record in records:
                    valoare = pd.to_numeric(record.get('Valoare vouchere', 0), errors='coerce')
                    if pd.notna(valoare):
                        total += abs(valoare)
                
                print(f"DV {os.path.basename(file_path)}: {len(records)} înregistrări, total: {total:.2f}")
            
            # NU aplicăm TVA - valorile se folosesc direct în formula eMag
            return total

        def calculate_dvs_total_from_records(dvs_file_data):
            """Calculează totalul voucher-elor stornate din înregistrările mapate - CU TVA INCLUS!
            DVS este deja cu TVA în fișier"""
            total = 0.0
            for file_info in dvs_file_data:
                file_path = file_info['file']
                records = file_info['records']
                
                for record in records:
                    valoare = pd.to_numeric(record.get('Valoare vouchere', 0), errors='coerce')
                    if pd.notna(valoare):
                        total += abs(valoare)
                
                print(f"DVS {os.path.basename(file_path)}: {len(records)} înregistrări, total: {total:.2f}")
            
            # DVS vine CU TVA inclus, nu aplicăm nimic
            return total

        def calculate_dy_total_from_records(dy_file_data, period):
            """Calculează totalul discount voucher-elor din W2.
            ATENȚIE: DY vine DEJA CU TVA INCLUS în fișier (153.07 este valoarea finală)!"""
            total = 0.0
            
            for file_info in dy_file_data:
                file_path = file_info['file']
                records = file_info['records']
                
                if len(records) > 0:
                    # Pentru DY, citim din celula W2 (totalul este acolo, nu în înregistrări)
                    try:
                        df_no_header = pd.read_excel(file_path, header=None)
                        if df_no_header.shape[1] > 22 and df_no_header.shape[0] > 1:
                            w2_value = pd.to_numeric(df_no_header.iloc[1, 22], errors='coerce')
                            if pd.notna(w2_value):
                                total += abs(w2_value)
                                print(f"DY {os.path.basename(file_path)}: {abs(w2_value):.2f} (din W2, deja cu TVA inclus)")
                    except Exception as e:
                        print(f"Eroare la citirea DY {file_path}: {e}")
            
            # DY vine DEJA cu TVA, NU aplicăm TVA din nou!
            print(f"DY: Total cu TVA inclus: {total:.2f}")
            return total

        def calculate_commission_from_records(file_data, file_type, period):
            """Calculează comisioanele din înregistrările mapate cu TVA aplicat.
            Citește 'Comision Net' (fără TVA) și aplică TVA pentru formula finală.
            Pentru DCCO și DCCD, citește direct din celula T2 (total per fișier).
            Pentru DCS, citește din celula W2 (deja cu TVA inclus, ca DY)."""
            total = 0.0
            tva_rate = get_tva_rate(period)
            
            # DCS citește din T2 (coloana 19 = "Comision Net") - valoare NET fără TVA, ca DCCO/DCCD
            if file_type == 'dcs':
                for file_info in file_data:
                    file_path = file_info['file']
                    try:
                        print(f"DEBUG: Citesc DCS din {os.path.basename(file_path)}...")
                        df = pd.read_excel(file_path, header=None)
                        print(f"DEBUG: DCS shape={df.shape}")
                        
                        if df.shape[1] > 19 and df.shape[0] > 1:
                            value = pd.to_numeric(df.iloc[1, 19], errors='coerce')  # T2 = coloana 19 (Comision Net)
                            if not pd.isna(value):
                                # DCS este negativ (storno), păstrăm valoarea absolută
                                total += abs(value)
                                print(f"DCS {os.path.basename(file_path)}: valoare din T2 (Comision Net) = {abs(value):.2f} (net fără TVA)")
                            else:
                                print(f"DCS {os.path.basename(file_path)}: valoare NaN în T2")
                        else:
                            print(f"DCS {os.path.basename(file_path)}: fișier incomplet (coloane={df.shape[1]}, rânduri={df.shape[0]})")
                    except Exception as e:
                        print(f"EROARE citire DCS {os.path.basename(file_path)}: {e}")
                
                # DCS vine net din T2, aplicăm TVA și returnăm
                total_cu_tva = total * tva_rate
                print(f"DCS: Total net {total:.2f} → cu TVA {tva_rate}: {total_cu_tva:.2f}")
                return total_cu_tva
            
            # DCCO și DCCD au totale în celula T2, nu trebuie sumate din înregistrări
            if file_type in ['dcco', 'dccd']:
                for file_info in file_data:
                    file_path = file_info['file']
                    try:
                        print(f"DEBUG: Citesc {file_type.upper()} din {os.path.basename(file_path)}...")
                        df = pd.read_excel(file_path, header=None)
                        print(f"DEBUG: {file_type.upper()} shape={df.shape}")
                        
                        if df.shape[1] > 19 and df.shape[0] > 1:
                            value = pd.to_numeric(df.iloc[1, 19], errors='coerce')
                            if not pd.isna(value):
                                total += abs(value)
                                print(f"{file_type.upper()} {os.path.basename(file_path)}: valoare din T2 = {abs(value):.2f}")
                            else:
                                print(f"{file_type.upper()} {os.path.basename(file_path)}: valoare NaN în T2")
                        else:
                            print(f"{file_type.upper()} {os.path.basename(file_path)}: fișier incomplet (coloane={df.shape[1]}, rânduri={df.shape[0]})")
                    except Exception as e:
                        print(f"EROARE citire {file_type.upper()} {os.path.basename(file_path)}: {e}")
                
                total_cu_tva = total * tva_rate
                print(f"{file_type.upper()}: Total net {total:.2f} → cu TVA {tva_rate}: {total_cu_tva:.2f}")
                return total_cu_tva
            
            # Pentru celelalte tipuri (DC, DED), citește din înregistrări
            for file_info in file_data:
                file_path = file_info['file']
                records = file_info['records']
                
                for record in records:
                    # DED folosește 'Valoare produs', altele folosesc 'Comision Net'
                    if file_type == 'ded':
                        valoare = pd.to_numeric(record.get('Valoare produs', 0), errors='coerce')
                    else:
                        valoare = pd.to_numeric(record.get('Comision Net', 0), errors='coerce')
                    
                    if pd.notna(valoare):
                        total += abs(valoare)
                
                print(f"{file_type.upper()} {os.path.basename(file_path)}: {len(records)} înregistrări, total net: {total:.2f}")
            
            # Aplicăm TVA pentru a obține valoarea facturată folosită în formula
            total_cu_tva = total * tva_rate
            print(f"{file_type.upper()}: Total net {total:.2f} → cu TVA {tva_rate}: {total_cu_tva:.2f}")
            return total_cu_tva

        def get_tva_rate(period):
            """Obține rata TVA pentru o anumită perioadă"""
            # Perioadele sunt în format "2025-09-01_2025-09-15"
            if period:
                try:
                    year_month = period.split('_')[0][:7]  # Extract "2025-09"
                    return tva_rates.get(year_month, 1.21)
                except:
                    return 1.21
            return 1.21

        def calculate_commission_with_tva(file_path, file_type, period):
            """Calculează comisionul cu TVA pentru un fișier"""
            try:
                tva_rate = tva_rates.get(period, 1.21)

                if file_type == 'dc':
                    df = pd.read_excel(file_path, header=None)
                    if df.shape[1] > 19 and df.shape[0] > 1:
                        value = pd.to_numeric(df.iloc[1, 19], errors='coerce')
                        if not pd.isna(value):
                            total = abs(value) * tva_rate
                            print(f"DC {os.path.basename(file_path)}: {abs(value):.2f} * {tva_rate} = {total:.2f}")
                            return total

                elif file_type == 'dcco':
                    print(f"DEBUG: Citesc DCCO din {os.path.basename(file_path)}...")
                    df = pd.read_excel(file_path, header=None)
                    print(f"DEBUG: DCCO shape={df.shape}")
                    if df.shape[1] > 19 and df.shape[0] > 1:
                        value = pd.to_numeric(df.iloc[1, 19], errors='coerce')
                        if not pd.isna(value):
                            total = abs(value) * tva_rate
                            print(f"DCCO {os.path.basename(file_path)}: {abs(value):.2f} * {tva_rate} = {total:.2f}")
                            return total
                        else:
                            print(f"DCCO {os.path.basename(file_path)}: valoare NaN în T2 - returnez None")
                    else:
                        print(f"DCCO {os.path.basename(file_path)}: nu are suficiente coloane/rânduri - returnez None")

                elif file_type == 'dccd':
                    print(f"DEBUG: Citesc DCCD din {os.path.basename(file_path)}...")
                    df = pd.read_excel(file_path, header=None)
                    print(f"DEBUG: DCCD shape={df.shape}")
                    if df.shape[1] > 19 and df.shape[0] > 1:
                        value = pd.to_numeric(df.iloc[1, 19], errors='coerce')
                        if not pd.isna(value):
                            total = abs(value) * tva_rate
                            print(f"DCCD {os.path.basename(file_path)}: {abs(value):.2f} * {tva_rate} = {total:.2f}")
                            return total
                        else:
                            print(f"DCCD {os.path.basename(file_path)}: valoare NaN în T2 - returnez None")
                    else:
                        print(f"DCCD {os.path.basename(file_path)}: nu are suficiente coloane/rânduri - returnez None")

                elif file_type == 'ded':
                    df = pd.read_excel(file_path, header=None)
                    print(f"DEBUG DED {os.path.basename(file_path)}: shape={df.shape}")
                    if df.shape[1] > 12 and df.shape[0] > 1:
                        value = pd.to_numeric(df.iloc[1, 12], errors='coerce')
                        print(f"DEBUG DED value din M2: {value}")
                        if not pd.isna(value):
                            total = abs(value) * tva_rate
                            print(f"DED {os.path.basename(file_path)}: {abs(value):.2f} * {tva_rate} = {total:.2f}")
                            return total
                        else:
                            print(f"DED {os.path.basename(file_path)}: valoare NaN în M2 - returnez 0")
                    else:
                        print(f"DED {os.path.basename(file_path)}: nu are suficiente coloane ({df.shape[1]}) sau randuri ({df.shape[0]}) - returnez 0")

                elif file_type == 'dcs':
                    df = pd.read_excel(file_path, dtype=str)
                    if 'Comision Net' in df.columns:
                        # Pentru DCS, folosește doar primul rând de date (nu toate)
                        # DCS este storno (valoare negativă în fișier) - păstrăm semnul și aplicăm TVA
                        values = pd.to_numeric(df['Comision Net'], errors='coerce')
                        if len(values) > 0:
                            first_value = values.iloc[0] if not pd.isna(values.iloc[0]) else 0
                            # Păstrăm semnul original și aplicăm TVA
                            total = first_value * tva_rate
                            print(f"DCS {os.path.basename(file_path)}: {first_value:.2f} * {tva_rate} = {total:.2f} (storno)")
                            return total
                    elif df.shape[1] > 3:
                        # Fallback: prima valoare din coloana D
                        if df.shape[0] > 1:
                            first_value = pd.to_numeric(df.iloc[1, 3], errors='coerce')
                            if not pd.isna(first_value):
                                # Păstrăm semnul original și aplicăm TVA
                                total = first_value * tva_rate
                                print(f"DCS {os.path.basename(file_path)}: {first_value:.2f} * {tva_rate} = {total:.2f} (storno)")
                                return total

            except Exception as e:
                print(f"Eroare la procesarea {file_type} {file_path}: {e}")

            return 0.0

        def extract_cluster_key(file_path):
            """Extrage o cheie de cluster din numele fișierului pentru a diferenția seturile din aceeași lună.
            Format tipic: nortia_<tip>_<MMYYYY>_<REPORTID>_v1.xlsx
            Cheia corectă de cluster este <REPORTID> (nu <MMYYYY>). Dacă nu o găsim, 'default'."""
            base = os.path.basename(file_path)
            # Caută patternul cu luna și apoi un al doilea grup numeric mare
            m = re.search(r'_\d{6}_(\d{6,})', base)
            if m:
                digits = m.group(1)
                # Folosim primele 5 cifre pentru a grupa fisierele care apartin aceleiasi perioade
                # (ex: 1758101xxx, 1758102xxx -> 17581; 1759110xxx -> 17591; 1759369xxx, 1759371xxx -> 17593)
                return digits[:5]
            # fallback: încearcă orice secvență numerică lungă
            m2 = re.search(r'(\d{7,})', base)
            if m2:
                return m2.group(1)[:5]
            return 'default'

        # Procesează fișierele
        periods = group_files_by_period()
        results = []

        for period_key, period_data in sorted(periods.items()):
            period_start = period_data['start']
            period_end = period_data['end']
            dp_files = period_data['dp_files']
            dp_total = period_data['dp_total']
            other_files = period_data['other_files']
            
            print(f"\n{'='*100}")
            print(f"PROCESARE PERIOADA: {period_start} → {period_end}")
            print(f"{'='*100}")
            print(f"  DP: {len(dp_files)} fisiere | Total: {dp_total:.2f} RON")
            print(f"  DC: {len(other_files.get('dc', []))} grupe de inregistrari")
            print(f"  DV: {len(other_files.get('dv', []))} grupe de inregistrari")
            print(f"  DVS: {len(other_files.get('dvs', []))} grupe de inregistrari")
            print(f"  DY: {len(other_files.get('dy', []))} grupe de inregistrari")
            print(f"  DCS: {len(other_files.get('dcs', []))} grupe de inregistrari")
            print(f"  DCCO: {len(other_files.get('dcco', []))} grupe de inregistrari")
            print(f"  DCCD: {len(other_files.get('dccd', []))} grupe de inregistrari")
            print(f"  DED: {len(other_files.get('ded', []))} grupe de inregistrari")
            
            # Calculăm totalurile pentru această perioadă
            print(f"\n=== Calcul pentru perioada {period_start} → {period_end} ===")
            
            dv_total = calculate_dv_total_from_records(other_files.get('dv', []))
            dvs_total = calculate_dvs_total_from_records(other_files.get('dvs', []))
            dy_total = calculate_dy_total_from_records(other_files.get('dy', []), period_key)
            dc_total = calculate_commission_from_records(other_files.get('dc', []), 'dc', period_key)
            dcco_total = calculate_commission_from_records(other_files.get('dcco', []), 'dcco', period_key)
            dccd_total = calculate_commission_from_records(other_files.get('dccd', []), 'dccd', period_key)
            ded_total = calculate_commission_from_records(other_files.get('ded', []), 'ded', period_key)
            dcs_total = calculate_commission_from_records(other_files.get('dcs', []), 'dcs', period_key)
            
            # Formula eMag: (DP) + (DV - DVS) - (DC + DCCD + DCCO + DY + DED) + DCS
            comisioane_pozitive = dc_total + dccd_total + dcco_total + dy_total + ded_total
            total_final = dp_total + (dv_total - dvs_total) - comisioane_pozitive + dcs_total

            print(f"\n--- Rezumat {period_start} → {period_end} ---")
            print(f"DP Total: {dp_total:.2f}")
            print(f"DV Total: {dv_total:.2f}")
            print(f"DVS Total: {dvs_total:.2f} (voucher storno - se scade din DV)")
            print(f"DY Total: {dy_total:.2f} (cu TVA)")
            print(f"DC Total: {dc_total:.2f} (cu TVA)")
            print(f"DCCO Total: {dcco_total:.2f} (cu TVA)")
            print(f"DCCD Total: {dccd_total:.2f} (cu TVA)")
            print(f"DED Total: {ded_total:.2f} (cu TVA)")
            print(f"DCS Total: {dcs_total:.2f} (comision stornat - se adauga, cu TVA)")
            print(f"Comisioane pozitive (DC+DCCO+DCCD+DY+DED): {comisioane_pozitive:.2f}")
            print(f"Formula: {dp_total:.2f} + ({dv_total:.2f} - {dvs_total:.2f}) - {comisioane_pozitive:.2f} + {dcs_total:.2f}")
            print(f"TOTAL FINAL: {total_final:.2f}")

            results.append({
                'period': f"{period_start} - {period_end}",
                'period_key': period_key,
                'dp_total': dp_total,
                'dv_total': dv_total,
                'dvs_total': dvs_total,
                'dy_total': dy_total,
                'dcco_total': dcco_total,
                'dccd_total': dccd_total,
                'dc_total': dc_total,
                'ded_total': ded_total,
                'dcs_total': dcs_total,
                'comisioane_total': comisioane_pozitive,
                'total_final': total_final
            })

        print(f"\n{'='*100}")
        print(f"REZUMAT FINAL - {len(results)} PERIOADE PROCESATE")
        print(f"{'='*100}\n")

        return results

    def _proceseaza_comisioane_emag(self, folder_emag, erori):
        """Procesează fișierele de comision eMag și returnează un dicționar luna -> comision cu TVA"""
        comisioane_per_luna = {}
        voucher_per_luna = {}  # Pentru DV (voucher) - se adună
        storno_per_luna = {}  # Pentru DCS (storno) - se adună
        
        print(f"eMag Comision DEBUG: Începe procesarea comisioanelor din {folder_emag}")
        
        def _determina_tva_din_perioada_referinta(file_path):
            """Determină rata TVA bazată pe perioada de referință din fișier, nu pe numele fișierului"""
            try:
                # Încearcă să citească cu header pentru a găsi coloana 'Luna'
                df_header = pd.read_excel(file_path, dtype=str)
                if 'Luna' in df_header.columns and len(df_header) > 0:
                    # Extrage luna din coloana 'Luna' (ex: "2025-07")
                    luna_referinta = str(df_header['Luna'].iloc[0]).strip()
                    if luna_referinta and luna_referinta != 'nan':
                        if luna_referinta <= "2025-07":
                            return 1.19, "19%", luna_referinta  # TVA 19% pentru rapoane din iulie
                        else:
                            return 1.21, "21%", luna_referinta  # TVA 21% pentru rapoarte din august+
                
                # Fallback: încearcă să găsești perioada de referință în header-uri
                # Citește fără header pentru a verifica primele rânduri
                df_no_header = pd.read_excel(file_path, header=None, dtype=str)
                for i in range(min(5, len(df_no_header))):  # Verifică primele 5 rânduri
                    for j in range(min(10, df_no_header.shape[1])):  # Verifică primele 10 coloane
                        cell_value = str(df_no_header.iloc[i, j])
                        if '2025-07' in cell_value:
                            return 1.19, "19%", "2025-07"
                        elif '2025-08' in cell_value:
                            return 1.21, "21%", "2025-08"
                        elif '2025-09' in cell_value:
                            return 1.21, "21%", "2025-09"
                
                # Ultimul fallback: bazat pe numele fișierului (logica veche)
                import re
                match = re.search(r'(\d{2})(\d{4})', file_path)
                if match:
                    month = match.group(1)  # 07
                    year = match.group(2)   # 2025
                    luna_din_nume = f"{year}-{month}"
                    if luna_din_nume <= "2025-07":
                        return 1.19, "19%", luna_din_nume
                    else:
                        return 1.21, "21%", luna_din_nume
                        
                # Default pentru cazuri necunoscute
                return 1.19, "19%", "nedeterminat"
                
            except Exception as e:
                print(f"eMag Comision: Eroare la determinarea TVA pentru {file_path}: {e}")
                return 1.19, "19%", "eroare"
        
        for file in os.listdir(folder_emag):
            print(f"eMag Comision DEBUG: Analizez fișierul {file}")
            if not file.endswith('.xlsx'):
                print(f"  -> Sărit (nu e .xlsx)")
                continue
            
            # Procesează fișierele DC (nortia_dc_*) - comisionul principal din coloana T, rândul 2
            if file.startswith('nortia_dc_') and not file.startswith('nortia_dccd_'):
                print(f"  -> GĂSIT fișier DC: {file}")
                try:
                    path_comision = os.path.join(folder_emag, file)
                    print(f"eMag Comision DC: Procesează {file}")
                    
                    # Determină TVA bazat pe perioada de referință din fișier
                    tva_rate, tva_percent, luna_referinta = _determina_tva_din_perioada_referinta(path_comision)
                    
                    # Citește cu header=None
                    comision_df = pd.read_excel(path_comision, header=None)
                    print(f"eMag Comision DC: {file} - Shape: {comision_df.shape}")
                    
                    # Verifică dacă există coloana T (index 19) și rândul 2 (index 1)
                    if comision_df.shape[1] > 19 and comision_df.shape[0] > 1:
                        # Extrage direct din rândul 2 (index 1), coloana T (index 19)
                        comision_value = comision_df.iloc[1, 19]  # Rândul 2, Coloana T
                        print(f"eMag Comision DC: {file} - Valoare din rândul 2, coloana T: {comision_value}")
                        
                        if pd.notna(comision_value):
                            comision_numeric = pd.to_numeric(comision_value, errors='coerce')
                            if not pd.isna(comision_numeric) and comision_numeric != 0:
                                # Aplică TVA pe comisionul NET bazat pe perioada de referință
                                comision_cu_tva = abs(comision_numeric) * tva_rate
                                
                                if luna_referinta in comisioane_per_luna:
                                    comisioane_per_luna[luna_referinta] += comision_cu_tva
                                else:
                                    comisioane_per_luna[luna_referinta] = comision_cu_tva
                                
                                print(f"eMag Comision DC: {file} - ✓ Luna {luna_referinta}")
                                print(f"  Net (T2): {abs(comision_numeric):.2f}, cu TVA {tva_percent}: {comision_cu_tva:.2f}")
                            else:
                                print(f"eMag Comision DC: {file} - Valoare comision invalidă: {comision_numeric}")
                        else:
                            print(f"eMag Comision DC: {file} - Valoare comision lipsă")
                    else:
                        print(f"eMag Comision DC: {file} - Fișier prea mic: {comision_df.shape}")
                    
                except Exception as e:
                    erori.append(f"Eroare la procesarea fișierului DC eMag {file}: {e}")
                    print(f"eMag Comision DC: EXCEPȚIE la {file}: {e}")
            
            # Procesează fișierele DED (nortia_ded_*) - comisionul DED din coloana M, rândul 2 (NET fără TVA)
            elif file.startswith('nortia_ded_'):
                print(f"  -> GĂSIT fișier DED: {file}")
                try:
                    path_comision = os.path.join(folder_emag, file)
                    print(f"eMag Comision DED: Procesează {file}")
                    
                    # Determină TVA bazat pe perioada de referință din fișier
                    tva_rate, tva_percent, luna_referinta = _determina_tva_din_perioada_referinta(path_comision)
                    
                    # Citește cu header=None
                    comision_df = pd.read_excel(path_comision, header=None)
                    print(f"eMag Comision DED: {file} - Shape: {comision_df.shape}")
                    
                    # Verifică dacă există coloana M (index 12) și rândul 2 (index 1)
                    if comision_df.shape[1] > 12 and comision_df.shape[0] > 1:
                        # Extrage direct din rândul 2 (index 1), coloana M (index 12) NET
                        comision_value = comision_df.iloc[1, 12]  # Rândul 2, Coloana M (NET)
                        print(f"eMag Comision DED: {file} - Valoare NET din rândul 2, coloana M: {comision_value}")
                        
                        if pd.notna(comision_value):
                            comision_numeric = pd.to_numeric(comision_value, errors='coerce')
                            if not pd.isna(comision_numeric) and comision_numeric != 0:
                                # DED este NET în M2 (Valoare produs), aplicăm TVA pe net
                                comision_cu_tva = abs(comision_numeric) * tva_rate
                                
                                if luna_referinta in comisioane_per_luna:
                                    comisioane_per_luna[luna_referinta] += comision_cu_tva
                                else:
                                    comisioane_per_luna[luna_referinta] = comision_cu_tva
                                
                                print(f"eMag Comision DED: {file} - ✓ Luna {luna_referinta}")
                                print(f"  Net (M2): {abs(comision_numeric):.2f}, cu TVA {tva_percent}: {comision_cu_tva:.2f}")
                            else:
                                print(f"eMag Comision DED: {file} - Valoare comision invalidă: {comision_numeric}")
                        else:
                            print(f"eMag Comision DED: {file} - Valoare comision lipsă")
                    else:
                        print(f"eMag Comision DED: {file} - Fișier prea mic: {comision_df.shape}")
                    
                except Exception as e:
                    erori.append(f"Eroare la procesarea fișierului DED eMag {file}: {e}")
                    print(f"eMag Comision DED: EXCEPȚIE la {file}: {e}")
            
            # Procesează fișierele DCCO (nortia_dcco_*) - comisionul DCCO din coloana T, rândul 2
            elif file.startswith('nortia_dcco_'):
                print(f"  -> GĂSIT fișier DCCO: {file}")
                try:
                    path_comision = os.path.join(folder_emag, file)
                    print(f"eMag Comision DCCO: Procesează {file}")
                    
                    # Determină TVA bazat pe perioada de referință din fișier
                    tva_rate, tva_percent, luna_referinta = _determina_tva_din_perioada_referinta(path_comision)
                    
                    comision_df = pd.read_excel(path_comision, header=None)
                    print(f"eMag Comision DCCO: {file} - Shape: {comision_df.shape}")
                    
                    if comision_df.shape[1] > 19 and comision_df.shape[0] > 1:
                        comision_value = comision_df.iloc[1, 19]  # Rândul 2, Coloana T
                        print(f"eMag Comision DCCO: {file} - Valoare din rândul 2, coloana T: {comision_value}")
                        
                        if pd.notna(comision_value):
                            comision_numeric = pd.to_numeric(comision_value, errors='coerce')
                            if not pd.isna(comision_numeric) and comision_numeric != 0:
                                # Aplică TVA pe comisionul NET DCCO bazat pe perioada de referință
                                comision_cu_tva = abs(comision_numeric) * tva_rate
                                
                                if luna_referinta in comisioane_per_luna:
                                    comisioane_per_luna[luna_referinta] += comision_cu_tva
                                else:
                                    comisioane_per_luna[luna_referinta] = comision_cu_tva
                                
                                print(f"eMag Comision DCCO: {file} - ✓ Luna {luna_referinta}")
                                print(f"  Net (T2): {abs(comision_numeric):.2f}, cu TVA {tva_percent}: {comision_cu_tva:.2f}")
                    
                except Exception as e:
                    erori.append(f"Eroare la procesarea fișierului DCCO eMag {file}: {e}")
                    print(f"eMag Comision DCCO: EXCEPȚIE la {file}: {e}")
            
            # Procesează fișierele DCCD (nortia_dccd_*) - comisionul DCCD din coloana T, rândul 2
            elif file.startswith('nortia_dccd_'):
                print(f"  -> GĂSIT fișier DCCD: {file}")
                try:
                    path_comision = os.path.join(folder_emag, file)
                    print(f"eMag Comision DCCD: Procesează {file}")
                    
                    # Determină TVA bazat pe perioada de referință din fișier
                    tva_rate, tva_percent, luna_referinta = _determina_tva_din_perioada_referinta(path_comision)
                    
                    comision_df = pd.read_excel(path_comision, header=None)
                    print(f"eMag Comision DCCD: {file} - Shape: {comision_df.shape}")
                    
                    if comision_df.shape[1] > 19 and comision_df.shape[0] > 1:
                        comision_value = comision_df.iloc[1, 19]  # Rândul 2, Coloana T
                        print(f"eMag Comision DCCD: {file} - Valoare din rândul 2, coloana T: {comision_value}")
                        
                        if pd.notna(comision_value):
                            comision_numeric = pd.to_numeric(comision_value, errors='coerce')
                            if not pd.isna(comision_numeric) and comision_numeric != 0:
                                # Aplică TVA pe comisionul NET DCCD bazat pe perioada de referință
                                comision_cu_tva = abs(comision_numeric) * tva_rate
                                
                                if luna_referinta in comisioane_per_luna:
                                    comisioane_per_luna[luna_referinta] += comision_cu_tva
                                else:
                                    comisioane_per_luna[luna_referinta] = comision_cu_tva
                                
                                print(f"eMag Comision DCCD: {file} - ✓ Luna {luna_referinta}")
                                print(f"  Net (T2): {abs(comision_numeric):.2f}, cu TVA {tva_percent}: {comision_cu_tva:.2f}")
                    
                except Exception as e:
                    erori.append(f"Eroare la procesarea fișierului DCCD eMag {file}: {e}")
                    print(f"eMag Comision DCCD: EXCEPȚIE la {file}: {e}")
            
            # Procesează fișierele DCS (nortia_dcs_*) - storno comision din coloana D, rândul 2 (net fără TVA)
            elif file.startswith('nortia_dcs_'):
                print(f"  -> GĂSIT fișier DCS (storno): {file}")
                try:
                    path_comision = os.path.join(folder_emag, file)
                    print(f"eMag Comision DCS: Procesează {file}")
                    
                    # Încercare 1: detectează coloana 'Comision Net' în header și citește valoarea din rândul 2
                    comision_header_df = pd.read_excel(path_comision)
                    dcs_numeric = None
                    if not comision_header_df.empty:
                        header_map = {str(c).strip().lower(): i for i, c in enumerate(comision_header_df.columns)}
                        if 'comision net' in header_map:
                            col_idx = header_map['comision net']
                            # Recitește cu header=None ca să putem extrage rândul 2 real
                            comision_df_noh = pd.read_excel(path_comision, header=None)
                            if comision_df_noh.shape[0] > 1 and comision_df_noh.shape[1] > col_idx:
                                dcs_val = comision_df_noh.iloc[1, col_idx]
                                dcs_numeric = pd.to_numeric(dcs_val, errors='coerce')
                                print(f"eMag Comision DCS: {file} - Găsit 'Comision Net' la coloana index {col_idx}, valoare r2: {dcs_val}")
                    # Încercare 2: fallback la T2
                    if pd.isna(dcs_numeric) or dcs_numeric is None or dcs_numeric == 0:
                        try:
                            df_noh = pd.read_excel(path_comision, header=None)
                            if df_noh.shape[1] > 19 and df_noh.shape[0] > 1:
                                val_t2 = df_noh.iloc[1, 19]
                                dcs_numeric = pd.to_numeric(val_t2, errors='coerce')
                                print(f"eMag Comision DCS: {file} - Fallback T2: {val_t2}")
                        except Exception as _:
                            pass
                    # Încercare 3: fallback la D2
                    if pd.isna(dcs_numeric) or dcs_numeric is None or dcs_numeric == 0:
                        try:
                            df_noh = pd.read_excel(path_comision, header=None)
                            if df_noh.shape[1] > 3 and df_noh.shape[0] > 1:
                                val_d2 = df_noh.iloc[1, 3]
                                dcs_numeric = pd.to_numeric(val_d2, errors='coerce')
                                print(f"eMag Comision DCS: {file} - Fallback D2: {val_d2}")
                        except Exception as _:
                            pass

                    if not pd.isna(dcs_numeric) and dcs_numeric is not None and dcs_numeric != 0:
                        # DCS (storno) cu TVA - se ADUNĂ la calcul (este negativ deci e retur)
                        tva_rate, tva_percent, luna_referinta = _determina_tva_din_perioada_referinta(path_comision)
                        storno_cu_tva = abs(float(dcs_numeric)) * tva_rate
                        storno_per_luna[luna_referinta] = storno_per_luna.get(luna_referinta, 0) + storno_cu_tva
                        print(f"eMag Storno DCS: {file} - ✓ Luna {luna_referinta}")
                        print(f"  Net (detected): {float(dcs_numeric):.2f}, cu TVA {tva_percent}: +{storno_cu_tva:.2f} (se adună)")
                    
                except Exception as e:
                    erori.append(f"Eroare la procesarea fișierului DCS eMag {file}: {e}")
                    print(f"eMag Comision DCS: EXCEPȚIE la {file}: {e}")
            
            # Procesează fișierele DV (nortia_dv_*) - voucher din coloana X "Valoare vouchere"
            elif file.startswith('nortia_dv_'):
                print(f"  -> GĂSIT fișier DV (voucher): {file}")
                try:
                    path_voucher = os.path.join(folder_emag, file)
                    print(f"eMag Voucher DV: Procesează {file}")
                    
                    # Citește cu header pentru a găsi coloana "Valoare vouchere"
                    voucher_df = pd.read_excel(path_voucher)
                    print(f"eMag Voucher DV: {file} - Shape: {voucher_df.shape}")
                    print(f"eMag Voucher DV: {file} - Coloane: {list(voucher_df.columns)}")
                    
                    # Determină TVA din perioada de referință
                    _tva_rate, _tva_percent, luna_referinta = self._determina_tva_din_perioada_referinta(voucher_df)
                    
                    if 'Valoare vouchere' in voucher_df.columns:
                        # Adună toate valorile din coloana "Valoare vouchere"
                        voucher_df['Valoare vouchere'] = pd.to_numeric(voucher_df['Valoare vouchere'], errors='coerce')
                        total_voucher = voucher_df['Valoare vouchere'].dropna().sum()
                        
                        if total_voucher > 0:
                            # Voucher-ul se ADUNĂ (nu se scade) și nu are TVA
                            if luna_referinta not in voucher_per_luna:
                                voucher_per_luna[luna_referinta] = 0
                            voucher_per_luna[luna_referinta] += total_voucher
                            
                            print(f"eMag Voucher DV: {file} - ✓ Luna {luna_referinta}")
                            print(f"  Total voucher (X): {total_voucher:.2f} (se adună, fără TVA)")
                    else:
                        print(f"eMag Voucher DV: {file} - Nu găsește coloana 'Valoare vouchere'")
                    
                except Exception as e:
                    erori.append(f"Eroare la procesarea fișierului DV eMag {file}: {e}")
                    print(f"eMag Voucher DV: EXCEPȚIE la {file}: {e}")
            else:
                print(f"  -> Sărit (nu e DC, DED, DCCO, DCCD, DCS sau DV)")
        
        # TVA a fost deja aplicat individual pentru fiecare tip de fișier
        # Nu mai aplicăm TVA suplimentar aici
        print(f"eMag Comision DEBUG: Comisioane finale cu TVA aplicat individual: {comisioane_per_luna}")
        print(f"eMag Voucher DEBUG: Voucher-uri finale (fără TVA): {voucher_per_luna}")  
        print(f"eMag Storno DEBUG: Storno finale cu TVA aplicat: {storno_per_luna}")
        return comisioane_per_luna, voucher_per_luna, storno_per_luna

    def _cauta_order_id_in_gomag(self, numar_factura):
        """
        Caută Order ID-ul (Numar Comanda) în fișierul Gomag pe baza numărului facturii.
        Returnează Order ID-ul dacă îl găsește, altfel None.
        """
        try:
            if not hasattr(self, '_gomag_df') or self._gomag_df is None:
                # Încarcă fișierul Gomag doar o dată și îl păstrează în cache
                if not self.path_gomag.get():
                    return None
                    
                self._gomag_df = pd.read_excel(self.path_gomag.get())
                self._gomag_df.columns = self._gomag_df.columns.str.strip().str.lower()
                
                # Verifică dacă are coloanele necesare
                if 'numar factura' not in self._gomag_df.columns or 'numar comanda' not in self._gomag_df.columns:
                    print(f"Gomag: Coloanele necesare lipsesc. Coloane găsite: {list(self._gomag_df.columns)}")
                    return None
            
            # Caută factură în Gomag
            if numar_factura and str(numar_factura).strip() and str(numar_factura) != 'nan':
                # Încearcă căutarea cu numărul exact
                matching_rows = self._gomag_df[self._gomag_df['numar factura'] == numar_factura]
                
                if matching_rows.empty:
                    # Încearcă cu conversia la int (dacă factură este număr)
                    try:
                        factura_int = int(float(str(numar_factura)))
                        matching_rows = self._gomag_df[self._gomag_df['numar factura'] == factura_int]
                    except (ValueError, TypeError):
                        pass
                
                if not matching_rows.empty:
                    order_id = matching_rows.iloc[0]['numar comanda']
                    print(f"Gomag: Factură {numar_factura} → Order ID {order_id}")
                    return order_id
                    
            return None
            
        except Exception as e:
            print(f"Eroare la căutarea în Gomag pentru factură {numar_factura}: {e}")
            return None

    def _citeste_facturi_emag_api(self, data_start, data_end):
        """
        Citește facturile eMag din API pentru o anumită perioadă.
        Returnează un dicționar cu facturile grupate pe perioadă.
        """
        import requests
        import base64
        from datetime import datetime

        username = self.emag_api_username.get().strip()
        password = self.emag_api_password.get().strip()

        if not username or not password:
            print("eMag API: Credențiale lipsă - nu se pot citi facturile")
            return {}

        try:
            # Autentificare Basic Auth
            credentials = base64.b64encode(f'{username}:{password}'.encode()).decode()
            headers = {'Authorization': f'Basic {credentials}'}

            # Citește toate facturile (ultimele 100)
            response = requests.get('https://marketplace-api.emag.ro/api-3/invoice/read',
                                   headers=headers, timeout=30)

            if response.status_code != 200:
                print(f"eMag API: Eroare {response.status_code} - {response.text}")
                return {}

            data = response.json()

            if data.get('isError', False) or 'results' not in data:
                print(f"eMag API: Răspuns invalid - {data.get('messages', [])}")
                return {}

            invoices = data['results'].get('invoices', [])
            print(f"eMag API: {len(invoices)} facturi citite din API")

            # Filtrează facturile după perioadă
            filtered_invoices = []
            for inv in invoices:
                inv_date = datetime.strptime(inv['date'], '%Y-%m-%d')
                if data_start <= inv_date <= data_end:
                    filtered_invoices.append(inv)

            print(f"eMag API: {len(filtered_invoices)} facturi în perioada {data_start} - {data_end}")

            return {'invoices': filtered_invoices}

        except Exception as e:
            print(f"eMag API: Eroare la citirea facturilor: {e}")
            return {}

    def export_opuri(self, rezultate_gls, rezultate_sameday, tranzactii_netopia, rezultate_emag, cale_export, folder_netopia):
        import openpyxl
        import re
        from openpyxl.styles import PatternFill, Font
        from openpyxl import Workbook

        # Resetează cache-ul Gomag pentru a încărca datele fresh
        self._gomag_df = None

        referinte_op = extrage_referinte_op_din_extras(self.path_extras.get())

        wb = Workbook()
        ws = wb.active
        ws.title = "OP-uri"

        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        blue_fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")
        courier_red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        netopia_fill = PatternFill(start_color="FFDAEEF3", end_color="FFDAEEF3", fill_type="solid") # Culoare pentru Netopia

        # Modificare: Adăugare coloană "Nume Borderou", "Order ID", "Diferență eMag" și "Facturi Comision eMag"
        header = ["Data OP", "Număr OP", "Nume Borderou", "Curier", "Order ID", "Număr Factură", "Sumă", "Erori", "Diferență eMag", "Facturi Comision eMag"]
        ws.append(header) # Header-ul se adaugă o singură dată la început
        header_row_idx = ws.max_row

        # Colorare header: "Erori" cu roșu, "Diferență eMag" cu portocaliu, "Facturi Comision eMag" cu verde
        orange_fill = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")
        green_fill = PatternFill(start_color="FF90EE90", end_color="FF90EE90", fill_type="solid")
        for col, col_name in enumerate(header, 1):
            if col_name == "Erori":
                ws.cell(row=header_row_idx, column=col).fill = red_fill
            elif col_name == "Diferență eMag":
                ws.cell(row=header_row_idx, column=col).fill = orange_fill
            elif col_name == "Facturi Comision eMag":
                ws.cell(row=header_row_idx, column=col).fill = green_fill

        for rezultate, curier in [(rezultate_gls, "GLS"), (rezultate_sameday, "Sameday")]:
            for rez in rezultate:
                potrivite = rez.get('potrivite', pd.DataFrame())
                suma_total = rez.get('suma_total', None)
                nume_borderou = rez['borderou'] # Extragem numele borderoului

                facturi_ok = potrivite[~potrivite['numar factura'].isna() & (potrivite['numar factura'] != 0)]
                facturi_ko = potrivite[potrivite['numar factura'].isna() | (potrivite['numar factura'] == 0)]

                # --- SORTARE CRONOLOGICĂ DUPĂ NUMĂRUL FACTURII ---
                # Convertește numărul facturii la numeric pentru sortare corectă
                facturi_ok = facturi_ok.copy()
                facturi_ok['_numar_factura_numeric'] = pd.to_numeric(facturi_ok['numar factura'], errors='coerce')
                facturi_ok = facturi_ok.sort_values(by='_numar_factura_numeric', ascending=True)
                facturi_ok = facturi_ok.drop(columns=['_numar_factura_numeric'])

                erori_exist = not facturi_ko.empty
                erori_text = "DA" if erori_exist else "NU"

                numar_op = ""
                data_op = ""
                if suma_total is not None and referinte_op:
                    for op, suma, data, batchid_details, details_text in referinte_op:
                        if abs(suma - suma_total) < 0.1:
                            numar_op = op
                            data_op = data
                            break

                first_row_for_borderou = True # Flag pentru a scrie data OP, numar OP, nume borderou și curier o singură dată per borderou
                for idx, row in facturi_ok.iterrows():
                    suma = row.get('Sumă ramburs') or row.get('Suma ramburs') or ""
                    
                    # Convertește numărul facturii în întreg pentru a evita apostroful în Excel
                    numar_factura = row['numar factura']
                    if numar_factura and str(numar_factura).strip() and str(numar_factura) != 'nan':
                        try:
                            numar_factura = int(float(str(numar_factura)))
                        except (ValueError, TypeError):
                            numar_factura = str(numar_factura).strip()
                    else:
                        numar_factura = ""
                    
                    # Caută Order ID în Gomag pe baza numărului facturii
                    order_id = self._cauta_order_id_in_gomag(numar_factura) if numar_factura else ""
                    
                    row_data = [
                        data_op if first_row_for_borderou else "",
                        numar_op if first_row_for_borderou else "",
                        nume_borderou if first_row_for_borderou else "",
                        curier if first_row_for_borderou else "",
                        order_id,  # Order ID căutat în Gomag
                        numar_factura,
                        suma,
                        erori_text if first_row_for_borderou else "",
                        "",  # Diferența eMag - nu se aplică pentru GLS/Sameday
                        ""   # Facturi Comision eMag - nu se aplică pentru GLS/Sameday
                    ]
                    ws.append(row_data)
                    row_idx = ws.max_row

                    # Colorare doar celula curierului
                    curier_col_idx = header.index("Curier") + 1
                    if first_row_for_borderou:
                        if curier == "GLS":
                            ws.cell(row=row_idx, column=curier_col_idx).fill = blue_fill
                        elif curier == "Sameday":
                            ws.cell(row=row_idx, column=curier_col_idx).fill = courier_red_fill
                        if erori_exist:
                            erori_col_idx = header.index("Erori") + 1
                            ws.cell(row=row_idx, column=erori_col_idx).fill = red_fill
                    first_row_for_borderou = False

                if facturi_ok.empty:
                    row_data = [
                        data_op,
                        numar_op,
                        nume_borderou,
                        curier,
                        "",  # Order ID - gol dacă nu există facturi
                        "",  # Număr Factură - gol dacă nu există facturi
                        "",  # Sumă - goală dacă nu există facturi
                        erori_text,
                        ""   # Diferența eMag - nu se aplică pentru GLS/Sameday
                    ]
                    ws.append(row_data)
                    row_idx = ws.max_row
                    curier_col_idx = header.index("Curier") + 1
                    if curier == "GLS":
                        ws.cell(row=row_idx, column=curier_col_idx).fill = blue_fill
                    elif curier == "Sameday":
                        ws.cell(row=row_idx, column=curier_col_idx).fill = courier_red_fill
                    if erori_exist:
                        erori_col_idx = header.index("Erori") + 1
                        ws.cell(row=row_idx, column=erori_col_idx).fill = red_fill

                if erori_exist:
                    ws.append(["", "", "", "", "", "AWB-uri fără factură:", "", "", "", ""])
                    for idx, row in facturi_ko.iterrows():
                        suma = row.get('Sumă ramburs') or row.get('Suma ramburs') or ""
                        ws.append(["", "", "", "", "", str(row['AWB_normalizat']), suma, "", "", ""])

                # --- NOU: Adaugă rândul de Total pentru borderou ---
                if suma_total is not None:
                    total_row_data = ["", "", "", "", "", "Total", suma_total, "", "", ""]
                    ws.append(total_row_data)
                    total_row_idx = ws.max_row
                    ws.cell(row=total_row_idx, column=header.index("Număr Factură") + 1).font = Font(bold=True)
                    ws.cell(row=total_row_idx, column=header.index("Sumă") + 1).font = Font(bold=True)

                ws.append([]) # Rând gol între borderouri

        # --- NETOPIA ---
        if tranzactii_netopia:
            referinte_op = extrage_referinte_op_din_extras(self.path_extras.get())

            def get_batchid(fisier):
                m = re.search(r'batchId\.(\d+)', fisier)
                return m.group(1) if m else None

            from collections import defaultdict
            grupare_batch = defaultdict(list)
            for tranz in tranzactii_netopia:
                batchid = get_batchid(tranz['fisier'])
                tranz['batchid'] = batchid
                grupare_batch[batchid].append(tranz)

            for batchid, tranzactii in grupare_batch.items():
                fisier_csv = tranzactii[0]['fisier']
                path_csv = os.path.join(folder_netopia, fisier_csv)
                try:
                    df = pd.read_csv(path_csv, sep=',', encoding='utf-8', dtype=str)
                    df.columns = df.columns.str.strip().str.replace('"', '').str.replace("'", "")
                    df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
                    
                    # Filtrează după prima coloană (#) care conține batchid-ul
                    df_batch = df[df['#'] == batchid]
                    
                    if len(df_batch) == 0:
                        raise ValueError("No matching batch found")
                    
                    df_batch['Procesat'] = pd.to_numeric(df_batch['Procesat'].str.replace(',', '.'), errors='coerce').fillna(0)
                    df_batch['Creditat'] = pd.to_numeric(df_batch['Creditat'].str.replace(',', '.'), errors='coerce').fillna(0)
                    df_batch['Comision'] = pd.to_numeric(df_batch['Comision'].str.replace(',', '.'), errors='coerce').fillna(0)
                    df_batch['TVA'] = pd.to_numeric(df_batch['TVA'].str.replace(',', '.'), errors='coerce').fillna(0)

                    # Calculează totalul facturilor: suma din Procesat + suma din Creditat (care include refund-urile cu minus)
                    total_procesat = df_batch['Procesat'].sum() + df_batch['Creditat'].sum()
                    # Calculează comisioanele din coloanele Comision și TVA
                    total_comision = abs(df_batch['Comision'].sum()) + abs(df_batch['TVA'].sum())
                    total_net = total_procesat - total_comision
                    
                except Exception as e:
                    total_procesat = sum([float(tr['suma']) for tr in tranzactii])
                    total_comision = 0
                    total_net = total_procesat

                # Caută OP-ul potrivit după batchId și sumă netă
                op_gasit = ""
                data_op = ""
                for op, suma_op, data, batchid_details, details_text in referinte_op:
                    # Încearcă să potrivească prin batchid din details
                    if batchid and batchid_details and batchid == batchid_details:
                        if abs(float(suma_op) - total_net) < 1:
                            op_gasit = op
                            data_op = data
                            break
                    # Fallback: încearcă să potrivească prin sumă dacă batchid-ul nu se potrivește
                    elif batchid and ("NETOPIA" in details_text) and abs(float(suma_op) - total_net) < 1:
                        op_gasit = op
                        data_op = data
                        break

                # --- SORTARE CRONOLOGICĂ DUPĂ NUMĂRUL FACTURII ---
                tranzactii_sorted = sorted(tranzactii, key=lambda t: float(t['numar_factura']) if t['numar_factura'] and str(t['numar_factura']).strip() and str(t['numar_factura']) != 'nan' else float('inf'))

                first_row = True
                for tranz in tranzactii_sorted:
                    # Convertește numărul facturii în întreg pentru a evita apostroful în Excel
                    numar_factura = tranz['numar_factura']
                    if numar_factura and str(numar_factura).strip() and str(numar_factura) != 'nan':
                        try:
                            numar_factura = int(float(str(numar_factura)))
                        except (ValueError, TypeError):
                            numar_factura = str(numar_factura).strip()
                    else:
                        numar_factura = ""
                    
                    # Caută Order ID în Gomag pe baza numărului facturii
                    order_id = self._cauta_order_id_in_gomag(numar_factura) if numar_factura else ""
                    
                    ws.append([
                        data_op if first_row and op_gasit else "",
                        op_gasit if first_row and op_gasit else "",
                        tranz['fisier'] if first_row else "",
                        tranz['curier'] if first_row else "",
                        order_id,  # Order ID căutat în Gomag
                        numar_factura,
                        tranz['suma'],
                        "NU",
                        "",   # Diferența eMag - nu se aplică pentru Netopia
                        ""    # Facturi Comision eMag - nu se aplică pentru Netopia
                    ])
                    row_idx_netopia = ws.max_row
                    netopia_col_idx = header.index("Curier") + 1
                    ws.cell(row=row_idx_netopia, column=netopia_col_idx).fill = netopia_fill
                    first_row = False
                # Sub facturi, adaugă sumarul cerut:
                ws.append(["", "", "", "", "", "Comisioane:", round(total_comision, 2), "", "", ""])
                ws.append(["", "", "", "", "", "Total facturi:", round(total_procesat, 2), "", "", ""])
                total_op_row_data = ["", "", "", "", "", f"Total OP:", round(total_net, 2), "", "", ""]
                ws.append(total_op_row_data)
                total_op_row_idx = ws.max_row
                ws.cell(row=total_op_row_idx, column=header.index("Număr Factură") + 1).font = Font(bold=True)
                ws.cell(row=total_op_row_idx, column=header.index("Sumă") + 1).font = Font(bold=True)
                ws.append([]) # Rând gol după fiecare OP/batch
        else:
            pass  # Nu există tranzacții Netopia de procesat

        # --- eMag ---
        if rezultate_emag:
            referinte_op = extrage_referinte_op_din_extras(self.path_extras.get())
            emag_fill = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid") # Culoare portocalie pentru eMag
            
            # Tracking pentru OP-uri deja folosite (pentru a preveni alocare multiplă)
            ops_folosite = []

            for rezultat in rezultate_emag:
                fisier = rezultat['fisier']  # Numele fișierului eMag
                payout_date = rezultat['payout_date']
                ref_period = rezultat['ref_period']
                suma_platita = rezultat['suma_platita']  # Suma efectiv primită
                comision_cu_tva = rezultat['comision_cu_tva']  # Comisionul cu TVA
                voucher_total = rezultat.get('voucher_total', 0)  # DV (voucher)
                storno_total = rezultat.get('storno_total', 0)    # DCS (storno)
                suma_finala_pentru_op = rezultat['suma_finala_pentru_op']  # Pentru mapare cu OP (suma plătită - comision)
                comenzi = rezultat['comenzi']

                # Citește facturile de comision eMag din API pentru perioada acestui DP
                from datetime import datetime
                try:
                    # Parsare perioadă referință (ex: "2025-07-16 - 2025-07-31")
                    ref_parts = ref_period.split(' - ')
                    if len(ref_parts) == 2:
                        data_start = datetime.strptime(ref_parts[0].strip(), '%Y-%m-%d')
                        data_end = datetime.strptime(ref_parts[1].strip(), '%Y-%m-%d')

                        print(f"eMag API: Citesc facturi pentru perioada {data_start.date()} - {data_end.date()}")
                        facturi_api = self._citeste_facturi_emag_api(data_start, data_end)
                        facturi_comision = []

                        if facturi_api and 'invoices' in facturi_api:
                            # Filtrează doar facturile de comision (C-MKTP, E-MKTP, Y-MKTP, etc.)
                            for inv in facturi_api['invoices']:
                                inv_number = inv.get('number', '')
                                if any(prefix in inv_number for prefix in ['C-MKTP', 'E-MKTP', 'Y-MKTP', 'V-MKTP']):
                                    facturi_comision.append(f"{inv_number} ({inv.get('total_with_vat', 0):.2f} RON)")

                            print(f"eMag API: {len(facturi_comision)} facturi de comision găsite")

                        facturi_text = '; '.join(facturi_comision) if facturi_comision else ""
                    else:
                        print(f"eMag API: Format perioadă invalid: {ref_period}")
                        facturi_text = ""
                except Exception as e:
                    print(f"eMag: Eroare la citirea facturilor API pentru {ref_period}: {e}")
                    import traceback
                    traceback.print_exc()
                    facturi_text = ""
                
                # Caută OP-ul potrivit după suma finală și identificarea "DANTE INTERNATIONAL SA" în detalii
                # IMPORTANT: Verifică dacă OP-ul nu a fost deja folosit pentru altă perioadă
                op_gasit = ""
                data_op = ""
                for op, suma_op, data, batchid_details, details_text in referinte_op:
                    if "DANTE INTERNATIONAL SA" in details_text:
                        # Verifică dacă OP-ul a fost deja alocat altei perioade
                        if op in ops_folosite:
                            print(f"eMag: Skipping OP {op} - deja folosit pentru altă perioadă")
                            continue
                        
                        diff = abs(float(suma_op) - suma_finala_pentru_op)
                        if diff < 1:
                            op_gasit = op
                            data_op = data
                            ops_folosite.append(op)  # Marchează OP-ul ca folosit
                            print(f"eMag: OP alocat pentru perioada {ref_period}: {op} ({suma_op:.2f} RON)")
                            break

                # --- SORTARE CRONOLOGICĂ DUPĂ NUMĂRUL FACTURII ---
                # Sortează comenzile după numărul facturii (ignoră ANULATA/CANCELED)
                def extract_factura_numeric(comanda):
                    nf = comanda.get('numar_factura', '')
                    if nf and str(nf).strip() and str(nf) not in ['nan', 'ANULATA', 'CANCELED', 'Canceled']:
                        try:
                            return float(nf)
                        except (ValueError, TypeError):
                            return float('inf')
                    return float('inf')

                comenzi_sorted = sorted(comenzi, key=extract_factura_numeric)

                # Scrie fiecare factură
                first_row = True
                for comanda in comenzi_sorted:
                    # Convertește numărul facturii în întreg pentru a evita apostroful în Excel
                    numar_factura = comanda['numar_factura']
                    if numar_factura and str(numar_factura).strip() and str(numar_factura) != 'nan':
                        try:
                            numar_factura = int(float(str(numar_factura)))
                        except (ValueError, TypeError):
                            numar_factura = str(numar_factura).strip()
                    else:
                        numar_factura = ""
                    
                    # Formatează diferența pentru afișare
                    diferenta_text = ""
                    if 'diferenta' in comanda and comanda['diferenta'] is not None:
                        diferenta_val = comanda['diferenta']
                        if abs(diferenta_val) > 0.01:  # Doar dacă diferența e semnificativă
                            if diferenta_val > 0:
                                diferenta_text = f"+{diferenta_val:.2f}"
                            else:
                                diferenta_text = f"{diferenta_val:.2f}"
                    
                    ws.append([
                        data_op if first_row and op_gasit else "",
                        op_gasit if first_row and op_gasit else "",
                        fisier if first_row else "",
                        "eMag" if first_row else "",
                        comanda['order_id'],  # Order ID - AICI ESTE ORDER ID-ul REAL!
                        numar_factura,
                        comanda['valoare'],
                        "NU",
                        diferenta_text,
                        facturi_text if first_row else ""  # Facturile de comision eMag doar pe primul rând
                    ])
                    row_idx_emag = ws.max_row
                    emag_col_idx = header.index("Curier") + 1
                    ws.cell(row=row_idx_emag, column=emag_col_idx).fill = emag_fill
                    
                    # FORMATARE CU ROȘU pentru facturile ANULATE (acceptă ANULATA/CANCELED)
                    if str(numar_factura).upper() in ["ANULATA", "CANCELED"]:
                        factura_col_idx = header.index("Număr Factură") + 1
                        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                        ws.cell(row=row_idx_emag, column=factura_col_idx).fill = red_fill
                        print(f"eMag Export: Factură {numar_factura} colorată în roșu pentru Order ID {comanda.get('order_id', 'N/A')}")
                    
                    first_row = False
                
                # Adaugă rândurile de detalii pentru eMag (suma plătită, comision, suma finală)
                ws.append(["", "", "", "", "", f"Suma plătită eMag:", round(suma_platita, 2), "", "", ""])
                ws.append(["", "", "", "", "", f"Comision eMag cu TVA:", round(comision_cu_tva, 2), "", "", ""])
                
                # DEBUG: Afișează TOATE componentele formulei pentru validare
                # Extrage componentele individuale din rezultat
                dc_val = rezultat.get('dc_total', 0)
                dccd_val = rezultat.get('dccd_total', 0)
                dcco_val = rezultat.get('dcco_total', 0)
                dy_val = rezultat.get('dy_total', 0)
                ded_val = rezultat.get('ded_total', 0)
                dvs_val = rezultat.get('dvs_total', 0)
                
                print(f"\n{'='*80}")
                print(f"eMag EXPORT DEBUG pentru perioada {ref_period}:")
                print(f"{'='*80}")
                print(f"  DP (suma plătită):           {suma_platita:.2f} RON")
                print(f"  DV (voucher):                {voucher_total:.2f} RON")
                print(f"  DVS (voucher storno):        {dvs_val:.2f} RON")
                print(f"  DC (comision):               {dc_val:.2f} RON")
                print(f"  DCCD (comenzi anulate):      {dccd_val:.2f} RON")
                print(f"  DCCO (comision anulare):     {dcco_val:.2f} RON")
                print(f"  DY (discount voucher):       {dy_val:.2f} RON")
                print(f"  DED (alte facturi):          {ded_val:.2f} RON")
                print(f"  DCS (comision storno):       {storno_total:.2f} RON")
                print(f"  ---")
                print(f"  COMISION TOTAL (DC+DCCD+DCCO+DY+DED): {comision_cu_tva:.2f} RON")
                print(f"  Formula: {suma_platita:.2f} + ({voucher_total:.2f} - {dvs_val:.2f}) - {comision_cu_tva:.2f} + {storno_total:.2f}")
                print(f"  TOTAL OP:                    {suma_finala_pentru_op:.2f} RON")
                print(f"{'='*80}\n")
                
                # Adaugă rândul de total pentru această plată eMag (suma finală pentru OP)
                total_emag_row_data = ["", "", "", "", "", f"Total OP eMag:", round(suma_finala_pentru_op, 2), "", "", ""]
                ws.append(total_emag_row_data)
                total_emag_row_idx = ws.max_row
                ws.cell(row=total_emag_row_idx, column=header.index("Număr Factură") + 1).font = Font(bold=True)
                ws.cell(row=total_emag_row_idx, column=header.index("Sumă") + 1).font = Font(bold=True)
                ws.append([]) # Rând gol după fiecare plată eMag

        try:
            wb.save(cale_export)
            print(f"Export realizat: {cale_export}")
            print("Nr. rânduri exportate:", ws.max_row)
            
            # POST-PROCESARE: Completează comenzile anulate eMag în fișierul exportat
            self._completeaza_comenzi_anulate_emag(cale_export)
            
        except Exception as e:
            print(f"Eroare la salvarea fișierului: {e}")
            raise
    
    def _completeaza_comenzi_anulate_emag(self, cale_export):
        """
        Post-procesează fișierul de export pentru a înlocui celulele goale din coloana Factură 
        cu 'Canceled' pentru comenzile eMag care sunt anulate în easySales.
        """
        try:
            print("eMag: POST-PROCESARE - Completez comenzile anulate...")
            
            # Verifică dacă avem calea către easySales configurată
            path_easysales = self.path_easysales.get().strip()
            if not path_easysales or not os.path.exists(path_easysales):
                print("eMag: Nu pot post-procesa - fișierul easySales nu este configurat sau nu există")
                return
            
            # Citește fișierul easySales pentru statusuri și facturi
            easysales = pd.read_excel(path_easysales, dtype=str)
            easysales.columns = easysales.columns.str.strip()
            
            if 'ID comandă' not in easysales.columns or 'Status' not in easysales.columns:
                print("eMag: Nu pot post-procesa - coloanele 'ID comandă' sau 'Status' lipsesc din easySales")
                return
            
            # Normalizează datele easySales
            easysales['ID comandă'] = easysales['ID comandă'].astype(str).str.strip().str.lstrip("'").str.lstrip("`")
            easysales['Status'] = easysales['Status'].astype(str).str.strip()
            
            # Creează dicționar Order ID -> Status pentru căutare rapidă
            status_dict = {}
            for _, row in easysales.iterrows():
                order_id = row['ID comandă']
                status = row['Status'].strip() if pd.notna(row['Status']) else ''
                if order_id and order_id not in ['', 'nan']:
                    status_dict[order_id] = status
            
            print(f"eMag: Dicționar statusuri creat cu {len(status_dict)} comenzi")

            # Creează dicționar Order ID -> Numărul facturii din easySales (dacă există coloana)
            invoice_dict = {}
            if 'Numărul facturii' in easysales.columns:
                fact_col = 'Numărul facturii'
                easysales[fact_col] = easysales[fact_col].astype(str).str.strip().str.lstrip("'").str.lstrip("`")
                for _, row in easysales.iterrows():
                    oid = row['ID comandă']
                    inv = row.get(fact_col, '')
                    if pd.notna(oid) and str(oid).strip() not in ['', 'nan'] and pd.notna(inv) and str(inv).strip() not in ['', 'nan']:
                        invoice_dict[str(oid).strip()] = str(inv).strip()
                print(f"eMag: Dicționar facturi (easySales) creat: {len(invoice_dict)}")
            else:
                print("eMag: Atenție - easySales nu are coloana 'Numărul facturii' (nu pot completa pentru 'Return')")
            
            # Deschide fișierul de export cu openpyxl pentru modificare
            from openpyxl import load_workbook
            wb = load_workbook(cale_export)
            ws = wb.active
            
            # Găsește indicii coloanelor
            header_row = 1
            order_id_col = None
            factura_col = None
            curier_col = None
            
            for col_idx, cell in enumerate(ws[header_row], 1):
                # CORECT: Order ID este coloana cu ID-ul comenzii eMag
                if cell.value == "Order ID":
                    order_id_col = col_idx
                elif cell.value == "Număr Factură":
                    factura_col = col_idx
                elif cell.value == "Curier":
                    curier_col = col_idx
            
            if not all([order_id_col, factura_col, curier_col]):
                print("eMag: Nu pot găsi coloanele necesare pentru post-procesare")
                return
            
            print(f"eMag: Coloane găsite - Order ID: {order_id_col}, Factură: {factura_col}, Curier: {curier_col}")
            
            # Parcurge toate rândurile și identifică cele eMag cu facturi goale
            comenzi_procesate = 0
            comenzi_canceled = 0
            
            # În export, curierul e scris doar pe primul rând al grupului; transportă ultima valoare non-goală
            curier_anterior = None
            for row_idx in range(2, ws.max_row + 1):  # Start de la rândul 2 (după header)
                curier_cell = ws.cell(row=row_idx, column=curier_col)
                order_id_cell = ws.cell(row=row_idx, column=order_id_col)
                factura_cell = ws.cell(row=row_idx, column=factura_col)
                
                curier_val = curier_cell.value if curier_cell.value not in [None, "", " "] else curier_anterior
                if curier_cell.value not in [None, "", " "]:
                    curier_anterior = curier_cell.value

                # Verifică dacă este rând eMag (direct sau prin carry-over)
                if curier_val == "eMag":
                    order_id = str(order_id_cell.value).strip() if order_id_cell.value else ""
                    factura = str(factura_cell.value).strip() if factura_cell.value else ""
                    
                    # Dacă nu are factură sau are ANULATA și Order ID este valid
                    if (not factura or factura.upper() == "ANULATA") and order_id and order_id not in ['', 'None', 'nan']:
                        comenzi_procesate += 1
                        
                        # Verifică statusul în easySales
                        status = status_dict.get(order_id, '')
                        
                        if status == 'Canceled':
                            # Înlocuiește celula cu "Canceled" și colorează în roșu
                            factura_cell.value = "Canceled"
                            from openpyxl.styles import PatternFill
                            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                            factura_cell.fill = red_fill
                            comenzi_canceled += 1
                            print(f"eMag: ✓ Order ID {order_id} marcat ca 'Canceled' și colorat roșu")
                        elif status == 'Return':
                            # Dacă este Return și există factură în easySales, completează
                            inv = invoice_dict.get(order_id)
                            if inv:
                                factura_cell.value = inv
                                print(f"eMag: ✓ Order ID {order_id} (Return) completat cu factura '{inv}'")
                            else:
                                print(f"eMag: - Order ID {order_id} este 'Return' dar fără factură în easySales")
                        else:
                            print(f"eMag: - Order ID {order_id} status: '{status}' - rămâne gol")
            
            # Salvează modificările
            wb.save(cale_export)
            wb.close()
            
            print(f"eMag: POST-PROCESARE finalizată - {comenzi_procesate} comenzi procesate, {comenzi_canceled} marcate 'Canceled'")
            
        except Exception as e:
            print(f"eMag: Eroare la post-procesare: {e}")
    
    def _apeleaza_script_comenzi_anulate(self, cale_export):
        """Apelează script-ul separat pentru completarea comenzilor anulate"""
        try:
            import subprocess
            
            # Calea către script-ul separat
            script_path = "completeaza_comenzi_anulate.py"
            
            # Calea către easySales
            cale_easysales = self.path_easysales.get().strip()
            
            print("\n" + "="*50)
            print("🔄 APELAREA SCRIPT-ULUI POST-PROCESARE")
            print("="*50)
            
            if not cale_easysales or not os.path.exists(cale_easysales):
                print("⚠️ WARNING: Fișierul easySales nu este configurat - nu se poate face post-procesarea")
                return
            
            # Construiește comanda
            cmd = [
                "python", 
                script_path, 
                cale_export, 
                cale_easysales
            ]
            
            print(f"🚀 Rulez: {' '.join(cmd)}")
            
            # Rulează script-ul
            result = subprocess.run(cmd, 
                                  capture_output=True, 
                                  text=True, 
                                  cwd=os.path.dirname(os.path.abspath(__file__)))
            
            # Afișează output-ul
            if result.stdout:
                print("📄 OUTPUT:")
                print(result.stdout)
                
            if result.stderr:
                print("⚠️ ERRORS:")
                print(result.stderr)
            
            if result.returncode == 0:
                print("✅ Script-ul s-a executat cu succes!")
            else:
                print(f"❌ Script-ul a eșuat cu codul {result.returncode}")
                
        except Exception as e:
            print(f"❌ Eroare la apelarea script-ului: {e}")
    
    def _determina_tva_din_perioada_referinta(self, df_sau_file_path):
        """Determină rata TVA bazată pe perioada de referință din fișier/DataFrame"""
        try:
            # Dacă primește DataFrame direct
            if hasattr(df_sau_file_path, 'columns'):
                df_header = df_sau_file_path
            else:
                # Dacă primește calea fișierului
                df_header = pd.read_excel(df_sau_file_path, dtype=str)
            
            if 'Luna' in df_header.columns and len(df_header) > 0:
                # Extrage luna din coloana 'Luna' (ex: "2025-07")
                luna_referinta = str(df_header['Luna'].iloc[0]).strip()
                if luna_referinta and luna_referinta != 'nan':
                    if luna_referinta <= "2025-07":
                        return 1.19, "19%", luna_referinta  # TVA 19% pentru rapoarte din iulie
                    else:
                        return 1.21, "21%", luna_referinta  # TVA 21% pentru rapoarte din august+
            
            # Fallback: bazat pe numele fișierului dacă există
            if not hasattr(df_sau_file_path, 'columns'):
                import re
                match = re.search(r'(\d{2})(\d{4})', df_sau_file_path)
                if match:
                    month = match.group(1)  # 07
                    year = match.group(2)   # 2025
                    luna_din_nume = f"{year}-{month}"
                    if luna_din_nume <= "2025-07":
                        return 1.19, "19%", luna_din_nume
                    else:
                        return 1.21, "21%", luna_din_nume
            
            # Default pentru 2025
            return 1.19, "19%", "2025-07"  # Default la iulie cu 19%
            
        except Exception as e:
            print(f"Eroare la determinarea TVA: {e}")
            return 1.19, "19%", "2025-07"  # Fallback la 19%

def parse_mt940_file(file_path):
    """
    Parsează un fișier MT940 (Banca Transilvania) și extrage tranzacțiile credit.
    Returnează lista de tuple: (op, suma, data, batchid, details)
    """
    referinte = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # Verifică dacă este format MT940 (conține tag-uri specifice)
        if ':20:' not in content or ':61:' not in content:
            return []

        # Extrage toate blocurile de tranzacții (:61: + :86:)
        # Pattern pentru :61: - formatul: YYMMDDMMDD[C/D]suma,decimaleNTRF...//referinta
        lines = content.replace('\r\n', '\n').replace('\r', '\n').split('\n')

        i = 0
        while i < len(lines):
            line = lines[i].strip()

            if line.startswith(':61:'):
                # Parsează linia :61:
                # Format: :61:2511201120C529,38NTRFNONREF//044ZEXA2532403P0
                statement_line = line[4:]  # Elimină ':61:'

                # Extrage data (primele 6 caractere: YYMMDD)
                if len(statement_line) >= 6:
                    date_str = statement_line[:6]
                    try:
                        # Convertește YYMMDD în format citibil
                        year = 2000 + int(date_str[0:2])
                        month = int(date_str[2:4])
                        day = int(date_str[4:6])
                        data_op = f"{year}-{month:02d}-{day:02d}"
                    except:
                        data_op = ""
                else:
                    data_op = ""

                # Verifică dacă este Credit (C) sau Debit (D)
                # Poziția depinde de format - căutăm C sau D urmat de sumă
                is_credit = False
                suma = 0.0

                # Găsește poziția C/D și suma
                # Format tipic după dată: MMDDC529,38N sau D529,38N
                rest = statement_line[6:]  # După data YYMMDD

                # Saltă data secundară (4 cifre MMDD) dacă există
                if len(rest) >= 4 and rest[:4].isdigit():
                    rest = rest[4:]

                # Acum ar trebui să fie C sau D urmat de sumă
                if rest.startswith('C'):
                    is_credit = True
                    rest = rest[1:]
                elif rest.startswith('D'):
                    is_credit = False
                    rest = rest[1:]

                # Extrage suma (până la prima literă non-cifră, virgulă)
                suma_match = re.match(r'([\d,]+)', rest)
                if suma_match:
                    suma_str = suma_match.group(1).replace(',', '.')
                    try:
                        suma = float(suma_str)
                    except:
                        suma = 0.0

                # Extrage referința OP (după //)
                ref_match = re.search(r'//(\S+)', statement_line)
                op_ref = ref_match.group(1) if ref_match else ""

                # Citește liniile :86: (detalii) - pot fi mai multe linii
                details_lines = []
                j = i + 1
                while j < len(lines):
                    next_line = lines[j].strip()
                    if next_line.startswith(':86:'):
                        details_lines.append(next_line[4:])  # Elimină ':86:'
                        j += 1
                        # Continuă să citească liniile care fac parte din detalii
                        while j < len(lines) and not lines[j].strip().startswith(':'):
                            details_lines.append(lines[j].strip())
                            j += 1
                        break
                    elif next_line.startswith(':'):
                        break
                    j += 1

                details_text = ' '.join(details_lines)

                # Procesează doar tranzacțiile credit relevante
                if is_credit and suma > 0:
                    # Verifică dacă este o tranzacție relevantă (GLS, Sameday, Netopia, eMag)
                    is_relevant = any(keyword in details_text.upper() for keyword in [
                        'GLS', 'GENERAL LOGISTICS',
                        'DELIVERY SOLUTIONS', 'SAMEDAY',
                        'NETOPIA', 'BATCHID',
                        'DANTE INTERNATIONAL', 'EMAG'
                    ])

                    if is_relevant:
                        # Extrage batchId dacă există (pentru Netopia)
                        batchid_in_details = None
                        batch_match = re.search(r'BATCHID\s+(\d+)', details_text, re.IGNORECASE)
                        if batch_match:
                            batchid_in_details = batch_match.group(1)

                        referinte.append((op_ref, suma, data_op, batchid_in_details, details_text))
                        print(f"MT940: Găsit credit relevant: {suma:.2f} RON, ref: {op_ref}, detalii: {details_text[:80]}...")

                i = j if j > i else i + 1
            else:
                i += 1

    except Exception as e:
        print(f"Eroare la parsarea MT940 {file_path}: {e}")

    return referinte


def extrage_referinte_op_din_extras(extras_path):
    referinte = []
    try:
        # Verifică dacă este un folder (pentru MT940 multiple files)
        if os.path.isdir(extras_path):
            print(f"MT940: Procesez folder {extras_path}")
            mt940_files = [f for f in os.listdir(extras_path) if f.lower().endswith('.txt') and 'MT940' in f.upper()]
            print(f"MT940: Găsite {len(mt940_files)} fișiere MT940")

            for filename in mt940_files:
                file_path = os.path.join(extras_path, filename)
                file_refs = parse_mt940_file(file_path)
                referinte.extend(file_refs)

            print(f"MT940: Total {len(referinte)} tranzacții relevante extrase din folder")

        elif extras_path.lower().endswith('.xml'):
            tree = ET.parse(extras_path)
            root = tree.getroot()
            for movement in root.findall('.//movement'):
                ref_element = movement.find('ref')
                credit_element = movement.find('credit')
                value_date_element = movement.find('value_date')
                details_element = movement.find('details')
                details_text = details_element.text if details_element is not None else ""
                if ref_element is not None and credit_element is not None:
                    op = ref_element.text
                    parts = op.split('.')
                    if len(parts) > 2:
                        op = ".".join(parts[:2])
                    suma_str = credit_element.text
                    data_op = value_date_element.text if value_date_element is not None else ""

                    # Extrage batchId din details dacă există
                    batchid_in_details = None
                    if "BatchId" in details_text:
                        batch_match = re.search(r'BatchId\s+(\d+)', details_text)
                        if batch_match:
                            batchid_in_details = batch_match.group(1)

                    try:
                        suma_float = float(suma_str)
                        if suma_float > 0 and ("OLP1." in op or "DELIVERY SOLUTIONS SA" in details_text or "NETOPIA FINANCIAL SERVICES SA" in details_text or "DANTE INTERNATIONAL SA" in details_text):
                            referinte.append((op, suma_float, data_op, batchid_in_details, details_text))
                    except ValueError:
                        continue

        elif extras_path.lower().endswith('.txt'):
            # Verifică dacă este format MT940
            with open(extras_path, 'r', encoding='utf-8') as f:
                content = f.read()

            if ':20:' in content and ':61:' in content:
                # Este format MT940
                print(f"MT940: Detectat fișier MT940 individual: {extras_path}")
                referinte = parse_mt940_file(extras_path)
            else:
                # Format TXT vechi (pentru compatibilitate)
                matches = re.findall(r'(OLP1\.\d+)[^\d]{1,20}([\d\.,]{3,})', content)
                for op, suma in matches:
                    parts = op.split('.')
                    if len(parts) > 2:
                        op = ".".join(parts[:2])
                    suma = suma.replace('.', '').replace(',', '.')
                    try:
                        suma_float = float(suma)
                        referinte.append((op, suma_float, "", None, ""))  # Fără dată și batchid în TXT
                    except:
                        continue
        else:
            print(f"Tip de fișier extras bancar nevalid: {extras_path}. Se acceptă .xml, .txt (MT940), sau folder cu fișiere MT940.")
    except FileNotFoundError:
        print(f"Fișierul/folderul extras bancar nu a fost găsit: {extras_path}")
    except ET.ParseError as e:
        print(f"Eroare la parsarea fișierului XML {extras_path}: {e}")
    except Exception as e:
        print(f"Eroare la citirea extrasului bancar: {e}")
        import traceback
        traceback.print_exc()
    print(f"Extracted OP references: {referinte}") # Debug print
    return referinte

if __name__ == "__main__":
    app = FacturiApp()
    app.mainloop()