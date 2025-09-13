import os
import sys
import pandas as pd
from openpyxl import load_workbook

# Use the top-level 'eMag' folder (conține fișierele pentru iulie 2025 în acest workspace)
BASE = sys.argv[1] if len(sys.argv) > 1 else os.path.join(os.getcwd(), 'eMag')
if not os.path.isdir(BASE):
    raise SystemExit(f"Nu găsesc folderul eMag: {BASE}")

def to_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        s = val.strip().replace('\xa0', ' ').replace(' ', '')
        s = s.replace(',', '.')
        try:
            return float(s)
        except ValueError:
            return None
    return None

def read_cell_any_sheet(file, cell_addr):
    path = os.path.join(BASE, file)
    wb = load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        v = ws[cell_addr].value
        fv = to_float(v)
        if fv is not None:
            return fv
    # if no sheet has a numeric, return 0.0 to avoid crash
    return 0.0

def row2_maxabs_any_sheet(file):
    path = os.path.join(BASE, file)
    wb = load_workbook(path, data_only=True, read_only=True)
    best = None
    for ws in wb.worksheets:
        # scan columns A..Z in row 2
        for col_idx in range(1, 27):
            v = ws.cell(row=2, column=col_idx).value
            fv = to_float(v)
            if fv is None:
                continue
            if best is None or abs(fv) > abs(best):
                best = fv
    return best if best is not None else 0.0

def print_cell_all_sheets(file, cell):
    path = os.path.join(BASE, file)
    wb = load_workbook(path, data_only=True, read_only=True)
    print(f"  Sheet values for {file} at {cell}:")
    for ws in wb.worksheets:
        v = ws[cell].value
        print(f"    - {ws.title}: {v}")

# Try to get DCS net by header 'Comision Net', otherwise fallbacks
def read_dcs_net(file):
    path = os.path.join(BASE, file)
    try:
        df = pd.read_excel(path)
        cols = {str(c).strip().lower(): c for c in df.columns}
        if 'comision net' in cols:
            # read row 2 from headerless
            df_noh = pd.read_excel(path, header=None)
            col_idx = list(df.columns).index(cols['comision net'])
            if df_noh.shape[0] > 1 and df_noh.shape[1] > col_idx:
                return to_float(df_noh.iloc[1, col_idx]) or 0.0
    except Exception:
        pass
    # fallback T2 then D2
    v = read_cell_any_sheet(file, 'T2')
    if v is None or abs(v) < 1e-9:
        v = read_cell_any_sheet(file, 'D2')
    return v or 0.0

# Helpers for specific cells
get_T2 = lambda f: read_cell_any_sheet(f, 'T2')
get_M2 = lambda f: read_cell_any_sheet(f, 'M2')
get_D2 = lambda f: read_cell_any_sheet(f, 'D2')

# DV sum from column "Valoare vouchere"
def sum_dv(file):
    df = pd.read_excel(os.path.join(BASE,file))
    col = None
    for c in df.columns:
        if str(c).strip().lower() == 'valoare vouchere':
            col = c
            break
    if col is None:
        raise RuntimeError('Coloana "Valoare vouchere" lipsă în DV')
    return pd.to_numeric(df[col], errors='coerce').dropna().sum()

# DP sum from column "Fraction value"
def sum_dp(file):
    df = pd.read_excel(os.path.join(BASE,file))
    col = None
    for c in df.columns:
        if str(c).strip().lower() == 'fraction value':
            col = c
            break
    if col is None:
        raise RuntimeError('Coloana "Fraction value" lipsă în DP')
    return pd.to_numeric(df[col], errors='coerce').dropna().sum()

def find_first(pattern_substr):
    for name in os.listdir(BASE):
        if name.lower().endswith('.xlsx') and pattern_substr in name.lower():
            return name
    return None

def find_all(pattern_substr):
    return [name for name in os.listdir(BASE) if name.lower().endswith('.xlsx') and pattern_substr in name.lower()]

if __name__ == '__main__':
    # Files: auto-detect for July 2025 ('072025')
    dcco_f = find_first('nortia_dcco_072025')
    dccd_f = find_first('nortia_dccd_072025')
    dc_f   = find_first('nortia_dc_072025')
    ded_f  = find_first('nortia_ded_072025')
    dcs_f  = find_first('nortia_dcs_072025')
    dv_f   = find_first('nortia_dv_072025')

    # DP: include all dp files (often daily) likely belonging to early August but covering July
    dp_files = find_all('nortia_dp_')
    if not dp_files:
        raise SystemExit('Nu s-au găsit fișiere DP în folderul eMag.')

    # Inspect DCS sheet values to confirm correct cell
    if dcs_f:
        print('--- Inspect DCS sheets ---')
        print_cell_all_sheets(dcs_f, 'D2')
        print_cell_all_sheets(dcs_f, 'T2')
    # Values
    # Print DP files and per-file sums
    print('DP files găsite:')
    dp_file_sums = []
    for f in dp_files:
        s = sum_dp(f)
        dp_file_sums.append((f, s))
        print(f"  - {f}: {s:.2f}")
    dp_total = sum(s for _, s in dp_file_sums)

    # În paralel: calculează suma doar pentru Cashing (exclude Refund) pentru comparație
    def sum_dp_cashings(file):
        df = pd.read_excel(os.path.join(BASE, file))
        cols = {c.strip().lower(): c for c in df.columns}
        if 'fraction type' not in cols or 'fraction value' not in cols:
            return 0.0
        ft = cols['fraction type']; fv = cols['fraction value']
        df[fv] = pd.to_numeric(df[fv], errors='coerce')
        mask = df[ft].astype(str).str.contains('Cashing', case=False, na=False)
        return df.loc[mask, fv].dropna().sum()
    dp_cashings_total = sum(sum_dp_cashings(f) for f in dp_files)
    dcco_net = float(get_T2(dcco_f)) if dcco_f else 0.0
    dccd_net = float(get_T2(dccd_f)) if dccd_f else 0.0
    dc_net   = float(get_T2(dc_f)) if dc_f else 0.0
    ded_net  = float(get_M2(ded_f)) if ded_f else 0.0
    # DCS: prefer D2, fallback to max-abs on row 2 if D2 yields 0
    dcs_net = float(read_dcs_net(dcs_f))
    dv_total = float(sum_dv(dv_f)) if dv_f else 0.0

    # July 2025 VAT
    tva = 1.19

    dcco_g = abs(dcco_net) * tva
    dccd_g = abs(dccd_net) * tva
    dc_g   = abs(dc_net) * tva
    ded_g  = abs(ded_net) * tva
    dcs_g  = abs(dcs_net) * tva

    total_comm = dcco_g + dccd_g + dc_g + ded_g
    total_op = dp_total - total_comm + dv_total + dcs_g

    print(f"DP total (net): {dp_total:.2f}")
    print(f"DP total (doar Cashing): {dp_cashings_total:.2f}")
    print(f"Comision DCCO net/gross: {dcco_net:.2f} / {dcco_g:.2f}")
    print(f"Comision DCCD net/gross: {dccd_net:.2f} / {dccd_g:.2f}")
    print(f"Comision DC   net/gross: {dc_net:.2f} / {dc_g:.2f}")
    print(f"Comision DED  net/gross: {ded_net:.2f} / {ded_g:.2f}")
    print(f"DV total: {dv_total:.2f}")
    print(f"DCS net/gross: {dcs_net:.2f} / {dcs_g:.2f}")
    print(f"Comision total (cu TVA): {total_comm:.2f}")
    print(f"Total OP (DP - comisioane + DV + DCS): {total_op:.2f}")
