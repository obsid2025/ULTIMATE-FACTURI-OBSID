#!/usr/bin/env python3
"""
Script pentru completarea comenzilor eMag anulate cu "Canceled" în fișierul opuri_export.xlsx
Acest script se rulează după generarea fișierului de export pentru a înlocui celulele goale
cu "Canceled" pentru comenzile eMag care sunt anulate în easySales.

Utilizare:
python completeaza_comenzi_anulate.py [cale_fisier_export] [cale_fisier_easysales]

Dacă nu se specifică căi, se folosesc căile implicite:
- opuri_export.xlsx 
- 8 August/Comenzi easySales.xlsx
"""

import pandas as pd
import os
import sys
from openpyxl import load_workbook

def completeaza_comenzi_anulate_emag(cale_export=None, cale_easysales=None):
    """
    Post-procesează fișierul de export pentru a înlocui celulele goale din coloana Factură 
    cu 'Canceled' pentru comenzile eMag care sunt anulate în easySales.
    """
    
    # Setează căile implicite dacă nu sunt specificate
    if cale_export is None:
        cale_export = "opuri_export.xlsx"
    if cale_easysales is None:
        cale_easysales = "8 August/Comenzi easySales.xlsx"
    
    try:
        print("=" * 60)
        print("🔄 COMPLETARE COMENZI ANULATE eMag")
        print("=" * 60)
        print(f"📁 Fișier export: {cale_export}")
        print(f"📁 Fișier easySales: {cale_easysales}")
        
        # Verifică existența fișierelor
        if not os.path.exists(cale_export):
            print(f"❌ EROARE: Fișierul de export nu există: {cale_export}")
            return False
            
        if not os.path.exists(cale_easysales):
            print(f"❌ EROARE: Fișierul easySales nu există: {cale_easysales}")
            return False
        
        print("✅ Ambele fișiere găsite")

        # Citește fișierul easySales pentru statusuri
        print("📖 Citesc fișierul easySales...")
        easysales = pd.read_excel(cale_easysales, dtype=str)
        easysales.columns = easysales.columns.str.strip()

        if 'ID comandă' not in easysales.columns or 'Status' not in easysales.columns:
            print("❌ EROARE: Coloanele 'ID comandă' sau 'Status' lipsesc din easySales")
            return False

        # Normalizează datele easySales
        easysales['ID comandă'] = easysales['ID comandă'].astype(str).str.strip().str.lstrip("'").str.lstrip("`")
        easysales['Status'] = easysales['Status'].astype(str).str.strip()
        
        # Creează dicționar Order ID -> Status pentru căutare rapidă
        status_dict = {}
        for _, row in easysales.iterrows():
            order_id = row['ID comandă']
            status = row['Status'].strip() if pd.notna(row['Status']) else ''
            if order_id and order_id not in ['', 'nan']:
                status_dict[order_id] = status
        
        print(f"📊 Dicționar statusuri creat cu {len(status_dict)} comenzi")
        
        # Afișează statistici statusuri
        status_counts = {}
        for status in status_dict.values():
            status_counts[status] = status_counts.get(status, 0) + 1
        
        print("📈 Statistici statusuri:")
        for status, count in sorted(status_counts.items()):
            print(f"   {status}: {count} comenzi")

        # Creează dicționar Order ID -> Numărul facturii (dacă există coloana în easySales)
        invoice_dict = {}
        if 'Numărul facturii' in easysales.columns:
            fact_col = 'Numărul facturii'
            easysales[fact_col] = easysales[fact_col].astype(str).str.strip().str.lstrip("'").str.lstrip("`")
            for _, row in easysales.iterrows():
                oid = row['ID comandă']
                inv = row.get(fact_col, '')
                if pd.notna(oid) and str(oid).strip() not in ['', 'nan'] and pd.notna(inv) and str(inv).strip() not in ['', 'nan']:
                    invoice_dict[str(oid).strip()] = str(inv).strip()
            print(f"📦 Dicționar facturi (din easySales) creat: {len(invoice_dict)} comenzi cu număr de factură")
        else:
            print("ℹ️  Atenție: easySales nu are coloana 'Numărul facturii' — completarea pentru 'Return' va fi sărită")
        
        # Deschide fișierul de export cu openpyxl pentru modificare
        print("📝 Deschid fișierul de export pentru modificare...")
        wb = load_workbook(cale_export)
        ws = wb.active
        
        # Găsește indicii coloanelor
        header_row = 1
        order_id_col = None
        factura_col = None
        curier_col = None
        
        print("🔍 Caut coloanele în header...")
        for col_idx, cell in enumerate(ws[header_row], 1):
            # Preferă "Order ID" (eMag), dar acceptă și "AWB" pentru alte curiere
            if cell.value == "Order ID":
                order_id_col = col_idx
                print(f"   Order ID găsit în coloana {col_idx}")
            elif cell.value == "AWB" and order_id_col is None:
                order_id_col = col_idx
                print(f"   AWB găsit în coloana {col_idx}")
            elif cell.value == "Număr Factură":
                factura_col = col_idx
                print(f"   Număr Factură găsit în coloana {col_idx}")
            elif cell.value == "Curier":
                curier_col = col_idx
                print(f"   Curier găsit în coloana {col_idx}")
        
        if not all([order_id_col, factura_col, curier_col]):
            print("❌ EROARE: Nu pot găsi toate coloanele necesare (Order ID/AWB, Număr Factură, Curier)")
            return False
        
        print("✅ Toate coloanele găsite")
        
        # Parcurge toate rândurile și identifică cele eMag cu facturi goale
        comenzi_procesate = 0
        comenzi_canceled = 0
        comenzi_emag_total = 0
        comenzi_return_completate = 0

        # Transportă curierul anterior când celula e goală (în export curierul e scris doar pe primul rând din grup)
        curier_anterior = None
        print("🔄 Procesez rândurile...")
        from openpyxl.styles import PatternFill
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        for row_idx in range(2, ws.max_row + 1):  # Start de la rândul 2 (după header)
            curier_cell = ws.cell(row=row_idx, column=curier_col)
            order_id_cell = ws.cell(row=row_idx, column=order_id_col)
            factura_cell = ws.cell(row=row_idx, column=factura_col)

            curier_val = curier_cell.value if curier_cell.value not in [None, "", " "] else curier_anterior
            if curier_cell.value not in [None, "", " "]:
                curier_anterior = curier_cell.value

            # Verifică dacă este rând eMag (direct sau prin curierul anterior)
            if curier_val == "eMag":
                comenzi_emag_total += 1
                order_id = str(order_id_cell.value).strip() if order_id_cell.value else ""
                factura = str(factura_cell.value).strip() if factura_cell.value else ""

                # Dacă factura este 'ANULATA', normalizează la 'Canceled'
                if factura.upper() == 'ANULATA':
                    factura_cell.value = "Canceled"
                    factura = "Canceled"
                    comenzi_canceled += 1
                    factura_cell.fill = red_fill
                    print(f"✅ Order ID {order_id} - 'ANULATA' normalizat la 'Canceled' (rândul {row_idx})")
                    continue

                # Dacă nu are factură și Order ID este valid
                if not factura and order_id and order_id not in ['', 'None', 'nan']:
                    comenzi_procesate += 1

                    # Verifică statusul în easySales
                    status = status_dict.get(order_id, '')

                    if status == 'Canceled':
                        # Înlocuiește celula goală cu "Canceled" și colorează în roșu
                        factura_cell.value = "Canceled"
                        factura_cell.fill = red_fill
                        comenzi_canceled += 1
                        print(f"✅ Order ID {order_id} marcat ca 'Canceled' (rândul {row_idx})")
                    elif status == 'Return':
                        # Dacă este Return și există factură în easySales, completează factura
                        inv = invoice_dict.get(order_id)
                        if inv:
                            factura_cell.value = inv
                            comenzi_return_completate += 1
                            print(f"✅ Order ID {order_id} (Return) completat cu factura '{inv}' (rândul {row_idx})")
                        else:
                            print(f"ℹ️  Order ID {order_id} este 'Return' dar nu are factură în easySales (rândul {row_idx})")
                    else:
                        print(f"ℹ️  Order ID {order_id} status: '{status}' - rămâne gol (rândul {row_idx})")
        
        # Salvează modificările
        print("💾 Salvez modificările...")
        wb.save(cale_export)
        wb.close()
        
        print("\n" + "=" * 60)
        print("✅ POST-PROCESARE FINALIZATĂ CU SUCCES!")
        print("=" * 60)
        print(f"📊 Statistici:")
        print(f"   Total comenzi eMag găsite: {comenzi_emag_total}")
        print(f"   Comenzi fără factură procesate: {comenzi_procesate}")
        print(f"   Comenzi marcate ca 'Canceled': {comenzi_canceled}")
        print(f"   Comenzi 'Return' completate cu factură: {comenzi_return_completate}")
        print(f"📁 Fișierul {cale_export} a fost actualizat")
        
        return True
        
    except Exception as e:
        print(f"❌ EROARE: {e}")
        return False

def main():
    """Funcția principală - poate fi apelată din linia de comandă"""
    
    # Parsează argumentele din linia de comandă
    cale_export = sys.argv[1] if len(sys.argv) > 1 else None
    cale_easysales = sys.argv[2] if len(sys.argv) > 2 else None
    
    # Apelează funcția de procesare
    success = completeaza_comenzi_anulate_emag(cale_export, cale_easysales)
    
    if success:
        print("\n🎉 Procesare completată cu succes!")
        sys.exit(0)
    else:
        print("\n💥 Procesarea a eșuat!")
        sys.exit(1)

if __name__ == "__main__":
    main()